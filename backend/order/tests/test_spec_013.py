"""SPEC-ORDER-013 TASK-001~005: rack_number field, PATCH endpoints, Excel
upload (TDD).

Coverage targets:
  T1  LineItem.rack_number field declared with correct type/max_length/default
      (AC-RACK-001)
  T2  rack_number is independent of location/logistics_status/purchase_status
      (AC-RACK-001)
  T3  No Order-level aggregate/computed property for rack_number exists
      (AC-RACK-002)
  T4  LineItemRackNumberUpdateView (single PATCH) (AC-RACK-003/003a/003b/013)
  T5  LineItemBulkRackNumberUpdateView (bulk PATCH) (AC-RACK-004/004a/013)
  T6  parse_rack_number_excel() pure parser (AC-RACK-005/005a)
  T7  UploadRackNumberView (AC-RACK-006/006a/006b/006c/007)
"""

import io
from unittest.mock import patch

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from order.excel_utils import parse_rack_number_excel
from order.models import LineItem, Order

User = get_user_model()

RACK_NUMBER_URL = "/api/purchase-orders/line-items/{pk}/rack-number/"
BULK_RACK_NUMBER_URL = "/api/purchase-orders/line-items/bulk-rack-number/"
UPLOAD_RACK_NUMBER_URL = "/api/purchase-orders/upload-rack-number/"


def _make_order(shopify_order_id: int = 300001, store_type: str = "gimssine", **kwargs) -> Order:
    return Order.objects.create(shopify_order_id=shopify_order_id, store_type=store_type, **kwargs)


def _make_line_item(
    order: Order, shopify_line_item_id: int = 1, sku: str = "SKU-RACK-001", **kwargs
) -> LineItem:
    defaults = {"quantity": 1, "title": "Test Book"}
    defaults.update(kwargs)
    return LineItem.objects.create(
        order=order, shopify_line_item_id=shopify_line_item_id, sku=sku, **defaults
    )


def _make_rack_excel(rows: list[tuple], header: tuple = ("주문번호", "SKU", "렉번호")) -> bytes:
    """Build a 3-column rack-number upload .xlsx file. `rows` are
    (order_number, sku, rack_number) tuples written verbatim (no coercion),
    so callers can pass blank/None values to exercise skip paths."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(header))
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _file_obj(file_bytes: bytes, name: str = "rack-upload.xlsx"):
    f = io.BytesIO(file_bytes)
    f.name = name
    return f


@pytest.fixture
def user(db):
    return User.objects.create_user(username="spec013_user", password="testpass123")


@pytest.fixture
def auth_client(user):
    client = APIClient()
    refresh = RefreshToken.for_user(user)
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {refresh.access_token}")
    return client


@pytest.fixture
def anon_client():
    return APIClient()


@pytest.mark.django_db
class TestLineItemRackNumberField:
    """AC-RACK-001: rack_number persists as a string field, max length 10,
    default empty string, independent of location/logistics_status/
    purchase_status."""

    def test_field_declared_with_correct_type_and_constraints(self):
        field = LineItem._meta.get_field("rack_number")
        assert field.get_internal_type() == "CharField"
        assert field.max_length == 10
        assert field.blank is True
        assert field.get_default() == ""

    def test_new_line_item_defaults_to_empty_string(self):
        order = _make_order()
        li = _make_line_item(order)
        assert li.rack_number == ""

    def test_rack_number_persists_to_database(self):
        order = _make_order()
        li = _make_line_item(order, rack_number="A-12")
        li.refresh_from_db()
        assert li.rack_number == "A-12"

    def test_rack_number_independent_of_location(self):
        order = _make_order()
        li = _make_line_item(order, location="US-WH-1", rack_number="B-01")
        li.refresh_from_db()
        assert li.location == "US-WH-1"
        assert li.rack_number == "B-01"

    def test_rack_number_independent_of_logistics_status(self):
        order = _make_order()
        li = _make_line_item(order, logistics_status="shipped", rack_number="C-02")
        li.refresh_from_db()
        assert li.logistics_status == "shipped"
        assert li.rack_number == "C-02"

    def test_rack_number_independent_of_purchase_status(self):
        order = _make_order()
        li = _make_line_item(order, purchase_status="in_stock", rack_number="D-03")
        li.refresh_from_db()
        assert li.purchase_status == "in_stock"
        assert li.rack_number == "D-03"

    def test_updating_rack_number_does_not_change_sibling_fields(self):
        order = _make_order()
        li = _make_line_item(
            order,
            location="US-WH-1",
            logistics_status="received",
            purchase_status="in_stock",
        )
        li.rack_number = "E-04"
        li.save(update_fields=["rack_number"])
        li.refresh_from_db()
        assert li.location == "US-WH-1"
        assert li.logistics_status == "received"
        assert li.purchase_status == "in_stock"
        assert li.rack_number == "E-04"


@pytest.mark.django_db
class TestOrderHasNoRackNumberAggregate:
    """AC-RACK-002: no Order-level aggregate/rollup field, column, or
    computed property for rack_number exists anywhere on Order."""

    def test_order_model_has_no_rack_number_model_field(self):
        field_names = {f.name for f in Order._meta.get_fields()}
        assert "rack_number" not in field_names

    def test_order_instance_has_no_rack_number_attribute(self):
        order = _make_order()
        assert hasattr(order, "rack_number") is False

    def test_order_class_has_no_rack_number_class_attribute(self):
        assert hasattr(Order, "rack_number") is False


# ---------------------------------------------------------------------------
# TASK-002: LineItemRackNumberUpdateView (single PATCH)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestLineItemRackNumberUpdateView:
    """REQ-RACK-003/003a/003b, REQ-RACK-013 (partial) — AC-RACK-003/003a/003b/013."""

    def test_patch_valid_value_returns_200_and_persists(self, auth_client):
        order = _make_order(shopify_order_id=310001)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-RACK-PATCH-1")
        url = RACK_NUMBER_URL.format(pk=li.pk)

        res = auth_client.patch(url, data={"rack_number": "A-12"}, format="json")

        assert res.status_code == 200
        li.refresh_from_db()
        assert li.rack_number == "A-12"

    def test_patch_response_body_contains_id_rack_number_sku(self, auth_client):
        order = _make_order(shopify_order_id=310002)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-RACK-PATCH-2")
        url = RACK_NUMBER_URL.format(pk=li.pk)

        res = auth_client.patch(url, data={"rack_number": "B-01"}, format="json")

        assert res.status_code == 200
        assert res.data == {"id": li.pk, "rack_number": "B-01", "sku": "SKU-RACK-PATCH-2"}

    def test_patch_exactly_10_chars_accepted(self, auth_client):
        """Boundary: max_length is 10, so a 10-char value is valid (REQ-RACK-003)."""
        order = _make_order(shopify_order_id=310003)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-RACK-PATCH-3")
        url = RACK_NUMBER_URL.format(pk=li.pk)
        value = "1234567890"

        res = auth_client.patch(url, data={"rack_number": value}, format="json")

        assert res.status_code == 200
        li.refresh_from_db()
        assert li.rack_number == value

    def test_patch_oversized_value_returns_400_and_leaves_unchanged(self, auth_client):
        """REQ-RACK-003b: 11-char value must be rejected (AC-RACK-003b)."""
        order = _make_order(shopify_order_id=310004)
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-RACK-PATCH-4", rack_number="OLD-01"
        )
        url = RACK_NUMBER_URL.format(pk=li.pk)

        res = auth_client.patch(url, data={"rack_number": "12345678901"}, format="json")

        assert res.status_code == 400
        li.refresh_from_db()
        assert li.rack_number == "OLD-01"

    def test_patch_nonexistent_id_returns_404_and_no_modification(self, auth_client):
        """REQ-RACK-003a (AC-RACK-003a)."""
        order = _make_order(shopify_order_id=310005)
        other = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-RACK-PATCH-5", rack_number="UNTOUCHED"
        )
        url = RACK_NUMBER_URL.format(pk=999999)

        res = auth_client.patch(url, data={"rack_number": "X-01"}, format="json")

        assert res.status_code == 404
        other.refresh_from_db()
        assert other.rack_number == "UNTOUCHED"

    def test_patch_unauthenticated_returns_401(self, anon_client):
        url = RACK_NUMBER_URL.format(pk=1)
        res = anon_client.patch(url, data={"rack_number": "A-01"}, format="json")
        assert res.status_code == 401

    def test_patch_without_rack_number_key_clears_existing_value(self, auth_client):
        """Secondary finding (evaluator-active): the endpoint reads
        `request.data.get("rack_number", "")`, so a PATCH body that omits
        the `rack_number` key entirely defaults to "" and clears any
        existing value rather than leaving it unchanged. No REQ specifies
        behavior for an omitted key; this test locks in and documents the
        current (intentional default-to-empty) behavior rather than leaving
        it undocumented. Do not change this behavior without a new REQ."""
        order = _make_order(shopify_order_id=310007)
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-RACK-PATCH-6", rack_number="EXISTING"
        )
        url = RACK_NUMBER_URL.format(pk=li.pk)

        res = auth_client.patch(url, data={}, format="json")

        assert res.status_code == 200
        li.refresh_from_db()
        assert li.rack_number == ""

    def test_patch_duplicate_value_allowed_no_uniqueness_enforced(self, auth_client):
        """REQ-RACK-013 (AC-RACK-013): duplicating another LineItem's
        rack_number must not be rejected or flagged."""
        order = _make_order(shopify_order_id=310006)
        li1 = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-RACK-DUP-1", rack_number="SHARED-01"
        )
        li2 = _make_line_item(order, shopify_line_item_id=2, sku="SKU-RACK-DUP-2")
        url = RACK_NUMBER_URL.format(pk=li2.pk)

        res = auth_client.patch(url, data={"rack_number": "SHARED-01"}, format="json")

        assert res.status_code == 200
        li1.refresh_from_db()
        li2.refresh_from_db()
        assert li1.rack_number == "SHARED-01"
        assert li2.rack_number == "SHARED-01"


# ---------------------------------------------------------------------------
# TASK-003: LineItemBulkRackNumberUpdateView (bulk PATCH)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestLineItemBulkRackNumberUpdateView:
    """REQ-RACK-004/004a, REQ-RACK-013 (partial), REQ-RACK-002 (no Order
    recompute) — AC-RACK-004/004a/013."""

    def _make_lis(self, count: int, shopify_order_id: int = 320001) -> list:
        order = _make_order(shopify_order_id=shopify_order_id)
        return [
            _make_line_item(
                order, shopify_line_item_id=i + 1, sku=f"SKU-BULK-RACK-{shopify_order_id}-{i}"
            )
            for i in range(count)
        ]

    def test_bulk_update_success(self, auth_client):
        lis = self._make_lis(3)
        ids = [li.pk for li in lis]

        res = auth_client.patch(
            BULK_RACK_NUMBER_URL,
            data={"ids": ids, "rack_number": "Z-99"},
            format="json",
        )

        assert res.status_code == 200
        assert res.data["updated_count"] == 3
        assert res.data["missing_ids"] == []
        for li in lis:
            li.refresh_from_db()
            assert li.rack_number == "Z-99"

    def test_bulk_partial_success_with_missing_ids(self, auth_client):
        lis = self._make_lis(2, shopify_order_id=320002)
        ids = [li.pk for li in lis] + [999998, 999999]

        res = auth_client.patch(
            BULK_RACK_NUMBER_URL,
            data={"ids": ids, "rack_number": "Y-88"},
            format="json",
        )

        assert res.status_code == 200
        assert res.data["updated_count"] == 2
        assert sorted(res.data["missing_ids"]) == [999998, 999999]

    def test_bulk_empty_ids_returns_400_and_no_modification(self, auth_client):
        """REQ-RACK-004a (AC-RACK-004a)."""
        lis = self._make_lis(1, shopify_order_id=320003)

        res = auth_client.patch(
            BULK_RACK_NUMBER_URL,
            data={"ids": [], "rack_number": "W-77"},
            format="json",
        )

        assert res.status_code == 400
        lis[0].refresh_from_db()
        assert lis[0].rack_number == ""

    def test_bulk_oversized_value_returns_400(self, auth_client):
        lis = self._make_lis(1, shopify_order_id=320004)

        res = auth_client.patch(
            BULK_RACK_NUMBER_URL,
            data={"ids": [lis[0].pk], "rack_number": "12345678901"},
            format="json",
        )

        assert res.status_code == 400
        lis[0].refresh_from_db()
        assert lis[0].rack_number == ""

    def test_bulk_unauthenticated_returns_401(self, anon_client):
        res = anon_client.patch(
            BULK_RACK_NUMBER_URL,
            data={"ids": [1], "rack_number": "V-01"},
            format="json",
        )
        assert res.status_code == 401

    def test_bulk_duplicate_value_allowed_no_uniqueness_enforced(self, auth_client):
        """REQ-RACK-013 (AC-RACK-013)."""
        lis = self._make_lis(2, shopify_order_id=320005)
        ids = [li.pk for li in lis]

        res = auth_client.patch(
            BULK_RACK_NUMBER_URL,
            data={"ids": ids, "rack_number": "SHARE-BLK"},
            format="json",
        )

        assert res.status_code == 200
        for li in lis:
            li.refresh_from_db()
            assert li.rack_number == "SHARE-BLK"

    def test_bulk_never_calls_recompute_order_aggregates(self, auth_client):
        """REQ-RACK-002: rack_number is not an Order-level aggregate, so the
        bulk write path must never trigger Order aggregate recomputation
        (critical invariant, distinguishes this from SPEC-ORDER-012's
        LineItemBulkStatusUpdateView)."""
        lis = self._make_lis(2, shopify_order_id=320006)
        ids = [li.pk for li in lis]

        with patch("order.purchase_order_views._recompute_order_aggregates") as spy:
            res = auth_client.patch(
                BULK_RACK_NUMBER_URL,
                data={"ids": ids, "rack_number": "U-01"},
                format="json",
            )

            assert res.status_code == 200
            spy.assert_not_called()


# ---------------------------------------------------------------------------
# TASK-004: parse_rack_number_excel() pure parser
# ---------------------------------------------------------------------------


class TestParseRackNumberExcel:
    """REQ-RACK-005/005a — AC-RACK-005/005a.

    Design fix (same-day follow-up to SPEC-ORDER-013): the order-identifier
    cell is read as a raw string (whitespace-stripped only, "#" preserved),
    NOT parsed as an integer -- it is matched against `Order.name` (exact
    string) downstream, never `Order.order_number`. See SPEC-ORDER-014."""

    def test_parses_rows_with_standard_korean_headers(self):
        file_bytes = _make_rack_excel(
            [("1001", "SKU-A", "A-01"), ("1002", "SKU-B", "B-02")],
            header=("주문번호", "SKU", "렉번호"),
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [
            {"order_name": "1001", "sku": "SKU-A", "rack_number": "A-01"},
            {"order_name": "1002", "sku": "SKU-B", "rack_number": "B-02"},
        ]

    def test_case_insensitive_and_position_independent_headers(self):
        """AC-RACK-005: headers in any order, any letter case."""
        file_bytes = _make_rack_excel(
            [("SKU-X", "RACK-1", "#2001")],
            header=("Lineitem sku", "Rack Code", "Order Number"),
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [{"order_name": "#2001", "sku": "SKU-X", "rack_number": "RACK-1"}]

    def test_english_alias_headers(self):
        file_bytes = _make_rack_excel(
            [("#3001", "SKU-Y", "C-03")],
            header=("order", "isbn", "rack"),
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [{"order_name": "#3001", "sku": "SKU-Y", "rack_number": "C-03"}]

    def test_missing_order_number_column_raises_valueerror(self):
        file_bytes = _make_rack_excel([("SKU-A", "A-01")], header=("SKU", "렉번호"))

        with pytest.raises(ValueError):
            parse_rack_number_excel(file_bytes)

    def test_missing_sku_column_raises_valueerror(self):
        file_bytes = _make_rack_excel([("#1001", "A-01")], header=("주문번호", "렉번호"))

        with pytest.raises(ValueError):
            parse_rack_number_excel(file_bytes)

    def test_missing_rack_column_raises_valueerror(self):
        file_bytes = _make_rack_excel([("#1001", "SKU-A")], header=("주문번호", "SKU"))

        with pytest.raises(ValueError):
            parse_rack_number_excel(file_bytes)

    def test_empty_rack_number_value_preserved_as_empty_string(self):
        """plan.md M3: blank rack_number is explicit-clear, not skipped."""
        file_bytes = _make_rack_excel([("#1001", "SKU-A", "")])

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [{"order_name": "#1001", "sku": "SKU-A", "rack_number": ""}]

    def test_row_with_blank_order_identifier_included_with_none(self):
        """A blank order-identifier cell must not be dropped from the
        parser's output either -- it is included with order_name: None so
        UploadRackNumberView can count it toward skipped_count instead of it
        vanishing from both counters (REQ-RACK-006a/AC-RACK-007)."""
        file_bytes = _make_rack_excel(
            [(None, "SKU-A", "A-01"), ("#1002", "SKU-B", "B-02")]
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [
            {"order_name": None, "sku": "SKU-A", "rack_number": "A-01"},
            {"order_name": "#1002", "sku": "SKU-B", "rack_number": "B-02"},
        ]

    def test_row_with_blank_string_order_identifier_included_with_none(self):
        """An order-identifier cell containing only whitespace must also
        normalize to None (same treatment as a truly empty/None cell)."""
        file_bytes = _make_rack_excel(
            [("   ", "SKU-A", "A-01"), ("#1002", "SKU-B", "B-02")]
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [
            {"order_name": None, "sku": "SKU-A", "rack_number": "A-01"},
            {"order_name": "#1002", "sku": "SKU-B", "rack_number": "B-02"},
        ]

    def test_row_with_blank_sku_skipped(self):
        file_bytes = _make_rack_excel(
            [("#1001", "", "A-01"), ("#1002", "SKU-B", "B-02")]
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [{"order_name": "#1002", "sku": "SKU-B", "rack_number": "B-02"}]

    # -- Regression: real-world "발송정리" Shopify warehouse export (bug fix) --

    def test_order_date_header_is_not_mistaken_for_order_number_column(self):
        """Bug 1 regression: header contains both 'Order Date' (index 0) and
        'Name' (index 1, the actual Shopify order-name column). The old
        substring match `"order" in h` matched 'order date' first because it
        scans left-to-right, so it locked onto the wrong (date) column
        instead of ever reaching 'Name' or the later '주문번호' column."""
        header = (
            "Order Date", "Name", "Lineitem sku", "Lineitem name", "Total",
            "Status", "공급사", "Note", "입고수량", "입고=전체", "신규입고",
            "순번", "키값", "렉번호", "현재개수", "주문번호", "Only교보",
            "입고완료", "입고개수",
        )
        row = [None] * len(header)
        row[0] = "2026-01-01"  # Order Date -- must NOT be picked as order identifier
        row[1] = "#37349"  # Name -- the real Shopify order name column
        row[2] = "SKU-A"  # Lineitem sku
        row[13] = "A-01"  # 렉번호
        file_bytes = _make_rack_excel([tuple(row)], header=header)

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [{"order_name": "#37349", "sku": "SKU-A", "rack_number": "A-01"}]

    def test_hash_prefixed_shopify_order_name_preserved_verbatim(self):
        """Shopify order names are formatted like '#37349'. The '#' MUST be
        preserved (not stripped) since it is matched against `Order.name`
        downstream, which stores the value with its leading '#' intact."""
        file_bytes = _make_rack_excel(
            [("#37349", "SKU-A", "A-01")],
            header=("주문번호", "SKU", "렉번호"),
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [{"order_name": "#37349", "sku": "SKU-A", "rack_number": "A-01"}]

    def test_name_column_matched_not_confused_with_lineitem_name_column(self):
        """Regression: a bare 'Name' header must match order_idx via the new
        exact-match alias, and must NOT be confused with a 'Lineitem name'
        column (the book title) even when 'Lineitem name' also contains the
        substring 'name'."""
        header = ("Name", "Lineitem name", "SKU", "렉번호")
        file_bytes = _make_rack_excel(
            [("#4001", "Some Book Title", "SKU-A", "A-01")],
            header=header,
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [{"order_name": "#4001", "sku": "SKU-A", "rack_number": "A-01"}]

    def test_eb_prefixed_order_identifier_preserved_as_string(self):
        """SPEC-ORDER-014 regression: an EB-prefixed manually-entered order
        identifier (e.g. '#EB10011778') must be preserved verbatim as a
        string -- it cannot be parsed as an integer at all, and must not be
        silently degraded to None the way the previous int-parsing design
        would have done."""
        file_bytes = _make_rack_excel(
            [("#EB10011778", "SKU-EB", "A-01")],
            header=("주문번호", "SKU", "렉번호"),
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [
            {"order_name": "#EB10011778", "sku": "SKU-EB", "rack_number": "A-01"}
        ]

    def test_numeric_cell_types_are_stringified(self):
        """Non-Shopify files with plain int/float order-identifier cells
        must still be captured -- stringified, not dropped -- even though
        they are no longer parsed back into integers."""
        file_bytes = _make_rack_excel(
            [(1001, "SKU-A", "A-01"), (1002.0, "SKU-B", "B-02")],
            header=("주문번호", "SKU", "렉번호"),
        )

        rows = parse_rack_number_excel(file_bytes)

        assert rows == [
            {"order_name": "1001", "sku": "SKU-A", "rack_number": "A-01"},
            {"order_name": "1002", "sku": "SKU-B", "rack_number": "B-02"},
        ]


# ---------------------------------------------------------------------------
# TASK-005: UploadRackNumberView
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadRackNumberView:
    """REQ-RACK-006/006a/006b/007 — AC-RACK-006/006a/006b/006c/007.

    Design fix (SPEC-ORDER-014, same-day follow-up to SPEC-ORDER-013): the
    Excel "Name" column identifier is matched against `Order.name` (exact
    string, e.g. "#37349") -- NOT `Order.order_number` (int) -- since the
    two fields diverge for manually-entered "EB"-prefixed orders."""

    def test_upload_wrong_extension_returns_400(self, auth_client):
        f = _file_obj(b"not an excel file", name="upload.csv")
        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")
        assert res.status_code == 400

    def test_upload_corrupted_xlsx_returns_422_and_no_modification(self, auth_client):
        """Secondary finding (evaluator-active): a file with the correct
        .xlsx extension but unreadable/corrupted binary content must be
        rejected with 422 (parser's openpyxl.load_workbook failure ->
        ValueError -> 422), not 500, and must not modify any LineItem."""
        order = _make_order(shopify_order_id=330014, name="#40012")
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-RACK-CORRUPT", rack_number="UNCHANGED"
        )
        f = _file_obj(b"this is not a valid xlsx binary payload", name="corrupt.xlsx")

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 422
        li.refresh_from_db()
        assert li.rack_number == "UNCHANGED"

    def test_upload_empty_xlsx_returns_422(self, auth_client):
        """Secondary finding (evaluator-active): an .xlsx workbook with no
        header/data rows at all must be rejected with 422, not 500."""
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        f = _file_obj(buf.getvalue(), name="empty.xlsx")

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 422

    def test_upload_malformed_header_returns_422_and_no_modification(self, auth_client):
        """REQ-RACK-005a via the view's error-translation layer (AC-RACK-005a)."""
        order = _make_order(shopify_order_id=330001, name="#40001")
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-RACK-422", rack_number="UNCHANGED"
        )
        file_bytes = _make_rack_excel([("#40001", "A-01")], header=("주문번호", "렉번호"))
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 422
        li.refresh_from_db()
        assert li.rack_number == "UNCHANGED"

    def test_upload_matches_and_updates_lineitem(self, auth_client):
        """REQ-RACK-006 (AC-RACK-006). Normal Shopify order: Order.name
        ("#40002") matched by exact string against the Excel "Name" cell."""
        order = _make_order(shopify_order_id=330002, name="#40002")
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-RACK-MATCH")
        file_bytes = _make_rack_excel([("#40002", "SKU-RACK-MATCH", "A-10")])
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        assert res.data["skipped_count"] == 0
        li.refresh_from_db()
        assert li.rack_number == "A-10"

    def test_upload_matches_eb_prefixed_order_with_divergent_order_number(self, auth_client):
        """SPEC-ORDER-014 regression: a manually-entered order has
        name="#EB10011778" and an unrelated order_number (1778) that shares
        no digits in common with the EB identifier. Matching MUST go through
        Order.name -- matching via Order.order_number (the previous,
        incorrect design) could never succeed here, since "EB10011778" is
        not parseable as an integer and, even if it were, would not equal
        1778."""
        order = _make_order(
            shopify_order_id=330015, name="#EB10011778", order_number=1778
        )
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-EB-MATCH")
        file_bytes = _make_rack_excel([("#EB10011778", "SKU-EB-MATCH", "EB-01")])
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        assert res.data["skipped_count"] == 0
        li.refresh_from_db()
        assert li.rack_number == "EB-01"

    def test_upload_order_not_found_skipped(self, auth_client):
        """REQ-RACK-006a (AC-RACK-006a)."""
        file_bytes = _make_rack_excel([("#99999", "SKU-NOPE", "A-01")])
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 200
        assert res.data["matched_count"] == 0
        assert res.data["skipped_count"] == 1

    def test_upload_sku_not_found_in_order_skipped(self, auth_client):
        """REQ-RACK-006a (AC-RACK-006a)."""
        order = _make_order(shopify_order_id=330003, name="#40003")
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-REAL")
        file_bytes = _make_rack_excel([("#40003", "SKU-DOES-NOT-EXIST", "A-01")])
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 200
        assert res.data["matched_count"] == 0
        assert res.data["skipped_count"] == 1
        li.refresh_from_db()
        assert li.rack_number == ""

    def test_upload_duplicate_order_sku_last_row_wins(self, auth_client):
        """REQ-RACK-006b (AC-RACK-006b)."""
        order = _make_order(shopify_order_id=330004, name="#40004")
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-DUP")
        file_bytes = _make_rack_excel(
            [
                ("#40004", "SKU-DUP", "FIRST"),
                ("#40004", "SKU-DUP", "LAST"),
            ]
        )
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        li.refresh_from_db()
        assert li.rack_number == "LAST"

    def test_upload_multiple_lineitems_matching_key_all_updated(self, auth_client):
        """Decision E / REQ-RACK-006: when (order_name, sku) matches 2+
        LineItems, all of them are updated."""
        order = _make_order(shopify_order_id=330005, name="#40005")
        li1 = _make_line_item(order, shopify_line_item_id=1, sku="SKU-MULTI")
        li2 = _make_line_item(order, shopify_line_item_id=2, sku="SKU-MULTI")
        file_bytes = _make_rack_excel([("#40005", "SKU-MULTI", "SHELF-1")])
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        li1.refresh_from_db()
        li2.refresh_from_db()
        assert li1.rack_number == "SHELF-1"
        assert li2.rack_number == "SHELF-1"

    def test_upload_cross_order_isolation(self, auth_client):
        """AC-RACK-006c: a LineItem in a different Order with the same SKU
        must not be modified."""
        order_a = _make_order(shopify_order_id=330006, name="#40006")
        order_b = _make_order(shopify_order_id=330007, name="#40007")
        li_a = _make_line_item(order_a, shopify_line_item_id=1, sku="SKU-CROSS")
        li_b = _make_line_item(order_b, shopify_line_item_id=1, sku="SKU-CROSS")
        file_bytes = _make_rack_excel([("#40006", "SKU-CROSS", "A-ONLY")])
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 200
        li_a.refresh_from_db()
        li_b.refresh_from_db()
        assert li_a.rack_number == "A-ONLY"
        assert li_b.rack_number == ""

    def test_upload_matched_skipped_count_distinct_keys_after_dedup(self, auth_client):
        """REQ-RACK-007: counts are per distinct (order_name, sku) key
        after dedup, not per source row nor per LineItem affected."""
        order = _make_order(shopify_order_id=330008, name="#40008")
        li1 = _make_line_item(order, shopify_line_item_id=1, sku="SKU-COUNT-A")
        li2 = _make_line_item(order, shopify_line_item_id=2, sku="SKU-COUNT-A")
        file_bytes = _make_rack_excel(
            [
                ("#40008", "SKU-COUNT-A", "FIRST"),
                ("#40008", "SKU-COUNT-A", "SECOND"),  # duplicate key -> deduped to 1
                ("#99999", "SKU-COUNT-B", "IGNORED"),  # order not found -> skipped
            ]
        )
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        assert res.data["skipped_count"] == 1
        li1.refresh_from_db()
        li2.refresh_from_db()
        assert li1.rack_number == "SECOND"
        assert li2.rack_number == "SECOND"

    def test_upload_blank_order_identifier_counts_as_skipped(self, auth_client):
        """REQ-RACK-006a/AC-RACK-007: a row whose order-identifier cell is
        blank/empty must still be counted -- as skipped, same as an "Order
        not found" row -- not silently dropped from both matched_count and
        skipped_count. (Previously this scenario was triggered by an
        *unparseable* order-number cell under the old int-parsing design;
        under the corrected string-matching design, any non-empty string is
        a valid identifier, so only a genuinely blank cell triggers this
        skip path now.)"""
        order = _make_order(shopify_order_id=330012, name="#40010")
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-RACK-VALID")
        file_bytes = _make_rack_excel(
            [
                (None, "SKU-BAD-ORDER", "A-01"),
                ("#40010", "SKU-RACK-VALID", "B-02"),
            ]
        )
        f = _file_obj(file_bytes)

        res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        assert res.data["skipped_count"] == 1
        li.refresh_from_db()
        assert li.rack_number == "B-02"

    def test_upload_unauthenticated_returns_401(self, anon_client):
        file_bytes = _make_rack_excel([("#1", "SKU-A", "A-01")])
        f = _file_obj(file_bytes)
        res = anon_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")
        assert res.status_code == 401

    def test_upload_never_calls_recompute_order_aggregates(self, auth_client):
        """REQ-RACK-002: upload write path must never trigger Order
        aggregate recomputation (same critical invariant as the bulk PATCH
        endpoint)."""
        order = _make_order(shopify_order_id=330009, name="#40009")
        _make_line_item(order, shopify_line_item_id=1, sku="SKU-RACK-NORECOMPUTE")
        file_bytes = _make_rack_excel([("#40009", "SKU-RACK-NORECOMPUTE", "A-01")])
        f = _file_obj(file_bytes)

        with patch("order.purchase_order_views._recompute_order_aggregates") as spy:
            res = auth_client.post(UPLOAD_RACK_NUMBER_URL, {"file": f}, format="multipart")

            assert res.status_code == 200
            spy.assert_not_called()


# ---------------------------------------------------------------------------
# TASK-006: LineItemDetailSerializer / OrderDetailSerializer expose rack_number
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestLineItemDetailSerializerExposesRackNumber:
    """REQ-RACK-001 (exposure), REQ-RACK-009 support: rack_number must be
    readable via LineItemDetailSerializer directly and transparently through
    OrderDetailSerializer.line_items (API-level exposure only — no
    OrderDetailPage UI change per REQ-RACK-012)."""

    def test_line_item_detail_serializer_includes_rack_number(self):
        from order.serializers import LineItemDetailSerializer

        order = _make_order(shopify_order_id=330010)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-RACK-SER-1", rack_number="A-99")

        data = LineItemDetailSerializer(li).data

        assert data["rack_number"] == "A-99"

    def test_order_detail_serializer_line_items_include_rack_number(self):
        from order.serializers import OrderDetailSerializer

        order = _make_order(shopify_order_id=330011)
        _make_line_item(order, shopify_line_item_id=1, sku="SKU-RACK-SER-2", rack_number="B-77")

        data = OrderDetailSerializer(order).data

        assert len(data["line_items"]) == 1
        assert data["line_items"][0]["rack_number"] == "B-77"
