"""SPEC-ORDER-011: LineItem 물류 상태(입고/출고) 추적 및 Order 집계 상태 (TDD).

Coverage targets (backend only — T1~T10; frontend T11 is a separate cycle):
  T1   LineItem.logistics_status field + Order.status choices
  T4   _recompute_order_aggregates() aggregation + N+1 query-count regression
  T5   excel_utils (order_name, sku) parsers
  T6   UploadVendorShipmentView
  T7   UploadWarehouseReceiptView
  T8   LineItemLogisticsStatusUpdateView / BulkUpdateView
  T10  PurchaseOrder.status non-interaction (REQ-LOGI-014/014a/014b)

Shopify sync exclusion tests (T3) live in test_shopify_orders.py.
Backfill migration tests (T2/T10) live in test_backfill_order_status_migration.py.
"""

import io
from unittest.mock import patch

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import connection
from django.test.utils import CaptureQueriesContext
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from order.excel_utils import parse_vendor_shipment_excel, parse_warehouse_receipt_excel
from order.models import LineItem, Order, PurchaseOrder
from order.purchase_order_views import _recompute_order_aggregates

User = get_user_model()

UPLOAD_VENDOR_SHIPMENT_URL = "/api/purchase-orders/upload-vendor-shipment/"
UPLOAD_WAREHOUSE_RECEIPT_URL = "/api/purchase-orders/upload-warehouse-receipt/"
LOGISTICS_STATUS_URL = "/api/purchase-orders/line-items/{pk}/logistics-status/"
BULK_LOGISTICS_STATUS_URL = "/api/purchase-orders/line-items/bulk-logistics-status/"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def user(db):
    return User.objects.create_user(username="spec011_user", password="testpass123")


@pytest.fixture
def auth_client(user):
    client = APIClient()
    refresh = RefreshToken.for_user(user)
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {refresh.access_token}")
    return client


@pytest.fixture
def anon_client():
    return APIClient()


def _make_order(shopify_order_id: int = 80001, store_type: str = "gimssine", **kwargs) -> Order:
    return Order.objects.create(shopify_order_id=shopify_order_id, store_type=store_type, **kwargs)


def _make_line_item(order: Order, shopify_line_item_id: int = 1, sku: str = "SKU-001", **kwargs) -> LineItem:
    defaults = {"quantity": 1, "title": "Test Book"}
    defaults.update(kwargs)
    return LineItem.objects.create(
        order=order, shopify_line_item_id=shopify_line_item_id, sku=sku, **defaults
    )


def _make_sku_only_excel(skus: list, header=("SKU", "기타컬럼")) -> bytes:
    """Build a SKU-only .xlsx file (no order-name column) — used to test the
    missing-required-column validation path now that both upload endpoints
    require (order_name, SKU)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(header))
    for sku in skus:
        ws.append([sku, "ignored"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_name_sku_excel(rows: list, header=("주문번호", "SKU", "기타컬럼")) -> bytes:
    """Build an (order_name, SKU) .xlsx file. `rows` is a list of
    (name, sku) tuples; a `None` entry in either position produces a blank
    cell (used to test blank-row-skip behavior)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(header))
    for name, sku in rows:
        ws.append([name, sku, "ignored"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _file_obj(file_bytes: bytes, name: str = "upload.xlsx"):
    f = io.BytesIO(file_bytes)
    f.name = name
    return f


# ---------------------------------------------------------------------------
# T1: LineItem.logistics_status field + Order.status choices
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestLineItemLogisticsStatusField:
    def test_default_logistics_status_is_not_shipped(self):
        """REQ-LOGI-001/AC-LOGI-001: defaults to 'not_shipped' (시나리오0)."""
        order = _make_order(shopify_order_id=85001)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-DEF-1")
        assert li.logistics_status == "not_shipped"

    def test_all_five_choices_accepted(self):
        valid_choices = [
            "not_shipped", "shipment_confirmed", "received",
            "outbound_scheduled", "shipped",
        ]
        order = _make_order(shopify_order_id=85002)
        for i, choice in enumerate(valid_choices):
            li = LineItem.objects.create(
                order=order,
                shopify_line_item_id=100 + i,
                sku=f"SKU-LOGI-{i}",
                quantity=1,
                logistics_status=choice,
            )
            li.refresh_from_db()
            assert li.logistics_status == choice

    def test_invalid_choice_rejected_on_full_clean(self):
        order = _make_order(shopify_order_id=85003)
        li = LineItem(
            order=order,
            shopify_line_item_id=200,
            sku="SKU-LOGI-INVALID",
            quantity=1,
            logistics_status="invalid_status",
        )
        with pytest.raises(ValidationError):
            li.full_clean()

    def test_independent_of_purchase_status_and_fulfillment_status(self):
        """AC-LOGI-001: never inferred from/synced with purchase_status or
        fulfillment_status (시나리오0)."""
        order = _make_order(shopify_order_id=85004)
        li = LineItem.objects.create(
            order=order,
            shopify_line_item_id=1,
            sku="SKU-LOGI-INDEP",
            quantity=1,
            purchase_status="in_stock",
            fulfillment_status="fulfilled",
            logistics_status="received",
        )
        li.purchase_status = "on_hold"
        li.fulfillment_status = "unfulfilled"
        li.save(update_fields=["purchase_status", "fulfillment_status"])
        li.refresh_from_db()
        assert li.logistics_status == "received"


@pytest.mark.django_db
class TestOrderStatusChoices:
    def test_status_field_accepts_all_logistics_values_plus_partial(self):
        """Decision D: Order.status choices = 5 logistics codes + 'partial'."""
        valid_values = [
            "not_shipped", "shipment_confirmed", "received",
            "outbound_scheduled", "shipped", "partial",
        ]
        for i, value in enumerate(valid_values):
            order = Order.objects.create(
                shopify_order_id=86000 + i, store_type="gimssine", status=value
            )
            order.full_clean()
            assert order.status == value

    def test_status_field_still_nullable(self):
        """Column itself unchanged — still null=True, blank=True."""
        order = _make_order(shopify_order_id=86100)
        assert order.status is None

    def test_invalid_status_value_rejected_on_full_clean(self):
        order = Order(shopify_order_id=86101, store_type="gimssine", status="bogus")
        with pytest.raises(ValidationError):
            order.full_clean()


# ---------------------------------------------------------------------------
# T5: excel_utils (order_name, sku) parsers
# ---------------------------------------------------------------------------


class TestParseNameSkuExcel:
    def test_parse_vendor_shipment_excel_returns_name_sku_rows(self):
        file_bytes = _make_name_sku_excel([("#P1", "SKU-P1"), ("#P2", "SKU-P2")])
        rows = parse_vendor_shipment_excel(file_bytes)
        assert [(r["name"], r["sku"]) for r in rows] == [("#P1", "SKU-P1"), ("#P2", "SKU-P2")]

    def test_parse_warehouse_receipt_excel_returns_name_sku_rows(self):
        """REQ-LOGI-016: parse_warehouse_receipt_excel now delegates to
        _parse_name_sku_xlsx (no quantity column) — each row is one received
        unit, so a repeated (name, sku) pair simply appears as multiple
        identical rows."""
        file_bytes = _make_name_sku_excel([("#P3", "SKU-P3"), ("#P3", "SKU-P3")])
        rows = parse_warehouse_receipt_excel(file_bytes)
        assert rows == [
            {"name": "#P3", "sku": "SKU-P3"},
            {"name": "#P3", "sku": "SKU-P3"},
        ]

    def test_blank_sku_row_skipped(self):
        file_bytes = _make_name_sku_excel([("#P4", "SKU-P4"), ("#P4b", None), ("#P5", "SKU-P5")])
        rows = parse_vendor_shipment_excel(file_bytes)
        assert [r["sku"] for r in rows] == ["SKU-P4", "SKU-P5"]

    def test_blank_name_row_still_included(self):
        """A blank order-name cell on an otherwise-valid SKU row is NOT
        dropped by the parser — it flows through so the calling view can
        count it as skipped (no Order match), same convention as
        parse_outbound_excel/parse_rack_number_excel."""
        file_bytes = _make_name_sku_excel([(None, "SKU-P6")])
        rows = parse_vendor_shipment_excel(file_bytes)
        assert rows == [{"name": "", "sku": "SKU-P6"}]

    def test_isbn_header_also_matched_case_insensitive(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["주문번호", "Isbn", "기타"])
        ws.append(["#ISBN1", "9788900000000", "x"])
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_vendor_shipment_excel(buf.getvalue())
        assert rows == [{"name": "#ISBN1", "sku": "9788900000000"}]

    def test_missing_sku_or_isbn_header_raises_value_error(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["주문번호", "제목", "수량"])
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError):
            parse_vendor_shipment_excel(buf.getvalue())

    def test_missing_name_column_raises_value_error(self):
        """REQ-LOGI-003b/005b: the order-name column is now required, same
        dual-column validation as parse_outbound_excel."""
        file_bytes = _make_sku_only_excel(["SKU-NONAME"])
        with pytest.raises(ValueError):
            parse_vendor_shipment_excel(file_bytes)

    def test_empty_file_raises_value_error(self):
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError):
            parse_warehouse_receipt_excel(buf.getvalue())


# ---------------------------------------------------------------------------
# T4: _recompute_order_aggregates() helper
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestRecomputeOrderStatusHelper:
    def test_uniform_logistics_status_sets_shared_value(self):
        """AC-LOGI-008/009 — 시나리오3."""
        order = _make_order(shopify_order_id=87001)
        _make_line_item(order, shopify_line_item_id=1, sku="SKU-U1", logistics_status="received")
        _make_line_item(order, shopify_line_item_id=2, sku="SKU-U2", logistics_status="received")
        _recompute_order_aggregates([order.id])
        order.refresh_from_db()
        assert order.status == "received"

    def test_mixed_logistics_status_sets_partial(self):
        """AC-LOGI-008 — 시나리오4."""
        order = _make_order(shopify_order_id=87002)
        _make_line_item(order, shopify_line_item_id=1, sku="SKU-M1", logistics_status="not_shipped")
        _make_line_item(order, shopify_line_item_id=2, sku="SKU-M2", logistics_status="received")
        _recompute_order_aggregates([order.id])
        order.refresh_from_db()
        assert order.status == "partial"

    def test_no_trackable_lineitems_leaves_status_unset(self):
        """AC-LOGI-008: sku IS NULL on every child LineItem -> status stays None."""
        order = _make_order(shopify_order_id=87003, status="paid")
        LineItem.objects.create(
            order=order, shopify_line_item_id=1, sku=None, quantity=1, title="No SKU"
        )
        _recompute_order_aggregates([order.id])
        order.refresh_from_db()
        assert order.status is None

    def test_noop_on_empty_order_ids(self):
        """No-op — must not raise and must not issue any query."""
        with CaptureQueriesContext(connection) as ctx:
            _recompute_order_aggregates([])
        assert len(ctx.captured_queries) == 0

    def test_recompute_overwrites_stale_status(self):
        order = _make_order(shopify_order_id=87004, status="paid")
        _make_line_item(order, shopify_line_item_id=1, sku="SKU-S1", logistics_status="shipped")
        _recompute_order_aggregates([order.id])
        order.refresh_from_db()
        assert order.status == "shipped"


@pytest.mark.django_db
class TestRecomputeOrderStatusQueryCount:
    """AC-LOGI-010 — 시나리오5: query count for Order.status recomputation
    scales with distinct Order count, not LineItem count (SPEC-PURCHASE-
    ORDER-009 N+1-avoidance precedent)."""

    def _bulk_patch_and_capture(
        self, auth_client, order_count: int, items_per_order: int, id_seed: int
    ) -> int:
        ids = []
        for o in range(order_count):
            order = _make_order(shopify_order_id=id_seed + o)
            for i in range(items_per_order):
                li = _make_line_item(order, shopify_line_item_id=i + 1, sku=f"SKU-QC-{id_seed}-{o}-{i}")
                ids.append(li.pk)

        with CaptureQueriesContext(connection) as ctx:
            res = auth_client.patch(
                BULK_LOGISTICS_STATUS_URL,
                data={"ids": ids, "logistics_status": "received"},
                format="json",
            )
        assert res.status_code == 200
        return len(ctx.captured_queries)

    def test_query_count_does_not_scale_with_lineitem_count(self, auth_client):
        """5 Orders x 3 LineItems vs 5 Orders x 30 LineItems must issue the
        same (small, constant) number of queries."""
        small_count = self._bulk_patch_and_capture(
            auth_client, order_count=5, items_per_order=3, id_seed=88000
        )
        large_count = self._bulk_patch_and_capture(
            auth_client, order_count=5, items_per_order=30, id_seed=88100
        )
        assert small_count == large_count, (
            f"query count scaled with LineItem count: {small_count} vs {large_count}"
        )
        assert small_count < 15, f"expected a small constant query count, got {small_count}"


# ---------------------------------------------------------------------------
# T6: UploadVendorShipmentView
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadVendorShipmentView:
    def test_matches_eligible_lineitem_and_transitions(self, auth_client):
        """REQ-LOGI-003/AC-LOGI-003 — 시나리오1."""
        order = _make_order(shopify_order_id=89001, name="#89001")
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-VS-1", purchase_status="in_stock"
        )
        file_bytes = _make_name_sku_excel([("#89001", "SKU-VS-1")])

        res = auth_client.post(
            UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        li.refresh_from_db()
        assert li.logistics_status == "shipment_confirmed"

    def test_upload_response_counts(self, auth_client):
        """REQ-LOGI-004/AC-LOGI-004 — 시나리오1b: matched + skipped ==
        distinct (order_name, sku) pair count in the file."""
        order = _make_order(shopify_order_id=89002, name="#89002")
        for i in range(3):
            _make_line_item(
                order, shopify_line_item_id=i + 1, sku=f"SKU-VS-M{i}", purchase_status="in_stock"
            )
        file_bytes = _make_name_sku_excel(
            [
                ("#89002", "SKU-VS-M0"),
                ("#89002", "SKU-VS-M1"),
                ("#89002", "SKU-VS-M2"),
                ("#89002", "SKU-VS-UNMATCHED"),
            ]
        )

        res = auth_client.post(
            UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 3
        assert res.data["skipped_count"] == 1
        assert res.data["matched_count"] + res.data["skipped_count"] == 4

    def test_duplicate_sku_rows_deduplicated_before_matching(self, auth_client):
        """REQ-LOGI-003a/AC-LOGI-003a — 시나리오1c: duplicate (order_name, sku)
        rows collapse to a single distinct pair (set-based dedup happens
        before matching, so matched+skipped counts the pair once, not twice)."""
        order = _make_order(shopify_order_id=89003, name="#89003")
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-VS-DUP", purchase_status="in_stock"
        )
        file_bytes = _make_name_sku_excel([("#89003", "SKU-VS-DUP"), ("#89003", "SKU-VS-DUP")])

        res = auth_client.post(
            UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        assert res.data["skipped_count"] == 0
        li.refresh_from_db()
        assert li.logistics_status == "shipment_confirmed"

    def test_unordered_purchase_status_excluded(self, auth_client):
        """REQ-LOGI-003/AC-LOGI-003 — 시나리오2c: purchase_status='unordered'
        SKUs are not matched (still counted as skipped)."""
        order = _make_order(shopify_order_id=89004, name="#89004")
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-VS-UNORD", purchase_status="unordered"
        )
        file_bytes = _make_name_sku_excel([("#89004", "SKU-VS-UNORD")])

        res = auth_client.post(
            UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 0
        assert res.data["skipped_count"] == 1
        li.refresh_from_db()
        assert li.logistics_status == "not_shipped"

    def test_already_shipment_confirmed_not_rematched(self, auth_client):
        """REQ-LOGI-003: only logistics_status='not_shipped' is eligible —
        already-confirmed LineItems are skipped, not re-transitioned."""
        order = _make_order(shopify_order_id=89005, name="#89005")
        li = _make_line_item(
            order,
            shopify_line_item_id=1,
            sku="SKU-VS-ALREADY",
            purchase_status="in_stock",
            logistics_status="shipment_confirmed",
        )
        file_bytes = _make_name_sku_excel([("#89005", "SKU-VS-ALREADY")])

        res = auth_client.post(
            UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["skipped_count"] == 1
        li.refresh_from_db()
        assert li.logistics_status == "shipment_confirmed"

    def test_atomicity_no_partial_update_on_failure(self, auth_client):
        """REQ-LOGI-003b/AC-LOGI-003b — 시나리오1d: a mid-transaction failure
        leaves no LineItem partially updated."""
        order = _make_order(shopify_order_id=89006, name="#89006")
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-VS-ATOMIC", purchase_status="in_stock"
        )
        file_bytes = _make_name_sku_excel([("#89006", "SKU-VS-ATOMIC")])

        with patch(
            "order.purchase_order_views.LineItem.objects.bulk_update",
            side_effect=Exception("boom"),
        ):
            res = auth_client.post(
                UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
            )
        assert res.status_code == 500
        li.refresh_from_db()
        assert li.logistics_status == "not_shipped"

    def test_same_sku_different_orders_only_named_order_transitions(self, auth_client):
        """REQ-LOGI-003b safety-fix regression test: two LineItems share the
        same SKU across two different orders. An upload naming only ONE of
        those orders must transition only that order's LineItem — matching
        by SKU alone previously flipped both."""
        order_a = _make_order(shopify_order_id=89007, name="#89007A")
        order_b = _make_order(shopify_order_id=89008, name="#89007B")
        li_a = _make_line_item(
            order_a, shopify_line_item_id=1, sku="SKU-VS-SHARED", purchase_status="in_stock"
        )
        li_b = _make_line_item(
            order_b, shopify_line_item_id=1, sku="SKU-VS-SHARED", purchase_status="in_stock"
        )
        file_bytes = _make_name_sku_excel([("#89007A", "SKU-VS-SHARED")])

        res = auth_client.post(
            UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        li_a.refresh_from_db()
        li_b.refresh_from_db()
        assert li_a.logistics_status == "shipment_confirmed"
        assert li_b.logistics_status == "not_shipped"

    def test_missing_file_returns_400(self, auth_client):
        res = auth_client.post(UPLOAD_VENDOR_SHIPMENT_URL, data={}, format="multipart")
        assert res.status_code == 400

    def test_invalid_file_extension_returns_400(self, auth_client):
        f = io.BytesIO(b"not an excel file")
        f.name = "upload.csv"
        res = auth_client.post(UPLOAD_VENDOR_SHIPMENT_URL, data={"file": f}, format="multipart")
        assert res.status_code == 400

    def test_missing_sku_column_returns_422(self, auth_client):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["제목", "수량"])
        ws.append(["Book A", 1])
        buf = io.BytesIO()
        wb.save(buf)
        res = auth_client.post(
            UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(buf.getvalue())}, format="multipart"
        )
        assert res.status_code == 422

    def test_unauthenticated_returns_401(self, anon_client):
        file_bytes = _make_name_sku_excel([("#89099", "SKU-VS-ANON")])
        res = anon_client.post(
            UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 401


# ---------------------------------------------------------------------------
# T7: UploadWarehouseReceiptView
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadWarehouseReceiptView:
    """REQ-LOGI-016: row-per-unit model — the real warehouse-receiving file
    has NO quantity column. Each row represents exactly one received unit,
    so a repeated (order_name, sku) pair across N rows expresses a quantity
    of N (see `_make_name_sku_excel`). REQ-LOGI-015's accumulation/threshold
    semantics — received_quantity accumulates across uploads, received_at is
    stamped on every write, logistics_status only advances to "received"
    once received_quantity reaches LineItem.quantity — are unchanged."""

    def test_single_row_with_quantity_one_completes(self, auth_client):
        """REQ-LOGI-016: one row == one received unit. LineItem.quantity == 1
        -> a single row completes it in one upload."""
        order = _make_order(shopify_order_id=90001, name="#90001")
        li = _make_line_item(
            order,
            shopify_line_item_id=1,
            sku="SKU-WR-1",
            quantity=1,
            logistics_status="shipment_confirmed",
        )
        file_bytes = _make_name_sku_excel([("#90001", "SKU-WR-1")])

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        assert res.data["skipped_count"] == 0
        li.refresh_from_db()
        assert li.received_quantity == 1
        assert li.received_at is not None
        assert li.logistics_status == "received"

    def test_n_repeated_rows_equal_quantity_completes_in_one_upload(self, auth_client):
        """REQ-LOGI-016: N repeated rows for the same (order, sku) express N
        received units. N == LineItem.quantity -> completes in one upload —
        mirrors REQ-OUTBOUND-010's completion behavior under the row-count
        model."""
        order = _make_order(shopify_order_id=90002, name="#90002")
        li = _make_line_item(
            order,
            shopify_line_item_id=1,
            sku="SKU-WR-N",
            quantity=5,
            logistics_status="shipment_confirmed",
        )
        file_bytes = _make_name_sku_excel([("#90002", "SKU-WR-N")] * 5)

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        assert res.data["skipped_count"] == 0
        li.refresh_from_db()
        assert li.received_quantity == 5
        assert li.received_at is not None
        assert li.logistics_status == "received"

    def test_direct_path_from_not_shipped_to_received(self, auth_client):
        """REQ-LOGI-015a: skips shipment_confirmed when the vendor never
        sent a shipment confirmation (Decision C, unchanged)."""
        order = _make_order(shopify_order_id=90009, name="#90009")
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-WR-2", quantity=3
        )  # not_shipped default
        file_bytes = _make_name_sku_excel([("#90009", "SKU-WR-2")] * 3)

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        li.refresh_from_db()
        assert li.received_quantity == 3
        assert li.logistics_status == "received"

    def test_partial_receipt_accumulates_without_transitioning(self, auth_client):
        """REQ-LOGI-015a: fewer repeated rows than LineItem.quantity ->
        received_quantity accumulates, received_at stamped, logistics_status
        stays whatever it was — not yet 'received'."""
        order = _make_order(shopify_order_id=90010, name="#90010")
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-WR-PARTIAL", quantity=10
        )
        file_bytes = _make_name_sku_excel([("#90010", "SKU-WR-PARTIAL")] * 4)

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        assert res.data["skipped_count"] == 0
        li.refresh_from_db()
        assert li.received_quantity == 4
        assert li.received_at is not None
        assert li.logistics_status == "not_shipped"

    def test_two_uploads_complete_a_partial_receipt(self, auth_client):
        """REQ-LOGI-015a: first upload partial, second upload brings it to
        completion -> logistics_status flips only on the second."""
        order = _make_order(shopify_order_id=90011, name="#90011")
        li = _make_line_item(
            order,
            shopify_line_item_id=1,
            sku="SKU-WR-TWOSTEP",
            quantity=10,
            logistics_status="shipment_confirmed",
        )
        file_bytes_1 = _make_name_sku_excel([("#90011", "SKU-WR-TWOSTEP")] * 6)
        res1 = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes_1)}, format="multipart"
        )
        assert res1.status_code == 200
        li.refresh_from_db()
        assert li.received_quantity == 6
        assert li.logistics_status == "shipment_confirmed"

        file_bytes_2 = _make_name_sku_excel([("#90011", "SKU-WR-TWOSTEP")] * 4)
        res2 = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes_2)}, format="multipart"
        )
        assert res2.status_code == 200
        li.refresh_from_db()
        assert li.received_quantity == 10
        assert li.logistics_status == "received"

    def test_over_receipt_rejected_as_quantity_exceeded(self, auth_client):
        """REQ-LOGI-015a: repeated rows that would push received_quantity
        past quantity are rejected — no partial write, existing
        received_quantity unchanged."""
        order = _make_order(shopify_order_id=90012, name="#90012")
        li = _make_line_item(
            order,
            shopify_line_item_id=1,
            sku="SKU-WR-OVER",
            quantity=5,
            received_quantity=2,
        )
        file_bytes = _make_name_sku_excel([("#90012", "SKU-WR-OVER")] * 4)

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 0
        assert res.data["skipped_count"] == 1
        li.refresh_from_db()
        assert li.received_quantity == 2
        assert li.received_at is None
        assert li.logistics_status == "not_shipped"

    def test_blank_order_name_row_skipped(self, auth_client):
        """REQ-LOGI-016: a row with a blank order-name cell cannot form a
        grouping key and is rejected (invalid_row), no write."""
        order = _make_order(shopify_order_id=90024, name="#90024")
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-WR-BLANKNAME", quantity=1)
        file_bytes = _make_name_sku_excel([(None, "SKU-WR-BLANKNAME")])

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 0
        assert res.data["skipped_count"] == 1
        li.refresh_from_db()
        assert li.received_quantity == 0
        assert li.received_at is None

    def test_atomicity_no_partial_update_on_failure(self, auth_client):
        """All-or-nothing: a mid-transaction failure leaves no LineItem
        partially updated (mirrors SPEC-ORDER-015's atomicity guarantee)."""
        with patch(
            "order.purchase_order_views.LineItem.objects.bulk_update",
            side_effect=Exception("boom"),
        ):
            order2 = _make_order(shopify_order_id=90004, name="#90004")
            li2 = _make_line_item(
                order2, shopify_line_item_id=1, sku="SKU-WR-ATOMIC", quantity=5
            )
            file_bytes2 = _make_name_sku_excel([("#90004", "SKU-WR-ATOMIC")] * 5)
            res2 = auth_client.post(
                UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes2)}, format="multipart"
            )
        assert res2.status_code == 500
        li2.refresh_from_db()
        assert li2.logistics_status == "not_shipped"
        assert li2.received_quantity == 0
        assert li2.received_at is None

    def test_warehouse_stock_untouched(self, auth_client):
        """REQ-LOGI-006/AC-LOGI-006 — 시나리오2d: WarehouseStock.quantity is
        never modified by a logistics_status transition to 'received'."""
        from order.models import WarehouseStock

        WarehouseStock.objects.create(isbn="SKU-WR-STOCK", location="korea", quantity=10)
        order = _make_order(shopify_order_id=90005, name="#90005")
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-WR-STOCK", quantity=2)
        file_bytes = _make_name_sku_excel([("#90005", "SKU-WR-STOCK")] * 2)

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        li.refresh_from_db()
        assert li.logistics_status == "received"
        stock = WarehouseStock.objects.get(isbn="SKU-WR-STOCK", location="korea")
        assert stock.quantity == 10

    def test_already_received_or_shipped_not_rematched(self, auth_client):
        """Only not_shipped/shipment_confirmed are eligible source states —
        received/outbound_scheduled/shipped are skipped."""
        order = _make_order(shopify_order_id=90006, name="#90006")
        li = _make_line_item(
            order,
            shopify_line_item_id=1,
            sku="SKU-WR-SHIPPED",
            quantity=1,
            logistics_status="shipped",
        )
        file_bytes = _make_name_sku_excel([("#90006", "SKU-WR-SHIPPED")])

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["skipped_count"] == 1
        li.refresh_from_db()
        assert li.logistics_status == "shipped"
        assert li.received_quantity == 0

    def test_same_sku_different_orders_only_named_order_transitions(self, auth_client):
        """REQ-LOGI-005b safety-fix regression test: two LineItems share the
        same SKU across two different orders. An upload naming only ONE of
        those orders must transition only that order's LineItem — matching
        by SKU alone previously flipped both. Must not regress under the
        row-count model."""
        order_a = _make_order(shopify_order_id=90007, name="#90007A")
        order_b = _make_order(shopify_order_id=90008, name="#90007B")
        li_a = _make_line_item(order_a, shopify_line_item_id=1, sku="SKU-WR-SHARED", quantity=1)
        li_b = _make_line_item(order_b, shopify_line_item_id=1, sku="SKU-WR-SHARED", quantity=1)
        file_bytes = _make_name_sku_excel([("#90007A", "SKU-WR-SHARED")])

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 1
        li_a.refresh_from_db()
        li_b.refresh_from_db()
        assert li_a.logistics_status == "received"
        assert li_a.received_quantity == 1
        assert li_b.logistics_status == "not_shipped"
        assert li_b.received_quantity == 0

    def test_order_not_found_skipped_no_write(self, auth_client):
        """order_not_found -> skipped, no write."""
        file_bytes = _make_name_sku_excel([("#NOPE-90020", "SKU-WR-NOORDER")])

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 0
        assert res.data["skipped_count"] == 1

    def test_line_item_not_found_skipped_no_write(self, auth_client):
        """line_item_not_found -> skipped, no write."""
        _make_order(shopify_order_id=90021, name="#90021")
        file_bytes = _make_name_sku_excel([("#90021", "SKU-WR-MISSING")])

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 0
        assert res.data["skipped_count"] == 1

    def test_multiple_line_items_skipped_no_write(self, auth_client):
        """multiple_line_items (2+ candidates, ambiguous) -> skipped, no
        write."""
        order = _make_order(shopify_order_id=90022, name="#90022")
        li1 = _make_line_item(order, shopify_line_item_id=1, sku="SKU-WR-DUPE", quantity=5)
        li2 = _make_line_item(order, shopify_line_item_id=2, sku="SKU-WR-DUPE", quantity=5)
        file_bytes = _make_name_sku_excel([("#90022", "SKU-WR-DUPE")] * 3)

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert res.data["matched_count"] == 0
        assert res.data["skipped_count"] == 1
        li1.refresh_from_db()
        li2.refresh_from_db()
        assert li1.received_quantity == 0
        assert li2.received_quantity == 0

    def test_response_exposes_per_row_detail_lists(self, auth_client):
        """REQ-LOGI-017: the response carries the three per-row detail lists
        (matched / unmatched / quantity_exceeded) alongside the pre-existing
        counts, so the UI can show WHICH rows were skipped and why — same
        payload shape OutboundProcessView returns. `matched_count` /
        `skipped_count` keep their previous meaning."""
        order = _make_order(shopify_order_id=90023, name="#90023")
        _make_line_item(order, shopify_line_item_id=1, sku="SKU-WR-SHAPE", quantity=1)
        # One row that matches, one that names a nonexistent order, and one
        # that over-receives an already-full LineItem — so all three lists
        # come back non-empty in a single call.
        full_li = _make_line_item(
            order, shopify_line_item_id=2, sku="SKU-WR-FULL", quantity=1, received_quantity=1
        )
        file_bytes = _make_name_sku_excel([
            ("#90023", "SKU-WR-SHAPE"),
            ("#NOPE-90023", "SKU-WR-GHOST"),
            ("#90023", "SKU-WR-FULL"),
        ])

        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200
        assert set(res.data.keys()) == {
            "matched_count",
            "skipped_count",
            "matched",
            "unmatched",
            "unmatched_count",
            "quantity_exceeded",
            "quantity_exceeded_count",
        }

        # skipped_count stays the sum of the two failure categories.
        assert res.data["matched_count"] == 1
        assert res.data["unmatched_count"] == 1
        assert res.data["quantity_exceeded_count"] == 1
        assert res.data["skipped_count"] == 2

        matched_row = res.data["matched"][0]
        assert matched_row["name"] == "#90023"
        assert matched_row["sku"] == "SKU-WR-SHAPE"
        assert matched_row["received_count"] == 1
        assert matched_row["received_quantity"] == 1
        assert matched_row["quantity"] == 1
        assert matched_row["logistics_status"] == "received"

        unmatched_row = res.data["unmatched"][0]
        assert unmatched_row["sku"] == "SKU-WR-GHOST"
        assert unmatched_row["reason"] == "order_not_found"

        exceeded_row = res.data["quantity_exceeded"][0]
        assert exceeded_row["sku"] == "SKU-WR-FULL"
        assert exceeded_row["line_item_id"] == full_li.id
        assert exceeded_row["reason"] == "quantity_exceeded"
        assert exceeded_row["received_quantity"] == 1
        assert exceeded_row["quantity"] == 1

    def test_unauthenticated_returns_401(self, anon_client):
        file_bytes = _make_name_sku_excel([("#90099", "SKU-WR-ANON")])
        res = anon_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 401

    def test_missing_file_returns_400(self, auth_client):
        res = auth_client.post(UPLOAD_WAREHOUSE_RECEIPT_URL, data={}, format="multipart")
        assert res.status_code == 400

    def test_invalid_file_extension_returns_400(self, auth_client):
        f = io.BytesIO(b"not an excel file")
        f.name = "upload.csv"
        res = auth_client.post(UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": f}, format="multipart")
        assert res.status_code == 400

    def test_missing_sku_column_returns_422(self, auth_client):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["제목", "메모"])
        ws.append(["Book A", "ignored"])
        buf = io.BytesIO()
        wb.save(buf)
        res = auth_client.post(
            UPLOAD_WAREHOUSE_RECEIPT_URL, data={"file": _file_obj(buf.getvalue())}, format="multipart"
        )
        assert res.status_code == 422


# ---------------------------------------------------------------------------
# T8: LineItemLogisticsStatusUpdateView (single PATCH)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestLineItemLogisticsStatusUpdateView:
    def test_patch_valid_value_returns_200(self, auth_client):
        """REQ-LOGI-007/AC-LOGI-007a — 시나리오2e."""
        order = _make_order(shopify_order_id=91001)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-PATCH-1")
        url = LOGISTICS_STATUS_URL.format(pk=li.pk)

        res = auth_client.patch(url, data={"logistics_status": "shipment_confirmed"}, format="json")
        assert res.status_code == 200
        li.refresh_from_db()
        assert li.logistics_status == "shipment_confirmed"
        assert res.data["logistics_status"] == "shipment_confirmed"
        assert res.data["id"] == li.pk

    def test_patch_invalid_value_returns_400_and_leaves_unchanged(self, auth_client):
        """REQ-LOGI-007a/AC-LOGI-007b — 시나리오2e2."""
        order = _make_order(shopify_order_id=91002)
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-PATCH-2", logistics_status="shipment_confirmed"
        )
        url = LOGISTICS_STATUS_URL.format(pk=li.pk)

        res = auth_client.patch(url, data={"logistics_status": "bogus_value"}, format="json")
        assert res.status_code == 400
        assert "bogus_value" in res.data["error"]
        li.refresh_from_db()
        assert li.logistics_status == "shipment_confirmed"

    def test_patch_nonexistent_id_returns_404(self, auth_client):
        url = LOGISTICS_STATUS_URL.format(pk=999999)
        res = auth_client.patch(url, data={"logistics_status": "received"}, format="json")
        assert res.status_code == 404

    def test_patch_unauthenticated_returns_401(self, anon_client):
        url = LOGISTICS_STATUS_URL.format(pk=1)
        res = anon_client.patch(url, data={"logistics_status": "received"}, format="json")
        assert res.status_code == 401

    def test_patch_all_five_choices(self, auth_client):
        valid_choices = [
            "not_shipped", "shipment_confirmed", "received",
            "outbound_scheduled", "shipped",
        ]
        order = _make_order(shopify_order_id=91003)
        for i, choice in enumerate(valid_choices):
            li = _make_line_item(order, shopify_line_item_id=100 + i, sku=f"SKU-PATCH-ALL-{i}")
            url = LOGISTICS_STATUS_URL.format(pk=li.pk)
            res = auth_client.patch(url, data={"logistics_status": choice}, format="json")
            assert res.status_code == 200, f"Failed for choice: {choice}"
            li.refresh_from_db()
            assert li.logistics_status == choice

    def test_patch_triggers_order_status_recompute(self, auth_client):
        """AC-LOGI-009 — 시나리오3: Order.status is recomputed before the
        response is returned."""
        order = _make_order(shopify_order_id=91004)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-PATCH-RECOMPUTE")
        url = LOGISTICS_STATUS_URL.format(pk=li.pk)

        res = auth_client.patch(url, data={"logistics_status": "received"}, format="json")
        assert res.status_code == 200
        order.refresh_from_db()
        assert order.status == "received"


# ---------------------------------------------------------------------------
# T8: LineItemLogisticsStatusBulkUpdateView (bulk PATCH)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestLineItemLogisticsStatusBulkUpdateView:
    def _make_lis(self, count: int, shopify_order_id: int = 92001) -> list:
        order = _make_order(shopify_order_id=shopify_order_id)
        return [
            _make_line_item(order, shopify_line_item_id=i + 1, sku=f"SKU-BULK-LOGI-{shopify_order_id}-{i}")
            for i in range(count)
        ]

    def test_bulk_update_success(self, auth_client):
        lis = self._make_lis(3)
        ids = [li.pk for li in lis]
        res = auth_client.patch(
            BULK_LOGISTICS_STATUS_URL,
            data={"ids": ids, "logistics_status": "shipment_confirmed"},
            format="json",
        )
        assert res.status_code == 200
        assert res.data["updated_count"] == 3
        assert res.data["missing_ids"] == []
        for li in lis:
            li.refresh_from_db()
            assert li.logistics_status == "shipment_confirmed"

    def test_bulk_partial_success_with_missing_ids(self, auth_client):
        lis = self._make_lis(2, shopify_order_id=92002)
        ids = [li.pk for li in lis] + [999998, 999999]
        res = auth_client.patch(
            BULK_LOGISTICS_STATUS_URL,
            data={"ids": ids, "logistics_status": "received"},
            format="json",
        )
        assert res.status_code == 200
        assert res.data["updated_count"] == 2
        assert 999998 in res.data["missing_ids"]
        assert 999999 in res.data["missing_ids"]

    def test_bulk_empty_ids_returns_400(self, auth_client):
        res = auth_client.patch(
            BULK_LOGISTICS_STATUS_URL,
            data={"ids": [], "logistics_status": "received"},
            format="json",
        )
        assert res.status_code == 400

    def test_bulk_invalid_value_returns_400_and_leaves_unchanged(self, auth_client):
        lis = self._make_lis(1, shopify_order_id=92003)
        res = auth_client.patch(
            BULK_LOGISTICS_STATUS_URL,
            data={"ids": [lis[0].pk], "logistics_status": "invalid_choice"},
            format="json",
        )
        assert res.status_code == 400
        assert "invalid_choice" in res.data["error"]
        lis[0].refresh_from_db()
        assert lis[0].logistics_status == "not_shipped"

    def test_bulk_unauthenticated_returns_401(self, anon_client):
        res = anon_client.patch(
            BULK_LOGISTICS_STATUS_URL,
            data={"ids": [1], "logistics_status": "received"},
            format="json",
        )
        assert res.status_code == 401

    def test_bulk_recomputes_all_affected_orders(self, auth_client):
        """AC-LOGI-010 — 시나리오5: N distinct Orders' status all recomputed
        by one bulk change."""
        order_a = _make_order(shopify_order_id=92010)
        order_b = _make_order(shopify_order_id=92011)
        li_a1 = _make_line_item(order_a, shopify_line_item_id=1, sku="SKU-BULK-A1")
        li_a2 = _make_line_item(order_a, shopify_line_item_id=2, sku="SKU-BULK-A2")
        li_b1 = _make_line_item(order_b, shopify_line_item_id=1, sku="SKU-BULK-B1")

        res = auth_client.patch(
            BULK_LOGISTICS_STATUS_URL,
            data={"ids": [li_a1.pk, li_b1.pk], "logistics_status": "received"},
            format="json",
        )
        assert res.status_code == 200

        order_a.refresh_from_db()
        order_b.refresh_from_db()
        li_a2.refresh_from_db()
        # order_a: li_a1=received, li_a2=not_shipped (untouched) -> mixed -> partial
        assert li_a2.logistics_status == "not_shipped"
        assert order_a.status == "partial"
        # order_b: only trackable LineItem is now received -> uniform
        assert order_b.status == "received"


# ---------------------------------------------------------------------------
# T10: PurchaseOrder.status non-interaction (REQ-LOGI-014/014a/014b)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestPurchaseOrderStatusIndependence:
    """시나리오6b/6c."""

    def test_manual_single_patch_does_not_touch_purchase_order_status(self, auth_client):
        order = _make_order(shopify_order_id=93001)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-PO-1")
        po = PurchaseOrder.objects.create(
            sku="SKU-PO-1", title="Book", distributor="booxen", quantity=1, status="pending"
        )
        po.line_items.add(li)

        url = LOGISTICS_STATUS_URL.format(pk=li.pk)
        res = auth_client.patch(url, data={"logistics_status": "received"}, format="json")
        assert res.status_code == 200

        po.refresh_from_db()
        assert po.status == "pending"

    def test_bulk_patch_does_not_touch_purchase_order_status(self, auth_client):
        order = _make_order(shopify_order_id=93002)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-PO-2")
        po = PurchaseOrder.objects.create(
            sku="SKU-PO-2", title="Book", distributor="kyobo", quantity=1, status="confirmed"
        )
        po.line_items.add(li)

        res = auth_client.patch(
            BULK_LOGISTICS_STATUS_URL,
            data={"ids": [li.pk], "logistics_status": "shipped"},
            format="json",
        )
        assert res.status_code == 200

        po.refresh_from_db()
        assert po.status == "confirmed"

    def test_upload_transition_does_not_touch_purchase_order_status(self, auth_client):
        order = _make_order(shopify_order_id=93003, name="#93003")
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-PO-3", purchase_status="in_stock"
        )
        po = PurchaseOrder.objects.create(
            sku="SKU-PO-3", title="Book", distributor="booxen", quantity=1, status="pending"
        )
        po.line_items.add(li)

        file_bytes = _make_name_sku_excel([("#93003", "SKU-PO-3")])
        res = auth_client.post(
            UPLOAD_VENDOR_SHIPMENT_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert res.status_code == 200

        po.refresh_from_db()
        assert po.status == "pending"

    def test_logistics_status_never_derived_from_purchase_order_status(self):
        """AC-LOGI-014b: changing PurchaseOrder.status has no effect on
        logistics_status — verified operationally (no write path in this
        SPEC reads PurchaseOrder.status as an input)."""
        order = _make_order(shopify_order_id=93004)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-PO-4")
        po = PurchaseOrder.objects.create(
            sku="SKU-PO-4", title="Book", distributor="booxen", quantity=1, status="pending"
        )
        po.line_items.add(li)

        before = li.logistics_status
        po.status = "confirmed"
        po.save(update_fields=["status"])

        li.refresh_from_db()
        assert li.logistics_status == before
