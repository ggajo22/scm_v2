"""
TDD tests for SPEC-ORDER-019 — Daily Review 업로드 배포처 행 메모 유실.

Coverage targets:
  T1  AC-MEMO-001  REQ-MEMO-001/002/003/004/005/006  a distributor row's memo
                   becomes exactly one 발주 note, byte-identical content
  T2  AC-MEMO-002  REQ-MEMO-007  whitespace-only memo creates no note, but the
                   row is still confirmed
  T3  AC-MEMO-003  REQ-MEMO-008  a file with no memo column does not regress;
                   a legacy file WITH a memo column does create the note
  T4  AC-MEMO-004  REQ-MEMO-009/010  memo-bearing PO rows at 300/500 rows stay
                   under the same absolute query ceiling as the existing test
  T5  AC-MEMO-005  REQ-MEMO-011/012  re-uploading the same bytes keeps one note
                   (idempotency inherited from _reorder_candidate_filter)
  T6  AC-MEMO-006  REQ-MEMO-013  CS / warehouse / distributor branches each
                   produce their own note with their own assignee
  T7  AC-MEMO-007  REQ-MEMO-001/014/015  one row resolving to N LineItems
                   produces N notes, and the PO aggregation is untouched
  T8  AC-MEMO-008  REQ-MEMO-014/015/016  memo presence changes nothing except
                   the notes themselves; no model/migration change
  T9  AC-MEMO-009  REQ-MEMO-017/018  the note shows up in the unresolved list
                   and never leaks into the 타출판사 Excel export

T5~T9 partly characterize behaviour that must already hold: the CS/warehouse
assertions in T6, the PurchaseOrder assertions in T7 and T8 (a)~(c), and the
whole read path in T9 must not require any production change beyond the
distributor-branch note append.

Fixtures / helpers follow test_daily_review_upload.py (:28, :42-52, :55-74,
:77-, :743-822, :2027-2032, :2117-2125, :2128-2129/:2138, :2349-2353) and the
snapshot convention of test_spec_018.py:197-216.
"""

import io

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.db import connection
from django.test.utils import CaptureQueriesContext
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from order.models import LineItem, LineItemNote, Order, PurchaseOrder, WarehouseStock
from order.tests.test_daily_review_upload import (
    _ACTIONABLE_CS_SELECTS,
    _ACTIONABLE_WAREHOUSE_LOCATIONS,
    _ACTIONABLE_WAREHOUSE_STATUSES,
    _FILLER_SELECTS,
    _N_FILLER,
    _PO_BRANCH_SELECTS,
    _make_daily_review_excel,
    _make_new_template_excel,
)

User = get_user_model()

UPLOAD_DAILY_URL = "/api/purchase-orders/upload-daily-review/"
NOTES_UNRESOLVED_URL = "/api/orders/line-item-notes/"
NOTES_EXPORT_URL = "/api/orders/line-item-notes/export/"

# Same absolute ceiling the pre-existing REQ-PO9-010 test pins
# (test_daily_review_upload.py:2160). AC-MEMO-004 asserts against this fixed
# value, not against a second measurement — a with/without comparison cancels
# out any constant cost added to both sides.
UPLOAD_QUERY_CEILING = 35


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------


@pytest.fixture
def user(db):
    return User.objects.create_user(username="spec019_user", password="testpass123")


@pytest.fixture
def auth_client(user):
    client = APIClient()
    refresh = RefreshToken.for_user(user)
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {refresh.access_token}")
    return client


def _make_order(shopify_order_id: int, name: str) -> Order:
    return Order.objects.create(
        shopify_order_id=shopify_order_id, store_type="gimssine", name=name
    )


def _make_line_item(
    order: Order,
    sku: str,
    *,
    title: str = "테스트 도서",
    quantity: int = 2,
    shopify_line_item_id: int = 1,
) -> LineItem:
    return LineItem.objects.create(
        order=order,
        shopify_line_item_id=shopify_line_item_id,
        sku=sku,
        title=title,
        quantity=quantity,
    )


def _post_file(auth_client, file_bytes: bytes):
    file_obj = io.BytesIO(file_bytes)
    file_obj.name = "daily_review.xlsx"
    return auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")


# ---------------------------------------------------------------------------
# T1 — AC-MEMO-001
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestDistributorRowMemoBecomesPurchasingNote:
    """T1 (AC-MEMO-001) — REQ-MEMO-001/002/003/004/005/006."""

    def test_distributor_row_memo_creates_exactly_one_purchasing_note(self, auth_client):
        order = _make_order(shopify_order_id=190001, name="#M001")
        li = _make_line_item(
            order, sku="S19-9788936479497", quantity=1, shopify_line_item_id=1
        )

        file_bytes = _make_new_template_excel([
            {
                "order_name": "#M001",
                "sku": "S19-9788936479497",
                "selected": "BOOXEN",
                "status": "품절이지만 북센 시도",
                "bs_price": 9000,
            }
        ])

        res = _post_file(auth_client, file_bytes)
        assert res.status_code == 201

        notes = list(LineItemNote.objects.filter(line_item=li))
        # (a)
        assert len(notes) == 1, f"expected exactly 1 note for the distributor row, got {len(notes)}"
        note = notes[0]
        # (b) byte-for-byte identical — no distributor prefix/suffix, no truncation
        assert note.content == "품절이지만 북센 시도"
        # (c)
        assert note.assignee == "발주"
        # (d)
        assert note.note_type == ""
        # (e)
        assert note.author is None
        # (f) recorded as already-resolved — the memo is a record of why the row
        # was confirmed, not a task, so it must not enter the unresolved queue
        assert note.is_resolved is True


# ---------------------------------------------------------------------------
# T2 — AC-MEMO-002
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestBlankMemoCreatesNoNote:
    """T2 (AC-MEMO-002) — REQ-MEMO-007."""

    def test_whitespace_only_memo_creates_no_note_but_row_is_still_confirmed(
        self, auth_client
    ):
        order = _make_order(shopify_order_id=190002, name="#M002")
        li_blank = _make_line_item(order, sku="S19-BLANK", shopify_line_item_id=1)
        li_real = _make_line_item(order, sku="S19-REAL", shopify_line_item_id=2)

        file_bytes = _make_new_template_excel([
            {
                "order_name": "#M002",
                "sku": "S19-BLANK",
                "selected": "BOOXEN",
                "status": "   ",
                "bs_price": 9000,
            },
            {
                "order_name": "#M002",
                "sku": "S19-REAL",
                "selected": "BOOXEN",
                "status": "재고 확인 요청",
                "bs_price": 9000,
            },
        ])

        res = _post_file(auth_client, file_bytes)
        assert res.status_code == 201

        # (a)
        assert LineItemNote.objects.count() == 1
        # (b)
        note = LineItemNote.objects.get()
        assert note.line_item_id == li_real.pk
        assert note.content == "재고 확인 요청"
        # (c)
        assert LineItemNote.objects.filter(line_item=li_blank).exists() is False
        # (d) the blank-memo row is confirmed all the same
        li_blank.refresh_from_db()
        assert li_blank.confirmed_distributor == "booxen"
        assert PurchaseOrder.objects.filter(sku="S19-BLANK").exists() is True


# ---------------------------------------------------------------------------
# T3 — AC-MEMO-003
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestMissingMemoColumnDoesNotRegress:
    """T3 (AC-MEMO-003) — REQ-MEMO-008."""

    def _make_excel_without_memo_column(self, rows: list[list]) -> bytes:
        """
        Legacy-shaped header carrying ONLY 주문번호 / ISBN / 선택.

        ISBN + 선택 satisfy parse_daily_review_excel's header detection
        (excel_utils.py:784-791) and keep is_new_template False (:803), so the
        note column is looked up as '메모' — which is absent, leaving
        note_idx None (:825).
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["주문번호", "ISBN", "선택"])
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_absent_memo_column_is_inert_while_a_present_one_still_creates_the_note(
        self, auth_client
    ):
        """
        Both halves live in one test on purpose: parts (a)/(b) alone also pass
        against the un-implemented code, so only part (c) gives AC-MEMO-003 its
        inversion discriminating power.
        """
        # --- Part 1: no 메모 column at all ---------------------------------
        order_a = _make_order(shopify_order_id=190031, name="#M003A")
        li_nocol = _make_line_item(order_a, sku="S19-NOCOL", shopify_line_item_id=1)

        res_nocol = _post_file(
            auth_client, self._make_excel_without_memo_column([["#M003A", "S19-NOCOL", "북센"]])
        )

        # (a)
        assert res_nocol.status_code == 201
        assert LineItemNote.objects.filter(line_item=li_nocol).exists() is False
        # (b)
        li_nocol.refresh_from_db()
        assert li_nocol.confirmed_distributor == "booxen"
        assert PurchaseOrder.objects.filter(sku="S19-NOCOL").exists() is True
        assert res_nocol.data["confirmed_count"] == 1
        assert res_nocol.data["errors"] == []

        # --- Part 2: legacy file that DOES carry a 메모 column --------------
        order_b = _make_order(shopify_order_id=190032, name="#M003B")
        li_hascol = _make_line_item(order_b, sku="S19-HASCOL", shopify_line_item_id=1)

        res_hascol = _post_file(
            auth_client,
            _make_daily_review_excel([
                {
                    "order_name": "#M003B",
                    "isbn": "S19-HASCOL",
                    "selected": "북센",
                    "note": "레거시 메모",
                }
            ]),
        )
        assert res_hascol.status_code == 201

        # (c) — this is the scenario's inversion discriminator
        notes = list(LineItemNote.objects.filter(line_item=li_hascol))
        assert len(notes) == 1, f"expected 1 note from the legacy 메모 column, got {len(notes)}"
        assert notes[0].content == "레거시 메모"
        assert notes[0].assignee == "발주"


# ---------------------------------------------------------------------------
# T4 — AC-MEMO-004
# ---------------------------------------------------------------------------


def _make_memo_bearing_bulk_fixture(
    total_rows: int, sku_seed: str, order_name: str
) -> tuple[bytes, list[str], int]:
    """
    Same composition as _make_bulk_daily_review_fixture
    (test_daily_review_upload.py:2035-2108) with ONE deliberate difference:
    every PO-branch row carries its own distinct, non-empty memo instead of
    the `""` at :2083. That blank is exactly why the pre-existing ceiling test
    never measured this branch's note-writing cost (research.md §4.2).

    Returns (file_bytes, actionable_skus, po_row_count).
    """
    assert total_rows > 9 + _N_FILLER

    actionable_skus: list[str] = []
    rows: list[dict] = []

    for i, selected in enumerate(_ACTIONABLE_CS_SELECTS):
        sku = f"S19{sku_seed}CS{i}"
        actionable_skus.append(sku)
        rows.append({"sku": sku, "selected": selected, "status": "", "bs_price": 1000})

    for i, wh_status in enumerate(_ACTIONABLE_WAREHOUSE_STATUSES):
        sku = f"S19{sku_seed}WH{i}"
        actionable_skus.append(sku)
        rows.append({"sku": sku, "selected": "재고", "status": wh_status})

    n_po = total_rows - len(actionable_skus) - _N_FILLER
    for i in range(n_po):
        selected = _PO_BRANCH_SELECTS[i % len(_PO_BRANCH_SELECTS)]
        sku = f"S19{sku_seed}PO{i}"
        actionable_skus.append(sku)
        row = {"sku": sku, "selected": selected, "status": f"메모-{sku_seed}-{i}"}
        if i % 3 == 0:
            row["bs_price"] = 1200 + i
        if i % 4 == 0:
            row["ky_price"] = 3400 + i
        if i % 5 == 0:
            row["bs_stock"] = 10
        if selected == "BOOXEN" and "bs_price" not in row:
            row["bs_price"] = 5000
        rows.append(row)

    for i in range(_N_FILLER):
        selected = _FILLER_SELECTS[i % len(_FILLER_SELECTS)]
        rows.append({"sku": f"S19{sku_seed}F{i}", "selected": selected, "status": ""})

    for row in rows:
        row["order_name"] = order_name

    return _make_new_template_excel(rows), actionable_skus, n_po


@pytest.mark.django_db
class TestMemoBearingBulkUploadStaysUnderQueryCeiling:
    """T4 (AC-MEMO-004) — REQ-MEMO-009/010."""

    def _seed(self, actionable_skus: list[str], order_id: int) -> Order:
        order = _make_order(shopify_order_id=order_id, name=f"#{order_id}")
        for idx, sku in enumerate(actionable_skus):
            _make_line_item(order, sku=sku, quantity=2, shopify_line_item_id=idx + 1)
        warehouse_skus = actionable_skus[3:6]
        for sku, location in zip(warehouse_skus, _ACTIONABLE_WAREHOUSE_LOCATIONS):
            WarehouseStock.objects.create(isbn=sku, location=location, quantity=100)
        return order

    def _upload_and_capture(self, auth_client, total_rows: int, sku_seed: str, order_id: int):
        file_bytes, actionable_skus, n_po = _make_memo_bearing_bulk_fixture(
            total_rows, sku_seed, order_name=f"#{order_id}"
        )
        order = self._seed(actionable_skus, order_id)

        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        with CaptureQueriesContext(connection) as ctx:
            res = auth_client.post(
                UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart"
            )
        assert res.status_code == 201
        assert res.data["confirmed_count"] == len(actionable_skus)
        return order, len(ctx.captured_queries), n_po

    def test_memo_bearing_rows_do_not_add_per_row_queries(self, auth_client):
        order_300, q_300, n_po_300 = self._upload_and_capture(
            auth_client, 300, sku_seed="C3", order_id=190401
        )
        order_500, q_500, n_po_500 = self._upload_and_capture(
            auth_client, 500, sku_seed="C5", order_id=190501
        )

        # (b) absolute ceiling — a per-row create() lands here at ~n_po + base
        assert q_300 < UPLOAD_QUERY_CEILING, (
            f"expected under {UPLOAD_QUERY_CEILING} queries at 300 memo-bearing rows, "
            f"got {q_300}"
        )
        # (c)
        assert q_500 < UPLOAD_QUERY_CEILING, (
            f"expected under {UPLOAD_QUERY_CEILING} queries at 500 memo-bearing rows, "
            f"got {q_500}"
        )
        assert q_500 - q_300 <= 3, (
            f"query count scaled with row count: 300 -> {q_300}, 500 -> {q_500}"
        )

        # (d) the notes actually got written — one per PO-branch LineItem
        notes_300 = LineItemNote.objects.filter(
            line_item__order=order_300, assignee="발주"
        ).count()
        notes_500 = LineItemNote.objects.filter(
            line_item__order=order_500, assignee="발주"
        ).count()
        assert notes_300 == n_po_300, (
            f"expected {n_po_300} 발주 notes at 300 rows, got {notes_300} "
            f"(queries: 300 rows -> {q_300}, 500 rows -> {q_500})"
        )
        assert notes_500 == n_po_500, (
            f"expected {n_po_500} 발주 notes at 500 rows, got {notes_500} "
            f"(queries: 300 rows -> {q_300}, 500 rows -> {q_500})"
        )


# ---------------------------------------------------------------------------
# T5 — AC-MEMO-005
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestReuploadKeepsSingleNote:
    """T5 (AC-MEMO-005) — REQ-MEMO-011/012."""

    def test_same_file_uploaded_twice_keeps_exactly_one_note(self, auth_client):
        order = _make_order(shopify_order_id=190005, name="#M005")
        li = _make_line_item(order, sku="S19-IDEM", shopify_line_item_id=1)

        # The SAME bytes, wrapped twice — Django's test client consumes the
        # file object, so reusing the object (rather than the bytes) would
        # hand the second request an empty file and fake idempotency.
        file_bytes = _make_new_template_excel([
            {
                "order_name": "#M005",
                "sku": "S19-IDEM",
                "selected": "BOOXEN",
                "status": "중복 확인용 메모",
                "bs_price": 9000,
            }
        ])

        res1 = _post_file(auth_client, file_bytes)
        # (a)
        assert res1.status_code == 201
        assert res1.data["confirmed_count"] == 1
        assert res1.data["skipped_count"] == 0
        # (b)
        notes_after_first = list(LineItemNote.objects.filter(line_item=li))
        assert len(notes_after_first) == 1, (
            f"expected 1 note after the first upload, got {len(notes_after_first)}"
        )
        first_note_pk = notes_after_first[0].pk

        res2 = _post_file(auth_client, file_bytes)
        # (c)
        assert res2.status_code == 201
        assert res2.data["confirmed_count"] == 0
        assert res2.data["skipped_count"] == 1
        # (d)
        notes_after_second = list(LineItemNote.objects.filter(line_item=li))
        assert len(notes_after_second) == 1
        assert notes_after_second[0].pk == first_note_pk
        # (e)
        assert PurchaseOrder.objects.filter(sku="S19-IDEM").count() == 1


# ---------------------------------------------------------------------------
# T6 — AC-MEMO-006
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestThreeBranchesEachKeepTheirOwnAssignee:
    """T6 (AC-MEMO-006) — REQ-MEMO-013."""

    def test_cs_warehouse_and_distributor_rows_produce_three_distinct_notes(
        self, auth_client
    ):
        order = _make_order(shopify_order_id=190006, name="#M006")
        li_cs = _make_line_item(order, sku="S19-CS", shopify_line_item_id=1)
        li_wh = _make_line_item(order, sku="S19-WH", shopify_line_item_id=2)
        li_dist = _make_line_item(order, sku="S19-DIST", shopify_line_item_id=3)
        WarehouseStock.objects.create(isbn="S19-WH", location="korea", quantity=100)

        file_bytes = _make_new_template_excel([
            {
                "order_name": "#M006",
                "sku": "S19-CS",
                "selected": "주문취소",
                "status": "CS 쪽 메모",
            },
            {
                "order_name": "#M006",
                "sku": "S19-WH",
                "selected": "재고",
                "status": "한국재고",
            },
            {
                "order_name": "#M006",
                "sku": "S19-DIST",
                "selected": "BOOXEN",
                "status": "배포처 쪽 메모",
                "bs_price": 9000,
            },
        ])

        res = _post_file(auth_client, file_bytes)
        assert res.status_code == 201
        assert LineItemNote.objects.count() == 3, (
            f"expected 3 notes (one per branch), got {LineItemNote.objects.count()}"
        )

        # (a) CS branch — untouched by this SPEC
        note_cs = LineItemNote.objects.get(line_item=li_cs)
        assert note_cs.content == "CS 쪽 메모"
        assert note_cs.assignee == "CS"
        assert note_cs.note_type == "주문취소"
        li_cs.refresh_from_db()
        assert li_cs.purchase_status == "order_cancelled"

        # (b) warehouse branch — untouched by this SPEC
        note_wh = LineItemNote.objects.get(line_item=li_wh)
        assert note_wh.content == "한국재고"
        assert note_wh.assignee == "한국창고"
        assert note_wh.note_type == ""
        li_wh.refresh_from_db()
        assert li_wh.purchase_status == "in_stock"

        # (c) distributor branch — the new behaviour
        note_dist = LineItemNote.objects.get(line_item=li_dist)
        assert note_dist.content == "배포처 쪽 메모"
        assert note_dist.assignee == "발주"
        assert note_dist.note_type == ""
        li_dist.refresh_from_db()
        assert li_dist.confirmed_distributor == "booxen"

        # (d)
        assert len({note_cs.line_item_id, note_wh.line_item_id, note_dist.line_item_id}) == 3


# ---------------------------------------------------------------------------
# T7 — AC-MEMO-007
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestOneRowManyLineItemsGetsOneNoteEach:
    """T7 (AC-MEMO-007) — REQ-MEMO-001/014/015."""

    def test_note_is_attached_to_every_resolved_line_item(self, auth_client):
        order = _make_order(shopify_order_id=190007, name="#M007")
        li_a = _make_line_item(order, sku="S19-MULTI", quantity=2, shopify_line_item_id=1)
        li_b = _make_line_item(order, sku="S19-MULTI", quantity=3, shopify_line_item_id=2)

        memo_file = _make_new_template_excel([
            {
                "order_name": "#M007",
                "sku": "S19-MULTI",
                "selected": "BOOXEN",
                "status": "묶음 메모",
                "bs_price": 9000,
            }
        ])

        # Control: identical shape, blank memo, its own Order/SKU so the two
        # uploads cannot interfere through the shared (sku, distributor) group.
        order_n = _make_order(shopify_order_id=190017, name="#M007N")
        _make_line_item(order_n, sku="S19-MULTIN", quantity=2, shopify_line_item_id=1)
        _make_line_item(order_n, sku="S19-MULTIN", quantity=3, shopify_line_item_id=2)
        control_file = _make_new_template_excel([
            {
                "order_name": "#M007N",
                "sku": "S19-MULTIN",
                "selected": "BOOXEN",
                "status": "",
                "bs_price": 9000,
            }
        ])

        res_memo = _post_file(auth_client, memo_file)
        assert res_memo.status_code == 201
        res_control = _post_file(auth_client, control_file)
        assert res_control.status_code == 201

        notes = list(LineItemNote.objects.filter(line_item__in=[li_a, li_b]))
        # (a)
        assert len(notes) == 2, f"expected one note per resolved LineItem, got {len(notes)}"
        # (b)
        assert {n.line_item_id for n in notes} == {li_a.pk, li_b.pk}
        # (c)
        assert {n.content for n in notes} == {"묶음 메모"}
        assert {n.assignee for n in notes} == {"발주"}

        # (d) PO aggregation untouched
        pos = list(PurchaseOrder.objects.filter(sku="S19-MULTI"))
        assert len(pos) == 1
        assert pos[0].quantity == 5
        assert set(pos[0].line_items.all()) == {li_a, li_b}

        # (e) counters untouched
        assert res_memo.data["confirmed_count"] == res_control.data["confirmed_count"]
        assert res_memo.data["skipped_count"] == res_control.data["skipped_count"]


# ---------------------------------------------------------------------------
# T8 — AC-MEMO-008
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestMemoPresenceChangesNothingElse:
    """T8 (AC-MEMO-008) — REQ-MEMO-014/015/016."""

    _LI_VOLATILE_FIELDS = ("id", "order_id", "sku")
    _PO_VOLATILE_FIELDS = ("id", "sku", "created_at", "updated_at")

    def _build_set(self, shopify_order_id: int, order_name: str, prefix: str, memos: bool):
        order = _make_order(shopify_order_id=shopify_order_id, name=order_name)
        rows = []
        for idx, (selected, price_key, price) in enumerate(
            [
                ("BOOXEN", "bs_price", 9000),
                ("교보", "ky_price", 8500),
                ("YES24", "yes24_price", 8800),
            ],
            start=1,
        ):
            sku = f"{prefix}{idx}"
            _make_line_item(
                order, sku=sku, title="대조용 도서", quantity=2, shopify_line_item_id=idx
            )
            row = {
                "order_name": order_name,
                "sku": sku,
                "selected": selected,
                "status": f"메모 {idx}" if memos else "",
                price_key: price,
            }
            rows.append(row)
        return order, _make_new_template_excel(rows)

    def _snapshot(self, values, drop):
        return [
            {k: v for k, v in row.items() if k not in drop}
            for row in values
        ]

    def test_only_the_notes_differ(self, auth_client):
        order_a, file_a = self._build_set(190081, "#M008A", "S19-A", memos=True)
        order_b, file_b = self._build_set(190082, "#M008B", "S19-B", memos=False)

        res_a = _post_file(auth_client, file_a)
        assert res_a.status_code == 201
        li_a = self._snapshot(
            LineItem.objects.filter(order=order_a).order_by("pk").values(),
            self._LI_VOLATILE_FIELDS,
        )
        po_a = self._snapshot(
            PurchaseOrder.objects.filter(sku__startswith="S19-A").order_by("pk").values(),
            self._PO_VOLATILE_FIELDS,
        )

        res_b = _post_file(auth_client, file_b)
        assert res_b.status_code == 201
        li_b = self._snapshot(
            LineItem.objects.filter(order=order_b).order_by("pk").values(),
            self._LI_VOLATILE_FIELDS,
        )
        po_b = self._snapshot(
            PurchaseOrder.objects.filter(sku__startswith="S19-B").order_by("pk").values(),
            self._PO_VOLATILE_FIELDS,
        )

        # (a)
        assert res_a.data["confirmed_count"] == res_b.data["confirmed_count"]
        assert res_a.data["skipped_count"] == res_b.data["skipped_count"]
        assert res_a.data["errors"] == res_b.data["errors"]
        assert {
            code: len(items) for code, items in res_a.data["confirmed_by_distributor"].items()
        } == {
            code: len(items) for code, items in res_b.data["confirmed_by_distributor"].items()
        }
        # (b)
        assert li_a == li_b
        # (c)
        assert po_a == po_b
        # (d)
        assert LineItemNote.objects.filter(line_item__order=order_a).count() == 3
        assert LineItemNote.objects.filter(line_item__order=order_b).count() == 0

        # (e) no model change was needed to get here
        try:
            call_command("makemigrations", "order", check=True, dry_run=True, verbosity=0)
        except SystemExit as exc:  # pragma: no cover - only on a model change
            pytest.fail(f"makemigrations --check reported pending order model changes: {exc}")


# ---------------------------------------------------------------------------
# T9 — AC-MEMO-009
# ---------------------------------------------------------------------------


def _excel_strings(file_bytes: bytes) -> set[str]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    values: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    values.add(str(cell))
    return values


@pytest.mark.django_db
class TestNewNoteIsRecordedButNeverQueuedOrExported:
    """T9 (AC-MEMO-009) — REQ-MEMO-017/018."""

    def test_note_is_recorded_but_absent_from_unresolved_list_and_publisher_export(
        self, auth_client
    ):
        order = _make_order(shopify_order_id=190009, name="#M009")
        li = _make_line_item(
            order, sku="S19-SHOW", title="표시 확인용 도서", shopify_line_item_id=1
        )

        # Control: a genuine 타출판사 note on a different LineItem, so an empty
        # export cannot masquerade as "our note did not leak".
        control_order = _make_order(shopify_order_id=190019, name="#M009C")
        control_li = _make_line_item(control_order, sku="S19-CTRL", shopify_line_item_id=1)
        LineItemNote.objects.create(
            line_item=control_li, content="아가페", note_type="타출판사", is_resolved=False
        )

        file_bytes = _make_new_template_excel([
            {
                "order_name": "#M009",
                "sku": "S19-SHOW",
                "selected": "BOOXEN",
                "status": "표시 확인 메모",
                "bs_price": 9000,
            }
        ])
        assert _post_file(auth_client, file_bytes).status_code == 201

        # (a) the memo IS recorded, attached to the right line item
        note = LineItemNote.objects.get(line_item=li, content="표시 확인 메모")
        assert note.assignee == "발주"
        assert note.is_resolved is True

        # (b) ...and precisely because it is already resolved it never enters the
        # unresolved queue, which is the sole data source for the CS / 발주 /
        # 타출판사 tabs. A memo explaining why a row was confirmed is a record,
        # not a task for someone to work off.
        res_list = auth_client.get(NOTES_UNRESOLVED_URL)
        assert res_list.status_code == 200
        matching = [
            row for row in res_list.data if row["content"] == "표시 확인 메모"
        ]
        assert matching == [], (
            f"the memo must not reach the unresolved queue, found {len(matching)}"
        )

        # ...while the endpoint demonstrably works — the control note is returned,
        # so an empty result above cannot be a broken query masquerading as a pass.
        assert any(row["content"] == "아가페" for row in res_list.data)

        # (c) it stays readable in the line item's own note history, which the
        # order detail page serializes without an is_resolved filter
        res_history = auth_client.get(f"/api/orders/line-items/{li.pk}/notes/")
        assert res_history.status_code == 200
        history = [row for row in res_history.data if row["content"] == "표시 확인 메모"]
        assert len(history) == 1
        assert history[0]["assignee"] == "발주"

        # (2) 타출판사 export — publisher=other is the bucket a leaked note
        # (note_type="타출판사", content not 아가페/성서유니온) would land in.
        res_other = auth_client.get(NOTES_EXPORT_URL, {"publisher": "other"})
        assert res_other.status_code == 200
        other_strings = _excel_strings(res_other.content)
        # (c)
        assert "표시 확인 메모" not in other_strings
        assert "S19-SHOW" not in other_strings

        # ...and the export itself demonstrably works: the control note is
        # returned by its own publisher bucket.
        res_agape = auth_client.get(NOTES_EXPORT_URL, {"publisher": "agape"})
        assert res_agape.status_code == 200
        agape_strings = _excel_strings(res_agape.content)
        assert "아가페" in agape_strings
        assert "S19-CTRL" in agape_strings
