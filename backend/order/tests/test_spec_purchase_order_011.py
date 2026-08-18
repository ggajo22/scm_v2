"""
TDD tests for SPEC-PURCHASE-ORDER-011 — 파손 교환 신청 페이지 (backend M1-M5).

Each test class/method traces the REQ-DEX-XXX / AC-DEX-XXX ids from spec.md
via the Given/When/Then scenarios in acceptance.md. See plan.md for the
file-by-file implementation this suite exercises:

  M1    LineItem.damaged_quantity field + migration     -> AC-DEX-001
  M1.5  _recompute_order_aggregates characterization     -> see
        test_spec_012.py::TestRecomputeOrderAggregatesReadyToShip (not
        duplicated here — this SPEC only adds new damaged_exchange cases).
  M2    DamagedExchangeSearchView                        -> AC-DEX-003~007
  M3    DamagedExchangeSubmitView                         -> AC-DEX-008~011
  M4    UnorderedItemsView damaged_quantity substitution  -> AC-DEX-012~012c
  M4.5  _recompute_order_aggregates damaged_exchange
        short-circuit                                     -> AC-DEX-009c
  M4.6  ready_to_ship backfill migration (0040)            -> AC-DEX-013/013a
  (regression) five/six untouched reorder-quantity sites   -> AC-DEX-012d/e/f
"""

import io

import openpyxl
import pytest
from django.apps import apps as global_apps
from django.contrib.auth import get_user_model
from django.db import connection
from django.db.migrations.loader import MigrationLoader
from django.db.models import Sum
from django.test.utils import CaptureQueriesContext
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from order.models import (
    LineItem,
    Order,
    PurchaseOrder,
    Refund,
    VendorComparison,
    WarehouseStock,
)
from order.purchase_order_views import _recompute_order_aggregates

User = get_user_model()

SEARCH_URL = "/api/purchase-orders/line-items/damaged-exchange-search/"
SUBMIT_URL = "/api/purchase-orders/line-items/{pk}/damaged-exchange/"
UNORDERED_URL = "/api/purchase-orders/unordered/"
RUN_COMPARISON_URL = "/api/purchase-orders/run-comparison/"
GENERATE_URL = "/api/purchase-orders/generate-order-file/"
DAILY_REVIEW_EXCEL_URL = "/api/purchase-orders/daily-review-excel/"
UPLOAD_DAILY_URL = "/api/purchase-orders/upload-daily-review/"
VENDOR_COMPARISON_URL = "/api/purchase-orders/comparison/"
CONFIRM_URL = "/api/purchase-orders/confirm/"
LINE_ITEM_STATUS_URL = "/api/purchase-orders/line-items/{pk}/status/"
BULK_STATUS_URL = "/api/purchase-orders/line-items/bulk-status/"


# ---------------------------------------------------------------------------
# Fixtures / factories
# ---------------------------------------------------------------------------


@pytest.fixture
def user(db):
    return User.objects.create_user(username="dex_user", password="testpass123")


@pytest.fixture
def auth_client(user):
    client = APIClient()
    refresh = RefreshToken.for_user(user)
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {refresh.access_token}")
    return client


def _make_order(shopify_order_id: int, **kwargs) -> Order:
    return Order.objects.create(shopify_order_id=shopify_order_id, store_type="gimssine", **kwargs)


def _make_line_item(order: Order, shopify_line_item_id: int, sku, **kwargs) -> LineItem:
    defaults = {"quantity": 1, "title": "Test Book"}
    defaults.update(kwargs)
    return LineItem.objects.create(
        order=order, shopify_line_item_id=shopify_line_item_id, sku=sku, **defaults
    )


def _make_refund(
    order: Order, shopify_line_item_id: int, quantity: int, shopify_refund_id: int = 1
) -> Refund:
    return Refund.objects.create(
        order=order,
        shopify_refund_id=shopify_refund_id,
        line_item_id=shopify_line_item_id,
        quantity=quantity,
    )


def _make_daily_review_excel(rows: list[dict]) -> bytes:
    """Mirrors test_daily_review_upload.py's `_make_daily_review_excel`."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "주문번호", "ISBN", "제목", "수량", "주문위치", "메모",
        "재고(한국)", "재고(CA)", "재고(NJ)",
        "북센 공급가", "교보 공급가",
        "북센 재고수량", "북센 재고상태",
        "교보 재고수량", "교보 재고상태",
        "공급가차이", "북센 입고예정", "북센 반품가능여부",
        "교보 출고여부", "교보 반품가능여부",
        "가격차이알림", "출판사", "선택근거", "선택",
    ])
    for row in rows:
        ws.append([
            row.get("order_name", ""),
            row.get("isbn", ""),
            row.get("title", ""),
            row.get("quantity", 1),
            "",
            row.get("note", ""),
            row.get("korea_stock", 0),
            row.get("ca_stock", 0),
            row.get("nj_stock", 0),
            row.get("bs_price", ""),
            "", "", "", "", "", "", "", "", "", "", "", "", "",
            row.get("selected", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _file_obj(file_bytes: bytes, name: str = "upload.xlsx"):
    f = io.BytesIO(file_bytes)
    f.name = name
    return f


# ---------------------------------------------------------------------------
# M1 — AC-DEX-001
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestLineItemDamagedQuantityField:
    def test_defaults_to_zero_and_leaves_other_fields_untouched(self):
        order = _make_order(1100001)
        li = _make_line_item(order, 1, "SKU-DEX-001", quantity=10, purchase_status="unordered")
        li.refresh_from_db()
        assert li.damaged_quantity == 0
        assert li.quantity == 10
        assert li.purchase_status == "unordered"


# ---------------------------------------------------------------------------
# M2 — DamagedExchangeSearchView — AC-DEX-003~007
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestDamagedExchangeSearch:
    def test_exact_match_only_substring_excluded(self, auth_client):
        """AC-DEX-003."""
        order = _make_order(1100010)
        _make_line_item(order, 1, "9788956609959", quantity=5, purchase_status="unordered")
        resp = auth_client.get(SEARCH_URL, {"sku": "978895660995"})
        assert resp.status_code == 200
        assert resp.data == {"count": 0, "results": []}

    def test_unshipped_and_not_cancelled_scope_reused(self, auth_client):
        """AC-DEX-004: same 미출고 definition as LineItemRackNumberSummaryView."""
        order = _make_order(1100011)
        keep = _make_line_item(
            order, 1, "SKU-DEX-004", logistics_status="not_shipped", purchase_status="unordered"
        )
        _make_line_item(
            order, 2, "SKU-DEX-004", logistics_status="shipped", purchase_status="unordered"
        )
        _make_line_item(
            order, 3, "SKU-DEX-004",
            logistics_status="not_shipped", purchase_status="order_cancelled",
        )
        resp = auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-004"})
        assert resp.data["count"] == 1
        assert resp.data["results"][0]["id"] == keep.id

    def test_already_damaged_exchange_row_included_and_flagged(self, auth_client):
        """AC-DEX-004a."""
        order = _make_order(1100012)
        li = _make_line_item(
            order, 1, "SKU-DEX-004A",
            logistics_status="not_shipped", purchase_status="damaged_exchange",
        )
        resp = auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-004A"})
        assert resp.data["count"] == 1
        row = resp.data["results"][0]
        assert row["id"] == li.id
        assert row["is_damaged_exchange"] is True

    def test_order_total_quantity_sums_only_trackable_non_cancelled_rows(self, auth_client):
        """AC-DEX-005/005b."""
        order = _make_order(1100013)
        target = _make_line_item(order, 1, "SKU-DEX-005", quantity=4, purchase_status="unordered")
        _make_line_item(order, 2, "SKU-DEX-005-OTHER", quantity=9, purchase_status="unordered")
        _make_line_item(
            order, 3, "SKU-DEX-005-CANCELLED", quantity=100, purchase_status="order_cancelled"
        )
        _make_line_item(order, 4, None, quantity=50)
        resp = auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-005"})
        row = resp.data["results"][0]
        assert row["id"] == target.id
        assert row["quantity"] == 4
        assert row["order_total_quantity"] == 13

    def test_search_response_exposes_stored_rack_number(self, auth_client):
        """The search row carries LineItem.rack_number (SPEC-ORDER-013
        REQ-RACK-001) read-only, so the operator can locate the damaged copy
        without switching to /rack-number.

        Two rows under one SKU with *different* rack_number values: an
        implementation that hardcoded a constant, echoed the unassigned
        default, or read the wrong field would fail. The unassigned row also
        pins that an empty rack_number stays the empty string in the payload
        (the '미지정' substitution is a frontend concern, not an API one).
        """
        order = _make_order(1100016)
        assigned = _make_line_item(
            order, 1, "SKU-DEX-RACK", purchase_status="unordered", rack_number="B-12"
        )
        order_b = _make_order(1100017)
        unassigned = _make_line_item(
            order_b, 1, "SKU-DEX-RACK", purchase_status="unordered", rack_number=""
        )

        resp = auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-RACK"})

        assert resp.status_code == 200
        by_id = {row["id"]: row for row in resp.data["results"]}
        assert by_id[assigned.id]["rack_number"] == "B-12"
        assert by_id[unassigned.id]["rack_number"] == ""

    def test_search_does_not_write_rack_number(self, auth_client):
        """rack_number is read-only here — editing stays exclusive to
        SPEC-ORDER-013's endpoints. Serving a search must not touch it."""
        order = _make_order(1100018)
        li = _make_line_item(
            order, 1, "SKU-DEX-RACK-RO", purchase_status="unordered", rack_number="C-07"
        )

        auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-RACK-RO"})

        li.refresh_from_db()
        assert li.rack_number == "C-07"

    def test_ready_to_ship_null_not_coerced_to_false(self, auth_client):
        """AC-DEX-005a."""
        order = _make_order(1100014, ready_to_ship=None)
        _make_line_item(
            order, 1, "SKU-DEX-005A", purchase_status="unordered", logistics_status="not_shipped"
        )
        resp = auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-005A"})
        assert resp.data["results"][0]["order_ready_to_ship"] is None

    def test_order_total_quantity_all_null_reports_zero_not_none(self, auth_client):
        """AC-DEX-005c: SQL SUM ignores NULL — only an all-null fixture is
        discriminating (v1.3.0 D20)."""
        order = _make_order(1100015)
        target = _make_line_item(
            order, 1, "SKU-DEX-005C", quantity=None, purchase_status="unordered"
        )
        _make_line_item(order, 2, "SKU-DEX-005C-B", quantity=None, purchase_status="unordered")
        resp = auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-005C"})
        row = resp.data["results"][0]
        assert row["id"] == target.id
        assert row["order_total_quantity"] == 0

    def test_ready_to_ship_is_read_only_no_recompute_triggered(self, auth_client):
        """AC-DEX-006: stored value shown as-is; serving the search request
        must not call _recompute_order_aggregates."""
        order = _make_order(1100016, ready_to_ship=True)
        _make_line_item(
            order, 1, "SKU-DEX-006", purchase_status="unordered", logistics_status="not_shipped"
        )
        # Added AFTER ready_to_ship=True was stored — a fresh recompute would
        # now yield False via the pre-existing cs_required short-circuit.
        _make_line_item(order, 2, "SKU-DEX-006-CS", quantity=1, purchase_status="cs_required")

        resp = auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-006"})
        assert resp.data["results"][0]["order_ready_to_ship"] is True
        order.refresh_from_db()
        assert order.ready_to_ship is True

    def test_no_match_returns_empty_list_with_http_200(self, auth_client):
        """AC-DEX-007."""
        resp = auth_client.get(SEARCH_URL, {"sku": "0000000000000"})
        assert resp.status_code == 200
        assert resp.data == {"count": 0, "results": []}

    def test_query_count_fixed_regardless_of_matched_row_count(self, auth_client):
        """N+1 regression guard (plan.md 기술적 접근 > 검색 엔드포인트)."""
        order_a = _make_order(1100017)
        order_b = _make_order(1100018)
        _make_line_item(order_a, 1, "SKU-DEX-QC-SINGLE", purchase_status="unordered")
        for i in range(2, 12):
            _make_line_item(order_b, i, "SKU-DEX-QC-MULTI", purchase_status="unordered")

        with CaptureQueriesContext(connection) as ctx_single:
            resp_single = auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-QC-SINGLE"})
        assert resp_single.data["count"] == 1

        with CaptureQueriesContext(connection) as ctx_multi:
            resp_multi = auth_client.get(SEARCH_URL, {"sku": "SKU-DEX-QC-MULTI"})
        assert resp_multi.data["count"] == 10

        assert len(ctx_single.captured_queries) == len(ctx_multi.captured_queries)


# ---------------------------------------------------------------------------
# M3 — DamagedExchangeSubmitView — AC-DEX-008~011
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestDamagedExchangeSubmit:
    def test_rejects_value_above_quantity_server_side(self, auth_client):
        """AC-DEX-008."""
        order = _make_order(1100020)
        li = _make_line_item(order, 1, "SKU-DEX-008", quantity=5, purchase_status="unordered")
        resp = auth_client.post(SUBMIT_URL.format(pk=li.id), {"damaged_quantity": 6}, format="json")
        assert resp.status_code == 400
        li.refresh_from_db()
        assert li.purchase_status == "unordered"
        assert li.damaged_quantity == 0

    def test_valid_submission_sets_status_and_quantity_without_touching_ready_to_ship(
        self, auth_client
    ):
        """AC-DEX-009 (개정): the damaged_exchange short-circuit was removed —
        ready_to_ship now asks only "is every live item 입고/재고?", so a
        damaged_exchange row that REQ-DEX-011 leaves at logistics_status=
        "received" keeps the order 출고가능. The submission still writes
        purchase_status/damaged_quantity; it just no longer flips the badge."""
        order = _make_order(1100021)
        li = _make_line_item(
            order, 1, "SKU-DEX-009", quantity=8,
            purchase_status="unordered", logistics_status="received",
        )
        _recompute_order_aggregates([order.id])
        order.refresh_from_db()
        assert order.ready_to_ship is True  # pre-condition

        resp = auth_client.post(SUBMIT_URL.format(pk=li.id), {"damaged_quantity": 3}, format="json")
        assert resp.status_code == 200

        li.refresh_from_db()
        order.refresh_from_db()
        assert li.purchase_status == "damaged_exchange"
        assert li.damaged_quantity == 3
        assert li.logistics_status == "received"
        assert order.ready_to_ship is True

    def test_out_of_range_rejected_state_unchanged(self, auth_client):
        """AC-DEX-009a."""
        order = _make_order(1100022)
        li = _make_line_item(order, 1, "SKU-DEX-009A", quantity=5, purchase_status="unordered")
        resp = auth_client.post(SUBMIT_URL.format(pk=li.id), {"damaged_quantity": 0}, format="json")
        assert resp.status_code == 400
        li.refresh_from_db()
        assert li.purchase_status == "unordered"
        assert li.damaged_quantity == 0

        li_null_qty = _make_line_item(
            order, 2, "SKU-DEX-009A-NULL", quantity=None, purchase_status="unordered"
        )
        resp2 = auth_client.post(
            SUBMIT_URL.format(pk=li_null_qty.id), {"damaged_quantity": 1}, format="json"
        )
        assert resp2.status_code == 400
        li_null_qty.refresh_from_db()
        assert li_null_qty.purchase_status == "unordered"

    def test_resubmission_overwrites_not_accumulates(self, auth_client):
        """AC-DEX-009b."""
        order = _make_order(1100023)
        li = _make_line_item(
            order, 1, "SKU-DEX-009B", quantity=8,
            purchase_status="damaged_exchange", damaged_quantity=3,
        )
        resp = auth_client.post(SUBMIT_URL.format(pk=li.id), {"damaged_quantity": 2}, format="json")
        assert resp.status_code == 200
        li.refresh_from_db()
        assert li.damaged_quantity == 2

    def test_damaged_exchange_no_longer_short_circuits_ready_to_ship(self):
        """AC-DEX-009c (개정): both rows satisfy 입고/재고, so the order is
        출고가능. The short-circuit that used to force False here was removed
        on user instruction — the only remaining rule is "every live item is
        입고 or 재고"."""
        order = _make_order(1100024)
        _make_line_item(
            order, 1, "SKU-DEX-009C-A",
            purchase_status="damaged_exchange", logistics_status="received",
        )
        _make_line_item(
            order, 2, "SKU-DEX-009C-B",
            purchase_status="unordered", logistics_status="received",
        )
        _recompute_order_aggregates([order.id])
        order.refresh_from_db()
        assert order.ready_to_ship is True
        # REQ-DEX-009d: the status rule was never coupled to the short-circuit.
        assert order.status == "received"

    def test_note_auto_created_with_author_type_assignee_and_quantity(self, auth_client, user):
        """AC-DEX-010."""
        order = _make_order(1100025)
        li = _make_line_item(order, 1, "SKU-DEX-010", quantity=8, purchase_status="unordered")
        assert li.notes.count() == 0

        resp = auth_client.post(SUBMIT_URL.format(pk=li.id), {"damaged_quantity": 3}, format="json")
        assert resp.status_code == 200

        notes = list(li.notes.all())
        assert len(notes) == 1
        note = notes[0]
        assert note.note_type == "파손/교환"
        assert note.assignee == "발주"
        assert note.author_id == user.id
        assert "3" in note.content

    def test_logistics_status_and_shipped_quantity_untouched(self, auth_client):
        """AC-DEX-011."""
        order = _make_order(1100026)
        li = _make_line_item(
            order, 1, "SKU-DEX-011", quantity=8,
            logistics_status="not_shipped", shipped_quantity=0,
        )
        resp = auth_client.post(SUBMIT_URL.format(pk=li.id), {"damaged_quantity": 4}, format="json")
        assert resp.status_code == 200
        li.refresh_from_db()
        assert li.logistics_status == "not_shipped"
        assert li.shipped_quantity == 0


# ---------------------------------------------------------------------------
# M4 — UnorderedItemsView reorder-quantity correction — AC-DEX-012~012c
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUnorderedItemsViewDamagedQuantitySubstitution:
    def test_damaged_exchange_uses_damaged_quantity_not_quantity(self, auth_client):
        """AC-DEX-012."""
        order = _make_order(1100030)
        li = _make_line_item(
            order, 1, "SKU-DEX-012", quantity=10, damaged_quantity=3,
            purchase_status="damaged_exchange",
        )
        resp = auth_client.get(UNORDERED_URL)
        row = next(r for r in resp.data["results"] if r["id"] == li.id)
        assert row["quantity"] == 3

    def test_refund_subtraction_applies_on_top_of_damaged_quantity(self, auth_client):
        """AC-DEX-012a (결정 A)."""
        order = _make_order(1100031)
        li = _make_line_item(
            order, 1, "SKU-DEX-012A", quantity=10, damaged_quantity=3,
            purchase_status="damaged_exchange",
        )
        _make_refund(order, li.shopify_line_item_id, quantity=1)
        resp = auth_client.get(UNORDERED_URL)
        row = next(r for r in resp.data["results"] if r["id"] == li.id)
        assert row["quantity"] == 2

    def test_zero_net_quantity_excludes_row_via_damaged_quantity_base(self, auth_client):
        """AC-DEX-012b: quantity=10 pinned so an unmodified (quantity-based)
        implementation would compute max(10-1,0)=9 and wrongly include the row."""
        order = _make_order(1100032)
        li = _make_line_item(
            order, 1, "SKU-DEX-012B", quantity=10, damaged_quantity=1,
            purchase_status="damaged_exchange",
        )
        _make_refund(order, li.shopify_line_item_id, quantity=1)
        resp = auth_client.get(UNORDERED_URL)
        assert all(r["id"] != li.id for r in resp.data["results"])

    def test_non_damaged_exchange_row_unaffected_regression(self, auth_client):
        """AC-DEX-012c."""
        order = _make_order(1100033)
        li = _make_line_item(
            order, 1, "SKU-DEX-012C", quantity=10, damaged_quantity=0,
            purchase_status="unordered",
        )
        resp = auth_client.get(UNORDERED_URL)
        row = next(r for r in resp.data["results"] if r["id"] == li.id)
        assert row["quantity"] == 10


# ---------------------------------------------------------------------------
# M4.6 — ready_to_ship backfill migration (0040) — AC-DEX-013/013a
# ---------------------------------------------------------------------------


def _run_dex_backfill_migration():
    loader = MigrationLoader(connection)
    migration = loader.get_migration("order", "0040_backfill_ready_to_ship_damaged_exchange")
    operation = migration.operations[0]
    operation.code(global_apps, None)


@pytest.mark.django_db
class TestBackfillReadyToShipDamagedExchangeMigration:
    def test_stale_true_updated_to_false_for_damaged_exchange_order(self):
        """AC-DEX-013 (v1.3.1 D24: logistics_status pinned to 'received' —
        under 'not_shipped' the pre-existing rule alone already yields False,
        which would let a migration missing the short-circuit pass too)."""
        order = _make_order(1100040, ready_to_ship=True)
        _make_line_item(
            order, 1, "SKU-DEX-013",
            purchase_status="damaged_exchange", logistics_status="received",
        )
        _run_dex_backfill_migration()
        order.refresh_from_db()
        assert order.ready_to_ship is False

    def test_unrelated_order_without_damaged_exchange_is_unaffected(self):
        """AC-DEX-013a."""
        order = _make_order(1100041, ready_to_ship=True)
        _make_line_item(
            order, 1, "SKU-DEX-013A",
            purchase_status="in_stock", logistics_status="not_shipped",
        )
        _run_dex_backfill_migration()
        order.refresh_from_db()
        assert order.ready_to_ship is True

    def test_migration_reverse_code_is_noop(self):
        from django.db import migrations as dj_migrations

        loader = MigrationLoader(connection)
        migration = loader.get_migration("order", "0040_backfill_ready_to_ship_damaged_exchange")
        operation = migration.operations[0]
        assert operation.reverse_code is dj_migrations.RunPython.noop

    def test_migration_dependency_chain(self):
        loader = MigrationLoader(connection)
        m0039 = loader.disk_migrations[("order", "0039_lineitem_damaged_quantity")]
        m0040 = loader.disk_migrations[("order", "0040_backfill_ready_to_ship_damaged_exchange")]
        assert m0039.dependencies == [("order", "0038_refund_unique_per_line_item")]
        assert m0040.dependencies == [("order", "0039_lineitem_damaged_quantity")]

    def test_query_count_fixed_regardless_of_order_count(self):
        """1 SELECT + 1 bulk_update, independent of Order count. The
        MigrationLoader is constructed OUTSIDE the capture window — building
        it issues its own introspection/bookkeeping queries unrelated to the
        migration's own RunPython operation under test."""
        for i in range(5):
            order = _make_order(1100050 + i, ready_to_ship=True)
            _make_line_item(
                order, 1, f"SKU-DEX-QC-{i}",
                purchase_status="damaged_exchange", logistics_status="received",
            )
        loader = MigrationLoader(connection)
        migration = loader.get_migration("order", "0040_backfill_ready_to_ship_damaged_exchange")
        operation = migration.operations[0]
        with CaptureQueriesContext(connection) as ctx:
            operation.code(global_apps, None)
        assert len(ctx.captured_queries) <= 2


# ---------------------------------------------------------------------------
# Regression — the six untouched reorder-quantity sites (Exclusions,
# REQ-DEX-012b/012c) — AC-DEX-012d/012e/012f
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestFiveUntouchedReorderQuantitySites:
    """A damaged_exchange LineItem with quantity=10, damaged_quantity=3 must
    still be read as `quantity` (10), never `damaged_quantity` (3), at every
    site REQ-DEX-012b/012c explicitly excludes from the REQ-DEX-012
    substitution."""

    def test_run_comparison_view_uses_quantity(self, auth_client):
        """AC-DEX-012d (1/4)."""
        order = _make_order(1100060)
        _make_line_item(
            order, 1, "SKU-DEX-012D-A", quantity=10, damaged_quantity=3,
            purchase_status="damaged_exchange",
        )
        resp = auth_client.post(RUN_COMPARISON_URL)
        assert resp.status_code == 200
        row = next(r for r in resp.data["results"] if r["sku"] == "SKU-DEX-012D-A")
        assert row["total_qty"] == 10

    def test_generate_order_file_view_uses_quantity(self, auth_client):
        """AC-DEX-012d (2/4)."""
        order = _make_order(1100061)
        _make_line_item(
            order, 1, "SKU-DEX-012D-B", quantity=10, damaged_quantity=3,
            purchase_status="damaged_exchange",
        )
        resp = auth_client.post(
            GENERATE_URL,
            {"distributor": "kyobo", "skus": ["SKU-DEX-012D-B"]},
            format="json",
        )
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # kyobo layout: header row ["ISBN", "수량"], data starts row 2.
        assert ws.cell(row=2, column=1).value == "SKU-DEX-012D-B"
        assert ws.cell(row=2, column=2).value == 10

    def test_daily_review_excel_view_uses_quantity(self, auth_client):
        """AC-DEX-012d (3/4)."""
        order = _make_order(1100062)
        _make_line_item(
            order, 1, "SKU-DEX-012D-C", quantity=10, damaged_quantity=3,
            purchase_status="damaged_exchange",
        )
        resp = auth_client.get(DAILY_REVIEW_EXCEL_URL)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        sku_idx = header.index("ISBN")
        qty_idx = header.index("수량")
        matching = [r for r in rows[1:] if r[sku_idx] == "SKU-DEX-012D-C"]
        assert len(matching) == 1
        assert matching[0][qty_idx] == 10

    def test_upload_daily_review_view_uses_quantity_for_warehouse_deduction(self, auth_client):
        """AC-DEX-012d (4/4): total_qty (used to deduct WarehouseStock) is
        built from raw LineItem.quantity (10), not damaged_quantity (3)."""
        sku = "SKU-DEX-012D-D"
        order = _make_order(1100063, name="#DEX-012D-D")
        _make_line_item(
            order, 1, sku, quantity=10, damaged_quantity=3,
            purchase_status="damaged_exchange",
        )
        WarehouseStock.objects.create(isbn=sku, quantity=100, location="korea")

        file_bytes = _make_daily_review_excel([
            {"order_name": "#DEX-012D-D", "isbn": sku, "quantity": 10, "selected": "재고(한국)"},
        ])
        resp = auth_client.post(
            UPLOAD_DAILY_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert resp.status_code == 201

        stock = WarehouseStock.objects.get(isbn=sku, location="korea")
        assert stock.quantity == 90  # 100 - 10 (quantity), not 100 - 3 (damaged_quantity)

    def test_vendor_comparison_view_qty_by_sku_uses_quantity_whitebox(self, auth_client):
        """AC-DEX-012e (v1.3.1 D23 — deliberate white-box assertion:
        VendorComparisonView's response rows carry no quantity field at all,
        so qty_by_sku is the only observable this fixture can discriminate).
        Scoped to a PurchaseOrder-unlinked row (v1.3.0 D21)."""
        order = _make_order(1100064)
        _make_line_item(
            order, 1, "SKU-DEX-012E", quantity=10, damaged_quantity=3,
            purchase_status="damaged_exchange",
        )
        VendorComparison.objects.create(sku="SKU-DEX-012E")

        # White-box: reproduce VendorComparisonView's own qty_by_sku query
        # (purchase_order_views.py) directly — its response rows carry no
        # quantity field, so this is the only observable this fixture can
        # discriminate (v1.3.1 D23).
        qty_by_sku = dict(
            LineItem.objects.filter(purchase_orders__isnull=True, sku__isnull=False)
            .values("sku")
            .annotate(total=Sum("quantity"))
            .values_list("sku", "total")
        )
        assert qty_by_sku.get("SKU-DEX-012E") == 10

        # Also confirm the endpoint itself still responds successfully with
        # this fixture present (regression — no crash/behavior change).
        resp = auth_client.get(VENDOR_COMPARISON_URL)
        assert resp.status_code == 200
        assert any(r["sku"] == "SKU-DEX-012E" for r in resp.data["results"])

    def test_confirm_order_view_uses_request_body_quantity_not_line_item_fields(self, auth_client):
        """AC-DEX-012f: ConfirmOrderView never reads LineItem.quantity or
        LineItem.damaged_quantity — it uses the request body's quantity
        verbatim, regardless of what those two LineItem fields hold."""
        order = _make_order(1100065)
        _make_line_item(
            order, 1, "SKU-DEX-012F", quantity=10, damaged_quantity=3,
            purchase_status="damaged_exchange",
        )
        resp = auth_client.post(
            CONFIRM_URL,
            {
                "items": [
                    {
                        "sku": "SKU-DEX-012F",
                        "distributor": "booxen",
                        "quantity": 7,
                        "unit_price": "1000",
                    }
                ]
            },
            format="json",
        )
        assert resp.status_code == 201
        po = PurchaseOrder.objects.get(sku="SKU-DEX-012F")
        assert po.quantity == 7


# ---------------------------------------------------------------------------
# Legacy write-path blocking — damage intake is exclusive to
# DamagedExchangeSubmitView (coordinator decision following the
# REQ-DEX-012 vs SPEC-010 conflict this suite's full-run originally
# surfaced). Verifies the three still-live entry points that could
# previously set purchase_status="damaged_exchange" without ever setting
# damaged_quantity are now rejected — the field is unset only through paths
# blocked here, so a damaged_exchange row with damaged_quantity=0 is no
# longer a state production code can reach.
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestDamagedExchangeLegacyWritePathsBlocked:
    def test_single_status_update_rejects_damaged_exchange(self, auth_client):
        """LineItemStatusUpdateView PATCH .../status/ rejects
        purchase_status="damaged_exchange" with HTTP 400; every other
        choice is unaffected."""
        order = _make_order(1100070)
        li = _make_line_item(order, 1, "SKU-DEX-BLOCK-1", quantity=5, purchase_status="unordered")

        resp = auth_client.patch(
            LINE_ITEM_STATUS_URL.format(pk=li.id),
            {"purchase_status": "damaged_exchange"},
            format="json",
        )
        assert resp.status_code == 400

        li.refresh_from_db()
        assert li.purchase_status == "unordered"
        assert li.damaged_quantity == 0

        # Sanity: an unrelated valid choice still works on the same view.
        resp2 = auth_client.patch(
            LINE_ITEM_STATUS_URL.format(pk=li.id),
            {"purchase_status": "cs_required"},
            format="json",
        )
        assert resp2.status_code == 200
        li.refresh_from_db()
        assert li.purchase_status == "cs_required"

    def test_bulk_status_update_rejects_damaged_exchange(self, auth_client):
        """LineItemBulkStatusUpdateView PATCH .../bulk-status/ rejects
        purchase_status="damaged_exchange" with HTTP 400 for the whole
        batch; no row is updated."""
        order = _make_order(1100071)
        li1 = _make_line_item(order, 1, "SKU-DEX-BLOCK-2A", quantity=5, purchase_status="unordered")
        li2 = _make_line_item(order, 2, "SKU-DEX-BLOCK-2B", quantity=5, purchase_status="unordered")

        resp = auth_client.patch(
            BULK_STATUS_URL,
            {"ids": [li1.id, li2.id], "purchase_status": "damaged_exchange"},
            format="json",
        )
        assert resp.status_code == 400

        li1.refresh_from_db()
        li2.refresh_from_db()
        assert li1.purchase_status == "unordered"
        assert li2.purchase_status == "unordered"

    def test_daily_review_upload_rejects_damaged_exchange_selection(self, auth_client):
        """UploadDailyReviewView rejects a '파손/교환' 선택 cell — the row
        is reported explicitly in `errors` (not silently absorbed into an
        undifferentiated skipped_count), purchase_status is left unchanged,
        and no LineItemNote is created."""
        sku = "SKU-DEX-BLOCK-3"
        order = _make_order(1100072, name="#8001")
        li = _make_line_item(order, 1, sku, quantity=5, purchase_status="unordered")

        file_bytes = _make_daily_review_excel([
            {"order_name": "#8001", "isbn": sku, "selected": "파손/교환", "note": "파손 확인"},
        ])
        resp = auth_client.post(
            UPLOAD_DAILY_URL, data={"file": _file_obj(file_bytes)}, format="multipart"
        )
        assert resp.status_code == 201
        assert resp.data["skipped_count"] == 1
        assert resp.data["errors"] == [
            {"name": "#8001", "sku": sku, "reason": "damaged_exchange_requires_dedicated_page"}
        ]

        li.refresh_from_db()
        assert li.purchase_status == "unordered"
        assert li.damaged_quantity == 0
        assert li.notes.count() == 0
