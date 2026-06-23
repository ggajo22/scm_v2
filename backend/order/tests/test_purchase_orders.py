"""
TDD tests for SPEC-PURCHASE-ORDER-001 M2~M7 purchase order API endpoints.

Coverage targets:
  SC-PO-001  unordered aggregation + auto_distributor
  SC-PO-002  generate-order-file normal (Content-Type Excel)
  SC-PO-003  generate-order-file with unknown SKUs
  SC-PO-004  upload vendor file
  SC-PO-005  auto_select_distributor logic
  SC-PO-006  confirm order normal
  SC-PO-007  confirm order double-link prevention → 409
  SC-PO-008  vendor rules duplicate → 409
  SC-PO-009  vendor rules delete
  SC-PO-014  unauthenticated → 401
  SC-PO-015  invalid file format → 400
  EC-PO-001  empty skus → 400
  EC-PO-002  no unordered items → 200 empty
  EC-PO-003  only bookseen uploaded → kyobo fields null
  EC-PO-004  delete non-existent rule → 404
  EC-PO-005  invalid distributor → 400
  EC-PO-006  missing Excel columns → 422
"""

import io
from decimal import Decimal
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from order.models import (
    DistributorVendorRule,
    LineItem,
    Order,
    PurchaseOrder,
    Refund,
    VendorComparison,
)

User = get_user_model()

UNORDERED_URL = "/api/purchase-orders/unordered/"
GENERATE_URL = "/api/purchase-orders/generate-order-file/"
UPLOAD_URL = "/api/purchase-orders/upload-vendor-file/"
COMPARISON_URL = "/api/purchase-orders/comparison/"
CONFIRM_URL = "/api/purchase-orders/confirm/"
RULES_URL = "/api/purchase-orders/vendor-rules/"
PO_LIST_URL = "/api/purchase-orders/"
LINE_ITEM_STATUS_URL = "/api/purchase-orders/line-items/{pk}/status/"
BULK_STATUS_URL = "/api/purchase-orders/line-items/bulk-status/"

EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def user(db):
    return User.objects.create_user(username="po_test_user", password="testpass123")


@pytest.fixture
def auth_client(user):
    client = APIClient()
    refresh = RefreshToken.for_user(user)
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {refresh.access_token}")
    return client


@pytest.fixture
def anon_client():
    return APIClient()


def _make_order(
    shopify_order_id: int = 90001, store_type: str = "gimssine", name: str | None = None
) -> Order:
    return Order.objects.create(
        shopify_order_id=shopify_order_id, store_type=store_type, name=name
    )


def _make_line_item(
    order: Order,
    shopify_line_item_id: int = 1,
    sku: str = "9788901234567",
    title: str = "테스트 도서",
    vendor: str = "처음교육",
    quantity: int = 2,
) -> LineItem:
    return LineItem.objects.create(
        order=order,
        shopify_line_item_id=shopify_line_item_id,
        sku=sku,
        title=title,
        vendor=vendor,
        quantity=quantity,
    )


def _make_excel(rows: list[list], with_header: bool = True) -> bytes:
    """Build a generic .xlsx file (bookseen-generic format: header row + data)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    if with_header:
        ws.append(["ISBN", "재고여부", "단가"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_kyobo_excel(rows: list[dict]) -> bytes:
    """Build a Kyobo-format .xlsx file.

    Format: row 0 = '엑셀주문' title, row 1 = headers, row 2+ = data.
    Each dict in rows must have: isbn, status, available ('Y'/'N'),
    returnable ('Y'/'N'), qty, total_price, stock.
    출고가 = total_price / qty
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["엑셀주문"] + [None] * 17)
    ws.append([
        "순서", "ISBN", "도서명", "분야", "상품\n상태", "출고\n여부",
        "반품\n가능여부", "저자", "출판사", "출판년월", "정가",
        "주문\n수량", "출고율", "정가합", "출고가합", "보유\n재고", None, None,
    ])
    # col: 0=순서 1=ISBN 2=도서명 3=분야 4=상품상태 5=출고여부 6=반품가능여부
    #      7=저자 8=출판사 9=출판년월 10=정가 11=주문수량 12=출고율 13=정가합
    #      14=출고가합 15=보유재고
    for i, r in enumerate(rows, start=1):
        qty = float(r.get("qty", 1))
        total_price = float(r.get("total_price", 0))
        ws.append([
            float(i),                  # 0: 순서
            r["isbn"],                 # 1: ISBN
            r.get("title", ""),        # 2: 도서명
            r.get("category", ""),     # 3: 분야
            r.get("status", "정상"),   # 4: 상품상태
            r.get("available", "Y"),   # 5: 출고여부
            r.get("returnable", "Y"),  # 6: 반품가능여부
            r.get("author", ""),       # 7: 저자
            r.get("publisher", ""),    # 8: 출판사
            r.get("pub_date", ""),     # 9: 출판년월
            r.get("list_price", 0),    # 10: 정가
            qty,                       # 11: 주문수량
            r.get("rate", "70%"),      # 12: 출고율
            r.get("list_total", 0),    # 13: 정가합
            total_price,               # 14: 출고가합
            float(r.get("stock", 0)),  # 15: 보유재고
            None, None,
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# SC-PO-014: Unauthenticated requests → 401
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUnauthorizedAccess:
    """SC-PO-014: All endpoints require JWT authentication."""

    def test_unordered_requires_auth(self, anon_client):
        res = anon_client.get(UNORDERED_URL)
        assert res.status_code == 401

    def test_generate_requires_auth(self, anon_client):
        res = anon_client.post(GENERATE_URL, data={"distributor": "bookseen", "skus": []}, format="json")
        assert res.status_code == 401

    def test_upload_requires_auth(self, anon_client):
        res = anon_client.post(UPLOAD_URL)
        assert res.status_code == 401

    def test_comparison_requires_auth(self, anon_client):
        res = anon_client.get(COMPARISON_URL)
        assert res.status_code == 401

    def test_confirm_requires_auth(self, anon_client):
        res = anon_client.post(CONFIRM_URL, data={"items": []}, format="json")
        assert res.status_code == 401

    def test_rules_list_requires_auth(self, anon_client):
        res = anon_client.get(RULES_URL)
        assert res.status_code == 401

    def test_po_list_requires_auth(self, anon_client):
        res = anon_client.get(PO_LIST_URL)
        assert res.status_code == 401


# ---------------------------------------------------------------------------
# M2: GET /api/purchase-orders/unordered/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUnorderedItemsView:
    """SC-PO-001, EC-PO-002"""

    def test_returns_empty_when_no_line_items(self, auth_client):
        """EC-PO-002: No unordered items → 200 with empty results."""
        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        assert res.data["count"] == 0
        assert res.data["results"] == []

    def test_returns_individual_line_items(self, auth_client):
        """SC-PO-001: Each LineItem is returned as a separate row (no SKU aggregation)."""
        order1 = _make_order(shopify_order_id=91001, name="#1001")
        order2 = _make_order(shopify_order_id=91002, name="#1002")
        _make_line_item(order1, shopify_line_item_id=1, sku="SKU-A", quantity=3)
        _make_line_item(order2, shopify_line_item_id=2, sku="SKU-A", quantity=2)

        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        assert res.data["count"] == 2
        skus = [r["sku"] for r in res.data["results"]]
        assert skus.count("SKU-A") == 2
        quantities = {r["order_name"]: r["quantity"] for r in res.data["results"]}
        assert quantities["#1001"] == 3
        assert quantities["#1002"] == 2

    def test_excludes_line_items_linked_to_purchase_order(self, auth_client):
        """SC-PO-001: LineItems already linked to a PurchaseOrder are excluded."""
        order = _make_order(shopify_order_id=91003)
        li = _make_line_item(order, shopify_line_item_id=1, sku="SKU-B")
        po = PurchaseOrder.objects.create(
            sku="SKU-B", title="Book B", distributor="bookseen", quantity=2
        )
        po.line_items.add(li)

        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        skus = [r["sku"] for r in res.data["results"]]
        assert "SKU-B" not in skus

    def test_auto_distributor_from_vendor_rule(self, auth_client):
        """SC-PO-001: auto_distributor comes from DistributorVendorRule.publisher_name = vendor."""
        DistributorVendorRule.objects.create(
            publisher_name="처음교육", distributor="choeumgoyuk"
        )
        order = _make_order(shopify_order_id=91004)
        _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-C", vendor="처음교육"
        )

        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        result = next(r for r in res.data["results"] if r["sku"] == "SKU-C")
        assert result["auto_distributor"] == "choeumgoyuk"

    def test_auto_distributor_null_when_no_rule(self, auth_client):
        """SC-PO-001: auto_distributor is null when no DistributorVendorRule exists."""
        order = _make_order(shopify_order_id=91005)
        _make_line_item(
            order, shopify_line_item_id=1, sku="SKU-D", vendor="알수없는출판사"
        )

        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        result = next(r for r in res.data["results"] if r["sku"] == "SKU-D")
        assert result["auto_distributor"] is None

    def test_excludes_line_items_with_null_sku(self, auth_client):
        """LineItems with null SKU should be excluded from aggregation."""
        order = _make_order(shopify_order_id=91006)
        LineItem.objects.create(
            order=order,
            shopify_line_item_id=99,
            sku=None,
            title="No SKU Item",
            quantity=1,
        )
        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        assert res.data["count"] == 0

    def test_partial_refund_reduces_quantity(self, auth_client):
        """Partial refund: qty=2, refund=1 → response shows quantity=1."""
        order = _make_order(shopify_order_id=91010)
        li = _make_line_item(order, shopify_line_item_id=500, sku="SKU-R1", quantity=2)
        _make_refund(order, shopify_line_item_id=500, quantity=1, shopify_refund_id=701)

        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        result = next((r for r in res.data["results"] if r["sku"] == "SKU-R1"), None)
        assert result is not None
        assert result["quantity"] == 1

    def test_fully_refunded_line_item_excluded(self, auth_client):
        """Full refund: qty=3, refund=3 → LineItem excluded from unordered list."""
        order = _make_order(shopify_order_id=91011)
        _make_line_item(order, shopify_line_item_id=501, sku="SKU-R2", quantity=3)
        _make_refund(order, shopify_line_item_id=501, quantity=3, shopify_refund_id=702)

        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        skus = [r["sku"] for r in res.data["results"]]
        assert "SKU-R2" not in skus

    def test_no_refund_shows_original_quantity(self, auth_client):
        """No refund: quantity returned unchanged."""
        order = _make_order(shopify_order_id=91012)
        _make_line_item(order, shopify_line_item_id=502, sku="SKU-R3", quantity=5)

        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        result = next((r for r in res.data["results"] if r["sku"] == "SKU-R3"), None)
        assert result is not None
        assert result["quantity"] == 5


# ---------------------------------------------------------------------------
# M3: POST /api/purchase-orders/generate-order-file/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestGenerateOrderFileView:
    """SC-PO-002, SC-PO-003, EC-PO-001, EC-PO-005"""

    def _setup_sku(self, sku: str = "9788901234567", title: str = "테스트 도서", quantity: int = 3):
        order = _make_order(shopify_order_id=92001)
        _make_line_item(
            order, shopify_line_item_id=1, sku=sku, title=title, quantity=quantity
        )

    def test_returns_excel_for_valid_skus(self, auth_client):
        """SC-PO-002: Valid SKUs → returns Excel binary with correct Content-Type."""
        self._setup_sku("9788901234567", "Great Gatsby", 3)
        payload = {"distributor": "bookseen", "skus": ["9788901234567"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 200
        assert EXCEL_CONTENT_TYPE in res["Content-Type"]
        # Verify content is a valid xlsx
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "ISBN"
        assert rows[1][0] == "9788901234567"

    def test_filename_contains_distributor_and_date(self, auth_client):
        """SC-PO-002: Content-Disposition includes distributor name and date."""
        self._setup_sku()
        payload = {"distributor": "kyobo", "skus": ["9788901234567"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 200
        disposition = res.get("Content-Disposition", "")
        assert "kyobo" in disposition
        assert ".xlsx" in disposition

    def test_empty_skus_returns_400(self, auth_client):
        """EC-PO-001: skus=[] → HTTP 400."""
        payload = {"distributor": "bookseen", "skus": []}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 400

    def test_all_unknown_skus_returns_warning_json(self, auth_client):
        """SC-PO-003: All SKUs unknown → JSON warning (no Excel)."""
        payload = {"distributor": "bookseen", "skus": ["0000000000000"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 200
        assert "unknown_skus" in res.data
        assert "0000000000000" in res.data["unknown_skus"]
        assert EXCEL_CONTENT_TYPE not in res.get("Content-Type", "")

    def test_mixed_skus_returns_warning_json(self, auth_client):
        """SC-PO-003: Some unknown SKUs → JSON warning with unknown_skus list."""
        self._setup_sku("9788901234567", "Valid Book", 2)
        payload = {
            "distributor": "bookseen",
            "skus": ["9788901234567", "0000000000000"],
        }
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 200
        assert "unknown_skus" in res.data
        assert "0000000000000" in res.data["unknown_skus"]

    def test_invalid_distributor_returns_400(self, auth_client):
        """EC-PO-005: Invalid distributor value → HTTP 400."""
        payload = {"distributor": "invalid_dist", "skus": ["9788901234567"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 400

    def test_missing_distributor_returns_400(self, auth_client):
        """distributor field required."""
        payload = {"skus": ["9788901234567"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 400

    def test_partial_refund_reduces_excel_quantity(self, auth_client):
        """Excel quantity reflects net (original − refund): qty=2, refund=1 → Excel shows 1."""
        order = _make_order(shopify_order_id=92010)
        li = _make_line_item(order, shopify_line_item_id=600, sku="9788901299991", quantity=2)
        _make_refund(order, shopify_line_item_id=600, quantity=1, shopify_refund_id=801)
        payload = {"distributor": "bookseen", "skus": ["9788901299991"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 200
        assert EXCEL_CONTENT_TYPE in res["Content-Type"]
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        data_row = next((r for r in rows if r[0] == "9788901299991"), None)
        assert data_row is not None
        assert data_row[1] == 1  # quantity column = net 1

    def test_fully_refunded_sku_treated_as_unknown(self, auth_client):
        """All quantity refunded → SKU appears in unknown_skus (not in Excel)."""
        order = _make_order(shopify_order_id=92011)
        _make_line_item(order, shopify_line_item_id=601, sku="9788901299992", quantity=3)
        _make_refund(order, shopify_line_item_id=601, quantity=3, shopify_refund_id=802)
        payload = {"distributor": "bookseen", "skus": ["9788901299992"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 200
        assert "unknown_skus" in res.data
        assert "9788901299992" in res.data["unknown_skus"]


# ---------------------------------------------------------------------------
# M4a: POST /api/purchase-orders/upload-vendor-file/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadVendorFileView:
    """SC-PO-004, SC-PO-015, EC-PO-003, EC-PO-006"""

    def test_upload_bookseen_file(self, auth_client):
        """SC-PO-004: Upload bookseen Excel → VendorComparison created/updated."""
        excel_bytes = _make_excel([
            ["9788901234567", True, 12000],
            ["9788901234568", False, 0],
        ])
        file_obj = io.BytesIO(excel_bytes)
        file_obj.name = "bookseen.xlsx"
        res = auth_client.post(
            UPLOAD_URL,
            data={"distributor": "bookseen", "file": file_obj},
            format="multipart",
        )
        assert res.status_code == 200
        assert res.data["parsed_count"] == 2
        assert res.data["distributor"] == "bookseen"
        assert len(res.data["comparisons"]) == 2

        vc = VendorComparison.objects.get(sku="9788901234567")
        assert vc.bookseen_available is True
        assert vc.bookseen_price == Decimal("12000")

    def test_upload_kyobo_file(self, auth_client):
        """SC-PO-004: Upload kyobo Excel (교보 format) → VendorComparison kyobo fields updated."""
        excel_bytes = _make_kyobo_excel([{
            "isbn": "9788901234567", "qty": 2, "total_price": 23000.0,
            "stock": 10, "available": "Y", "returnable": "Y", "status": "정상",
            "publisher": "테스트출판사",
        }])
        file_obj = io.BytesIO(excel_bytes)
        file_obj.name = "kyobo.xlsx"
        res = auth_client.post(
            UPLOAD_URL,
            data={"distributor": "kyobo", "file": file_obj},
            format="multipart",
        )
        assert res.status_code == 200
        vc = VendorComparison.objects.get(sku="9788901234567")
        assert vc.kyobo_available is True
        assert vc.kyobo_price == Decimal("11500")  # 23000 / 2
        assert vc.kyobo_stock == 10
        assert vc.kyobo_returnable is True
        assert vc.kyobo_status == "정상"
        assert vc.kyobo_publisher == "테스트출판사"
        assert vc.kyobo_ordered_qty == 2
        assert vc.kyobo_total_price == Decimal("23000")

    def test_only_bookseen_uploaded_kyobo_fields_null(self, auth_client):
        """EC-PO-003: Only bookseen uploaded → kyobo fields remain null."""
        excel_bytes = _make_excel([["9788901234569", True, 10000]])
        file_obj = io.BytesIO(excel_bytes)
        file_obj.name = "bookseen.xlsx"
        auth_client.post(
            UPLOAD_URL,
            data={"distributor": "bookseen", "file": file_obj},
            format="multipart",
        )
        vc = VendorComparison.objects.get(sku="9788901234569")
        assert vc.kyobo_available is None
        assert vc.kyobo_price is None

    def test_invalid_file_format_returns_400(self, auth_client):
        """SC-PO-015: Non-.xlsx/.xls file → HTTP 400."""
        fake_csv = io.BytesIO(b"isbn,available,price\n9788901234567,true,10000")
        fake_csv.name = "data.csv"
        res = auth_client.post(
            UPLOAD_URL,
            data={"distributor": "bookseen", "file": fake_csv},
            format="multipart",
        )
        assert res.status_code == 400

    def test_upsert_updates_existing_record(self, auth_client):
        """SC-PO-004: Re-upload updates existing VendorComparison record."""
        VendorComparison.objects.create(
            sku="9788901234567", bookseen_available=False, bookseen_price=Decimal("9000")
        )
        excel_bytes = _make_excel([["9788901234567", True, 12000]])
        file_obj = io.BytesIO(excel_bytes)
        file_obj.name = "bookseen.xlsx"
        auth_client.post(
            UPLOAD_URL,
            data={"distributor": "bookseen", "file": file_obj},
            format="multipart",
        )
        vc = VendorComparison.objects.get(sku="9788901234567")
        assert vc.bookseen_available is True
        assert vc.bookseen_price == Decimal("12000")
        assert VendorComparison.objects.filter(sku="9788901234567").count() == 1


# ---------------------------------------------------------------------------
# SC-PO-005: auto_select_distributor logic
# ---------------------------------------------------------------------------


class TestAutoSelectDistributor:
    """SC-PO-005: Distributor auto-selection with new SPEC-AUTO-DIST-001 API.

    The function now accepts a VendorComparison instance and returns a dict.
    Detailed logic tests are in test_auto_dist.py.
    """

    def _vc(self, **kwargs) -> VendorComparison:
        defaults = {
            "sku": "TEST",
            "bookseen_stock": kwargs.pop("bookseen_stock", 10),
            "kyobo_stock": kwargs.pop("kyobo_stock", 10),
        }
        defaults.update(kwargs)
        return VendorComparison(**defaults)

    def test_both_available_lower_price_wins(self):
        from order.excel_utils import auto_select_distributor

        vc = self._vc(
            bookseen_price=Decimal("10000"),
            kyobo_price=Decimal("11000"),
        )
        result = auto_select_distributor(vc=vc, total_qty=5)
        assert result["selected_distributor"] == "bookseen"

    def test_both_available_kyobo_cheaper(self):
        from order.excel_utils import auto_select_distributor

        vc = self._vc(
            bookseen_price=Decimal("12000"),
            kyobo_price=Decimal("10000"),
        )
        result = auto_select_distributor(vc=vc, total_qty=5)
        assert result["selected_distributor"] == "kyobo"

    def test_only_bookseen_has_stock(self):
        from order.excel_utils import auto_select_distributor

        vc = self._vc(
            bookseen_stock=10,
            kyobo_stock=0,
            bookseen_price=None,
            kyobo_price=Decimal("11000"),
        )
        result = auto_select_distributor(vc=vc, total_qty=5)
        assert result["selected_distributor"] == "bookseen"

    def test_no_stock_check_required(self):
        from order.excel_utils import auto_select_distributor

        vc = self._vc(
            bookseen_stock=0,
            kyobo_stock=0,
            bookseen_price=None,
            kyobo_price=None,
            bookseen_status="품절",
            kyobo_status="품절",
        )
        result = auto_select_distributor(vc=vc, total_qty=5)
        assert result["selected_distributor"] == "check_required"


# ---------------------------------------------------------------------------
# M4b: GET /api/purchase-orders/comparison/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestVendorComparisonView:
    """EC-PO-003: comparison list"""

    def test_returns_empty_list(self, auth_client):
        res = auth_client.get(COMPARISON_URL)
        assert res.status_code == 200
        assert res.data["count"] == 0

    def test_returns_comparison_records(self, auth_client):
        VendorComparison.objects.create(
            sku="9788901234567",
            bookseen_available=True,
            bookseen_price=Decimal("12000"),
        )
        res = auth_client.get(COMPARISON_URL)
        assert res.status_code == 200
        assert res.data["count"] == 1
        record = res.data["results"][0]
        assert record["sku"] == "9788901234567"
        assert record["bookseen_available"] is True


# ---------------------------------------------------------------------------
# M5: POST /api/purchase-orders/confirm/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestConfirmOrderView:
    """SC-PO-006, SC-PO-007"""

    def test_confirm_creates_purchase_orders(self, auth_client):
        """SC-PO-006: Confirm creates PurchaseOrder and links LineItems."""
        order = _make_order(shopify_order_id=93001)
        _make_line_item(
            order, shopify_line_item_id=1, sku="ISBN-001", title="Book One", quantity=3
        )
        payload = {
            "items": [
                {
                    "sku": "ISBN-001",
                    "distributor": "bookseen",
                    "quantity": 3,
                    "unit_price": "10500.00",
                }
            ]
        }
        res = auth_client.post(CONFIRM_URL, data=payload, format="json")
        assert res.status_code == 201
        assert res.data["created_count"] == 1
        assert len(res.data["purchase_order_ids"]) == 1

        po = PurchaseOrder.objects.get(pk=res.data["purchase_order_ids"][0])
        assert po.sku == "ISBN-001"
        assert po.distributor == "bookseen"
        assert po.line_items.count() >= 1

    def test_confirm_multiple_skus(self, auth_client):
        """SC-PO-006: Confirm with multiple SKUs creates multiple PurchaseOrders."""
        order = _make_order(shopify_order_id=93002)
        _make_line_item(order, shopify_line_item_id=1, sku="ISBN-A", quantity=2)
        _make_line_item(order, shopify_line_item_id=2, sku="ISBN-B", quantity=1)
        payload = {
            "items": [
                {"sku": "ISBN-A", "distributor": "bookseen", "quantity": 2, "unit_price": "10000.00"},
                {"sku": "ISBN-B", "distributor": "kyobo", "quantity": 1, "unit_price": "8000.00"},
            ]
        }
        res = auth_client.post(CONFIRM_URL, data=payload, format="json")
        assert res.status_code == 201
        assert res.data["created_count"] == 2

    def test_double_confirm_returns_409(self, auth_client):
        """SC-PO-007: Confirming already-linked LineItems → HTTP 409."""
        order = _make_order(shopify_order_id=93003)
        li = _make_line_item(
            order, shopify_line_item_id=1, sku="ISBN-DUPE", quantity=2
        )
        # First confirmation
        po = PurchaseOrder.objects.create(
            sku="ISBN-DUPE", title="Dupe Book", distributor="bookseen", quantity=2
        )
        po.line_items.add(li)

        payload = {
            "items": [{"sku": "ISBN-DUPE", "distributor": "bookseen", "quantity": 2, "unit_price": "9000.00"}]
        }
        res = auth_client.post(CONFIRM_URL, data=payload, format="json")
        assert res.status_code == 409

    def test_sku_with_no_unordered_items_returns_400(self, auth_client):
        """EC-PO-002 variant: SKU has no unordered LineItem → HTTP 400."""
        payload = {
            "items": [{"sku": "NONEXISTENT-SKU", "distributor": "bookseen", "quantity": 1, "unit_price": "9000.00"}]
        }
        res = auth_client.post(CONFIRM_URL, data=payload, format="json")
        assert res.status_code == 400

    def test_confirm_empty_items_returns_400(self, auth_client):
        """Empty items list → HTTP 400."""
        res = auth_client.post(CONFIRM_URL, data={"items": []}, format="json")
        assert res.status_code == 400


# ---------------------------------------------------------------------------
# M6: /api/purchase-orders/vendor-rules/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestDistributorVendorRuleViews:
    """SC-PO-008, SC-PO-009, EC-PO-004"""

    def test_list_rules_empty(self, auth_client):
        res = auth_client.get(RULES_URL)
        assert res.status_code == 200
        assert res.data["count"] == 0
        assert res.data["results"] == []

    def test_create_rule(self, auth_client):
        """POST creates a new DistributorVendorRule."""
        payload = {"publisher_name": "처음교육출판사", "distributor": "choeumgoyuk"}
        res = auth_client.post(RULES_URL, data=payload, format="json")
        assert res.status_code == 201
        assert DistributorVendorRule.objects.filter(publisher_name="처음교육출판사").exists()

    def test_duplicate_publisher_name_returns_409(self, auth_client):
        """SC-PO-008: Duplicate publisher_name → HTTP 409."""
        DistributorVendorRule.objects.create(
            publisher_name="아가페출판사", distributor="agape"
        )
        payload = {"publisher_name": "아가페출판사", "distributor": "choeumgoyuk"}
        res = auth_client.post(RULES_URL, data=payload, format="json")
        assert res.status_code == 409

    def test_delete_rule(self, auth_client):
        """SC-PO-009: DELETE removes the rule successfully."""
        rule = DistributorVendorRule.objects.create(
            publisher_name="삭제출판사", distributor="agape"
        )
        res = auth_client.delete(f"{RULES_URL}{rule.pk}/")
        assert res.status_code == 204
        assert not DistributorVendorRule.objects.filter(pk=rule.pk).exists()

    def test_delete_nonexistent_rule_returns_404(self, auth_client):
        """EC-PO-004: Deleting non-existent rule → HTTP 404."""
        res = auth_client.delete(f"{RULES_URL}99999/")
        assert res.status_code == 404

    def test_invalid_distributor_in_rule_returns_400(self, auth_client):
        """EC-PO-005: Invalid distributor in vendor rule → HTTP 400."""
        payload = {"publisher_name": "테스트출판사", "distributor": "invalid_dist"}
        res = auth_client.post(RULES_URL, data=payload, format="json")
        assert res.status_code == 400


# ---------------------------------------------------------------------------
# M7: GET /api/purchase-orders/
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestPurchaseOrderListView:
    """M7: List PurchaseOrders with filters and pagination."""

    def _create_po(self, sku: str, distributor: str = "bookseen", status: str = "pending") -> PurchaseOrder:
        return PurchaseOrder.objects.create(
            sku=sku, title=f"Book {sku}", distributor=distributor, quantity=1, status=status
        )

    def test_list_returns_paginated_results(self, auth_client):
        for i in range(3):
            self._create_po(f"SKU-LIST-{i}")
        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        assert res.data["count"] == 3

    def test_filter_by_distributor(self, auth_client):
        self._create_po("SKU-BS", distributor="bookseen")
        self._create_po("SKU-KY", distributor="kyobo")
        res = auth_client.get(PO_LIST_URL, {"distributor": "bookseen"})
        assert res.status_code == 200
        assert res.data["count"] == 1
        assert res.data["results"][0]["distributor"] == "bookseen"

    def test_filter_by_status(self, auth_client):
        self._create_po("SKU-PEND", status="pending")
        self._create_po("SKU-CONF", status="confirmed")
        res = auth_client.get(PO_LIST_URL, {"status": "confirmed"})
        assert res.status_code == 200
        assert res.data["count"] == 1
        assert res.data["results"][0]["status"] == "confirmed"

    def test_ordered_by_created_at_desc(self, auth_client):
        po1 = self._create_po("SKU-FIRST")
        po2 = self._create_po("SKU-SECOND")
        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        ids = [r["id"] for r in res.data["results"]]
        assert ids.index(po2.pk) < ids.index(po1.pk)

    def test_pagination_page_size_50(self, auth_client):
        for i in range(55):
            self._create_po(f"SKU-PAG-{i}")
        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        assert res.data["count"] == 55
        assert len(res.data["results"]) == 50
        assert res.data["next"] is not None

    def test_filter_by_date_range(self, auth_client):
        """date_from and date_to filter on created_at."""
        self._create_po("SKU-DATE")
        res = auth_client.get(PO_LIST_URL, {"date_from": "2020-01-01", "date_to": "2099-12-31"})
        assert res.status_code == 200
        assert res.data["count"] >= 1


# ---------------------------------------------------------------------------
# excel_utils unit tests
# ---------------------------------------------------------------------------


class TestGenerateOrderExcel:
    """Unit tests for excel_utils.generate_order_excel."""

    def test_bookseen_column_format(self):
        from order.excel_utils import generate_order_excel

        data = [{"sku": "9780001", "title": "Test Book", "total_quantity": 5}]
        result = generate_order_excel(data, "bookseen")
        assert isinstance(result, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("ISBN", "주문수량", "도서명", "출판사", "저자", "정가")
        assert rows[1][0] == "9780001"
        assert rows[1][1] == 5

    def test_kyobo_column_format(self):
        from order.excel_utils import generate_order_excel

        data = [{"sku": "9780001", "title": "Test Book", "total_quantity": 3}]
        result = generate_order_excel(data, "kyobo")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("ISBN", "수량")
        assert rows[1][0] == "9780001"
        assert rows[1][1] == 3

    def test_empty_data_still_has_header(self):
        from order.excel_utils import generate_order_excel

        result = generate_order_excel([], "kyobo")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("ISBN", "수량")


class TestParseVendorExcel:
    """Unit tests for excel_utils.parse_vendor_excel."""

    def test_parses_standard_columns(self):
        from order.excel_utils import parse_vendor_excel

        excel_bytes = _make_excel([["9780001", True, 12000]])
        results = parse_vendor_excel(excel_bytes, "bookseen")
        assert len(results) == 1
        assert results[0]["sku"] == "9780001"
        assert results[0]["available"] is True
        assert results[0]["price"] == 12000.0

    def test_skips_empty_sku_rows(self):
        from order.excel_utils import parse_vendor_excel

        excel_bytes = _make_excel([["", True, 5000], ["9780002", False, 0]])
        results = parse_vendor_excel(excel_bytes, "bookseen")
        assert len(results) == 1
        assert results[0]["sku"] == "9780002"

    def test_handles_empty_file(self):
        from order.excel_utils import parse_vendor_excel

        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="Empty file"):
            parse_vendor_excel(buf.getvalue(), "bookseen")

    def test_parses_bookseen_xls_format(self):
        """Bookseen .xls: row 0 = timestamp, row 1 = headers, row 2+ = data.
        ISBN at col 14, 출고가 at col 6, 재고량 at col 7."""
        from order.excel_utils import _XLS_MAGIC, parse_vendor_excel

        rows = [[""] * 16, [""] * 16, [0.0] * 16]
        rows[0][0] = "조회기간 10건"
        rows[2][14] = 8809264180921.0  # ISBN as float (xlrd numeric cell)
        rows[2][6] = 15260.0           # 출고가
        rows[2][7] = 3.0               # 재고량 > 0 → available

        mock_sheet = MagicMock()
        mock_sheet.nrows = 3
        mock_sheet.row_values.side_effect = lambda n: rows[n]

        mock_wb = MagicMock()
        mock_wb.sheet_by_index.return_value = mock_sheet

        xls_magic_bytes = _XLS_MAGIC + b"\x00" * 100

        with patch("order.excel_utils.xlrd.open_workbook", return_value=mock_wb):
            results = parse_vendor_excel(xls_magic_bytes, "bookseen")

        assert len(results) == 1
        assert results[0]["sku"] == "8809264180921"
        assert results[0]["available"] is True
        assert results[0]["price"] == 15260.0

    def test_bookseen_xls_zero_stock_is_unavailable(self):
        """재고량 == 0 → available = False."""
        from order.excel_utils import _XLS_MAGIC, parse_vendor_excel

        rows = [[""] * 16, [""] * 16, [0.0] * 16]
        rows[2][14] = 9780000000000.0
        rows[2][6] = 12000.0
        rows[2][7] = 0.0

        mock_sheet = MagicMock()
        mock_sheet.nrows = 3
        mock_sheet.row_values.side_effect = lambda n: rows[n]

        mock_wb = MagicMock()
        mock_wb.sheet_by_index.return_value = mock_sheet

        with patch("order.excel_utils.xlrd.open_workbook", return_value=mock_wb):
            results = parse_vendor_excel(_XLS_MAGIC + b"\x00" * 100, "bookseen")

        assert results[0]["available"] is False


# ---------------------------------------------------------------------------
# SPEC-PURCHASE-ORDER-003: Refund exclusion & net_quantity field
# ---------------------------------------------------------------------------


def _make_refund(
    order: Order,
    shopify_line_item_id: int,
    quantity: int,
    shopify_refund_id: int = 1,
) -> Refund:
    return Refund.objects.create(
        order=order,
        shopify_refund_id=shopify_refund_id,
        line_item_id=shopify_line_item_id,
        quantity=quantity,
    )


@pytest.mark.django_db
class TestPurchaseOrderListRefundExclusion:
    """SPEC-PURCHASE-ORDER-003: PO list excludes fully-refunded POs and exposes net_quantity."""

    def test_ac01_fully_refunded_po_excluded(self, auth_client):
        """AC-01: LineItem qty=5, Refund qty=5 → PO excluded from response."""
        order = _make_order(shopify_order_id=80001)
        li = _make_line_item(order, shopify_line_item_id=10, quantity=5)
        po = PurchaseOrder.objects.create(sku=li.sku, title=li.title, distributor="bookseen", quantity=5)
        po.line_items.add(li)
        _make_refund(order, shopify_line_item_id=10, quantity=5, shopify_refund_id=101)

        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        ids = [item["id"] for item in res.data["results"]]
        assert po.id not in ids

    def test_ac02_partially_refunded_po_shown_with_net_quantity(self, auth_client):
        """AC-02: LineItem qty=5, Refund qty=3 → shown, net_quantity=2."""
        order = _make_order(shopify_order_id=80002)
        li = _make_line_item(order, shopify_line_item_id=20, quantity=5)
        po = PurchaseOrder.objects.create(sku=li.sku, title=li.title, distributor="bookseen", quantity=5)
        po.line_items.add(li)
        _make_refund(order, shopify_line_item_id=20, quantity=3, shopify_refund_id=201)

        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        po_data = next(item for item in res.data["results"] if item["id"] == po.id)
        assert po_data["net_quantity"] == 2

    def test_ac03_po_with_no_line_items_shown_with_original_quantity(self, auth_client):
        """AC-03: PO with no LineItems → shown, net_quantity = PO.quantity."""
        po = PurchaseOrder.objects.create(sku="9780000000001", title="노라인아이템", distributor="bookseen", quantity=7)

        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        po_data = next(item for item in res.data["results"] if item["id"] == po.id)
        assert po_data["net_quantity"] == 7

    def test_ac04_multiple_line_items_all_fully_refunded_excluded(self, auth_client):
        """AC-04: 2 LineItems (qty=3, qty=2) both fully refunded → PO excluded."""
        order = _make_order(shopify_order_id=80004)
        li1 = _make_line_item(order, shopify_line_item_id=41, sku="SKU-AC04-A", quantity=3)
        li2 = _make_line_item(order, shopify_line_item_id=42, sku="SKU-AC04-B", quantity=2)
        po = PurchaseOrder.objects.create(sku="SKU-AC04-A", title="멀티라인 전환불", distributor="bookseen", quantity=5)
        po.line_items.add(li1, li2)
        _make_refund(order, shopify_line_item_id=41, quantity=3, shopify_refund_id=401)
        _make_refund(order, shopify_line_item_id=42, quantity=2, shopify_refund_id=402)

        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        ids = [item["id"] for item in res.data["results"]]
        assert po.id not in ids

    def test_ac05_multiple_line_items_partially_refunded(self, auth_client):
        """AC-05: qty=3,qty=2; Refund=3,1 → shown, net_quantity=1 (2-1=1)."""
        order = _make_order(shopify_order_id=80005)
        li1 = _make_line_item(order, shopify_line_item_id=51, sku="SKU-AC05-A", quantity=3)
        li2 = _make_line_item(order, shopify_line_item_id=52, sku="SKU-AC05-B", quantity=2)
        po = PurchaseOrder.objects.create(sku="SKU-AC05-A", title="멀티라인 부분환불", distributor="bookseen", quantity=5)
        po.line_items.add(li1, li2)
        _make_refund(order, shopify_line_item_id=51, quantity=3, shopify_refund_id=501)
        _make_refund(order, shopify_line_item_id=52, quantity=1, shopify_refund_id=502)

        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        po_data = next(item for item in res.data["results"] if item["id"] == po.id)
        assert po_data["net_quantity"] == 1

    def test_ac06_refund_exceeds_quantity_treated_as_fully_refunded(self, auth_client):
        """AC-06: LineItem qty=5, Refund qty=6 → treated fully refunded, PO excluded."""
        order = _make_order(shopify_order_id=80006)
        li = _make_line_item(order, shopify_line_item_id=60, quantity=5)
        po = PurchaseOrder.objects.create(sku=li.sku, title=li.title, distributor="bookseen", quantity=5)
        po.line_items.add(li)
        _make_refund(order, shopify_line_item_id=60, quantity=6, shopify_refund_id=601)

        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        ids = [item["id"] for item in res.data["results"]]
        assert po.id not in ids

    def test_ac07_filter_with_refund_exclusion(self, auth_client):
        """AC-07: distributor filter + fully refunded PO → fully refunded PO excluded."""
        order = _make_order(shopify_order_id=80007)
        li = _make_line_item(order, shopify_line_item_id=70, quantity=3)
        po_fully_refunded = PurchaseOrder.objects.create(
            sku=li.sku, title=li.title, distributor="bookseen", quantity=3
        )
        po_fully_refunded.line_items.add(li)
        _make_refund(order, shopify_line_item_id=70, quantity=3, shopify_refund_id=701)

        po_normal = PurchaseOrder.objects.create(
            sku="SKU-AC07-NORMAL", title="일반 PO", distributor="bookseen", quantity=2
        )

        res = auth_client.get(PO_LIST_URL, {"distributor": "bookseen"})
        assert res.status_code == 200
        ids = [item["id"] for item in res.data["results"]]
        assert po_fully_refunded.id not in ids
        assert po_normal.id in ids

    def test_ac08_po_with_line_item_no_refund_shown(self, auth_client):
        """AC-08: PO with LineItem but no Refund records → shown, net_quantity=LineItem.quantity."""
        order = _make_order(shopify_order_id=80008)
        li = _make_line_item(order, shopify_line_item_id=80, quantity=4)
        po = PurchaseOrder.objects.create(sku=li.sku, title=li.title, distributor="bookseen", quantity=4)
        po.line_items.add(li)

        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        po_data = next(item for item in res.data["results"] if item["id"] == po.id)
        assert po_data["net_quantity"] == 4

    def test_ac09_net_quantity_field_present_in_response(self, auth_client):
        """AC-09: net_quantity field present in API response for non-excluded POs."""
        po = PurchaseOrder.objects.create(sku="9780000000009", title="필드존재확인", distributor="bookseen", quantity=3)

        res = auth_client.get(PO_LIST_URL)
        assert res.status_code == 200
        assert res.data["count"] >= 1
        po_data = next(item for item in res.data["results"] if item["id"] == po.id)
        assert "net_quantity" in po_data


# ---------------------------------------------------------------------------
# SPEC-PURCHASE-ORDER-004: UnorderedItemsView purchase_status filter
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUnorderedItemsViewPurchaseStatusFilter:
    """REQ-PO4-003, REQ-PO4-004: Filter by purchase_status and expose field in response."""

    def test_on_hold_item_excluded_from_unordered(self, auth_client):
        """REQ-PO4-003: LineItem with purchase_status='on_hold' is excluded from unordered list."""
        order = _make_order(shopify_order_id=94001)
        LineItem.objects.create(
            order=order,
            shopify_line_item_id=1,
            sku="SKU-HOLD",
            title="On Hold Book",
            quantity=2,
            purchase_status="on_hold",
        )
        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        skus = [r["sku"] for r in res.data["results"]]
        assert "SKU-HOLD" not in skus

    def test_unordered_item_included(self, auth_client):
        """REQ-PO4-003: LineItem with purchase_status='unordered' appears in unordered list."""
        order = _make_order(shopify_order_id=94002)
        LineItem.objects.create(
            order=order,
            shopify_line_item_id=1,
            sku="SKU-UNORDERED",
            title="Unordered Book",
            quantity=2,
            purchase_status="unordered",
        )
        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        skus = [r["sku"] for r in res.data["results"]]
        assert "SKU-UNORDERED" in skus

    def test_response_includes_purchase_status_field(self, auth_client):
        """REQ-PO4-004: Response item includes 'purchase_status' key."""
        order = _make_order(shopify_order_id=94003)
        _make_line_item(order, shopify_line_item_id=1, sku="SKU-FIELD")
        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        assert res.data["count"] >= 1
        result = next(r for r in res.data["results"] if r["sku"] == "SKU-FIELD")
        assert "purchase_status" in result

    def test_po_linked_unordered_item_excluded(self, auth_client):
        """M2M link takes priority: purchase_status='unordered' but linked to PO → excluded."""
        order = _make_order(shopify_order_id=94004)
        li = LineItem.objects.create(
            order=order,
            shopify_line_item_id=1,
            sku="SKU-PO-LINKED",
            title="PO Linked Book",
            quantity=2,
            purchase_status="unordered",
        )
        po = PurchaseOrder.objects.create(
            sku="SKU-PO-LINKED", title="Linked Book PO", distributor="bookseen", quantity=2
        )
        po.line_items.add(li)
        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        skus = [r["sku"] for r in res.data["results"]]
        assert "SKU-PO-LINKED" not in skus

    def test_all_non_unordered_statuses_excluded(self, auth_client):
        """All non-unordered statuses are excluded from the unordered list."""
        non_unordered = ["on_hold", "order_cancelled", "other_publisher", "cs_required", "in_stock"]
        order = _make_order(shopify_order_id=94005)
        for i, status in enumerate(non_unordered):
            LineItem.objects.create(
                order=order,
                shopify_line_item_id=10 + i,
                sku=f"SKU-NOLIST-{i}",
                title=f"Book {status}",
                quantity=1,
                purchase_status=status,
            )
        res = auth_client.get(UNORDERED_URL)
        assert res.status_code == 200
        result_skus = [r["sku"] for r in res.data["results"]]
        for i in range(len(non_unordered)):
            assert f"SKU-NOLIST-{i}" not in result_skus


# ---------------------------------------------------------------------------
# SPEC-PURCHASE-ORDER-004: LineItemStatusUpdateView (single PATCH)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestLineItemStatusUpdateView:
    """REQ-PO4-005, REQ-PO4-007: Single line item status update."""

    def _make_li(self, shopify_order_id: int = 95001, shopify_line_item_id: int = 1) -> LineItem:
        order = _make_order(shopify_order_id=shopify_order_id)
        return _make_line_item(order, shopify_line_item_id=shopify_line_item_id)

    def test_patch_valid_status_returns_200(self, auth_client):
        """REQ-PO4-005: PATCH with valid status updates the LineItem and returns 200."""
        li = self._make_li()
        url = LINE_ITEM_STATUS_URL.format(pk=li.pk)
        res = auth_client.patch(url, data={"purchase_status": "on_hold"}, format="json")
        assert res.status_code == 200
        li.refresh_from_db()
        assert li.purchase_status == "on_hold"

    def test_patch_response_contains_updated_status(self, auth_client):
        """Response body includes the updated purchase_status."""
        li = self._make_li(shopify_order_id=95002, shopify_line_item_id=2)
        url = LINE_ITEM_STATUS_URL.format(pk=li.pk)
        res = auth_client.patch(url, data={"purchase_status": "in_stock"}, format="json")
        assert res.status_code == 200
        assert res.data["purchase_status"] == "in_stock"
        assert res.data["id"] == li.pk

    def test_patch_invalid_status_returns_400(self, auth_client):
        """REQ-PO4-007: PATCH with invalid status value returns 400."""
        li = self._make_li(shopify_order_id=95003, shopify_line_item_id=3)
        url = LINE_ITEM_STATUS_URL.format(pk=li.pk)
        res = auth_client.patch(url, data={"purchase_status": "invalid_value"}, format="json")
        assert res.status_code == 400

    def test_patch_nonexistent_id_returns_404(self, auth_client):
        """REQ-PO4-007: PATCH on non-existent LineItem pk returns 404."""
        url = LINE_ITEM_STATUS_URL.format(pk=99999)
        res = auth_client.patch(url, data={"purchase_status": "on_hold"}, format="json")
        assert res.status_code == 404

    def test_patch_unauthenticated_returns_401(self, anon_client):
        """REQ-PO4-007: PATCH without auth returns 401."""
        url = LINE_ITEM_STATUS_URL.format(pk=1)
        res = anon_client.patch(url, data={"purchase_status": "on_hold"}, format="json")
        assert res.status_code == 401

    def test_patch_all_six_choices(self, auth_client):
        """All 6 valid purchase_status choices can be set via PATCH."""
        valid_choices = [
            "unordered", "on_hold", "order_cancelled",
            "other_publisher", "cs_required", "in_stock",
        ]
        order = _make_order(shopify_order_id=95010)
        for i, choice in enumerate(valid_choices):
            li = _make_line_item(order, shopify_line_item_id=100 + i, sku=f"SKU-ALL-{i}")
            url = LINE_ITEM_STATUS_URL.format(pk=li.pk)
            res = auth_client.patch(url, data={"purchase_status": choice}, format="json")
            assert res.status_code == 200, f"Failed for choice: {choice}"
            li.refresh_from_db()
            assert li.purchase_status == choice


# ---------------------------------------------------------------------------
# SPEC-PURCHASE-ORDER-004: LineItemBulkStatusUpdateView (bulk PATCH)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestLineItemBulkStatusUpdateView:
    """REQ-PO4-006, REQ-PO4-007: Bulk line item status update."""

    def _make_lis(self, count: int, shopify_order_id: int = 96001) -> list:
        order = _make_order(shopify_order_id=shopify_order_id)
        items = []
        for i in range(count):
            li = _make_line_item(order, shopify_line_item_id=i + 1, sku=f"SKU-BULK-{shopify_order_id}-{i}")
            items.append(li)
        return items

    def test_bulk_update_success(self, auth_client):
        """REQ-PO4-006: Bulk PATCH updates all given IDs and returns updated_count."""
        lis = self._make_lis(3)
        ids = [li.pk for li in lis]
        res = auth_client.patch(
            BULK_STATUS_URL,
            data={"ids": ids, "purchase_status": "on_hold"},
            format="json",
        )
        assert res.status_code == 200
        assert res.data["updated_count"] == 3
        assert res.data["missing_ids"] == []
        for li in lis:
            li.refresh_from_db()
            assert li.purchase_status == "on_hold"

    def test_bulk_partial_success_with_missing_ids(self, auth_client):
        """REQ-PO4-006: Missing IDs are reported in missing_ids; existing ones are updated."""
        lis = self._make_lis(2, shopify_order_id=96002)
        ids = [li.pk for li in lis] + [99998, 99999]
        res = auth_client.patch(
            BULK_STATUS_URL,
            data={"ids": ids, "purchase_status": "cs_required"},
            format="json",
        )
        assert res.status_code == 200
        assert res.data["updated_count"] == 2
        assert 99998 in res.data["missing_ids"]
        assert 99999 in res.data["missing_ids"]

    def test_bulk_empty_ids_returns_400(self, auth_client):
        """REQ-PO4-007: Empty ids list returns 400."""
        res = auth_client.patch(
            BULK_STATUS_URL,
            data={"ids": [], "purchase_status": "on_hold"},
            format="json",
        )
        assert res.status_code == 400

    def test_bulk_invalid_status_returns_400(self, auth_client):
        """REQ-PO4-007: Invalid purchase_status returns 400."""
        lis = self._make_lis(1, shopify_order_id=96003)
        res = auth_client.patch(
            BULK_STATUS_URL,
            data={"ids": [lis[0].pk], "purchase_status": "invalid_choice"},
            format="json",
        )
        assert res.status_code == 400

    def test_bulk_unauthenticated_returns_401(self, anon_client):
        """REQ-PO4-007: PATCH without auth returns 401."""
        res = anon_client.patch(
            BULK_STATUS_URL,
            data={"ids": [1], "purchase_status": "on_hold"},
            format="json",
        )
        assert res.status_code == 401
