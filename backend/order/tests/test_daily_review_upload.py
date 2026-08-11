"""
TDD tests for SPEC-PURCHASE-ORDER-005.

Covers:
  REQ-PO5-001  _DISTRIBUTOR_LABEL_MAP includes warehouse codes
  REQ-PO5-002  _DAILY_REVIEW_HEADERS uses location columns (재고(한국)/재고(CA)/재고(NJ))
  REQ-PO5-003  generate_daily_review_excel uses korea_stock/ca_stock/nj_stock keys
  REQ-PO5-004  UploadDailyReviewView deducts WarehouseStock
  REQ-PO5-005  UploadDailyReviewView sets LineItem.purchase_status = "in_stock"
  REQ-PO5-006  UploadDailyReviewView does NOT create PurchaseOrder for warehouse rows
  REQ-PO5-007  Upload response includes confirmed_by_distributor
  REQ-PO5-008  VALID_DISTRIBUTORS includes warehouse codes
"""

import io
from decimal import Decimal

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken

from order.models import LineItem, Order, PurchaseOrder, WarehouseStock

User = get_user_model()

UPLOAD_DAILY_URL = "/api/purchase-orders/upload-daily-review/"
GENERATE_URL = "/api/purchase-orders/generate-order-file/"
DAILY_REVIEW_EXCEL_URL = "/api/purchase-orders/daily-review-excel/"

EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def user(db):
    return User.objects.create_user(username="dr_test_user", password="testpass123")


@pytest.fixture
def auth_client(user):
    client = APIClient()
    refresh = RefreshToken.for_user(user)
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {refresh.access_token}")
    return client


def _make_order(shopify_order_id: int = 80001, name: str = "#8001") -> Order:
    return Order.objects.create(
        shopify_order_id=shopify_order_id, store_type="gimssine", name=name
    )


def _make_line_item(
    order: Order,
    sku: str = "9788901234567",
    title: str = "테스트 도서",
    quantity: int = 2,
    shopify_line_item_id: int = 1,
) -> LineItem:
    return LineItem.objects.create(
        order=order,
        shopify_line_item_id=shopify_line_item_id,
        sku=sku,
        title=title,
        quantity=quantity,
    )


def _make_daily_review_excel(rows: list[dict]) -> bytes:
    """
    Build a Daily Review .xlsx file with standard headers.

    Each dict in rows must have: isbn, selected (Korean label), and optionally note.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "주문번호", "ISBN", "제목", "수량", "주문위치", "메모",
        "재고(한국)", "재고(CA)", "재고(NJ)",
        "북센 공급가", "교보 공급가",
        "북센 재고수량", "북센 재고상태",
        "교보 재고수량", "교보 재고상태",
        "공급가차이", "북센 입고예정", "북센 반품가능여부",
        "교보 출고여부", "교보 반품가능여부",
        "가격차이알림", "출판사", "선택근거", "선택",
    ])
    for row in rows:
        ws.append([
            # REQ-PO8-018: default matches _make_order()'s default name
            # ("#8001") so existing single-order tests that never pass
            # order_name still exercise the same LineItem/order pairing
            # they did before the (order_name, sku) fix.
            row.get("order_name", "#8001"),
            row.get("isbn", ""),
            row.get("title", ""),
            row.get("quantity", 1),
            "",
            row.get("note", ""),
            row.get("korea_stock", 0),
            row.get("ca_stock", 0),
            row.get("nj_stock", 0),
            row.get("bs_price", ""),
            "", "", "", "", "", "", "", "", "", "", "", "", "",
            row.get("selected", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_legacy_daily_review_excel(rows: list[dict]) -> bytes:
    """Build a Daily Review .xlsx file with the OLD headers (창고재고수량/창고재고위치)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "주문번호", "ISBN", "제목", "수량", "주문위치", "메모",
        "창고재고수량", "창고재고위치",
        "북센 공급가", "교보 공급가",
        "북센 재고수량", "북센 재고상태",
        "교보 재고수량", "교보 재고상태",
        "공급가차이", "북센 입고예정", "북센 반품가능여부",
        "교보 출고여부", "교보 반품가능여부",
        "가격차이알림", "출판사", "선택근거", "선택",
    ])
    for row in rows:
        ws.append([
            # REQ-PO8-018: same default-name rationale as _make_daily_review_excel.
            row.get("order_name", "#8001"),
            row.get("isbn", ""),
            row.get("title", ""),
            row.get("quantity", 1),
            "",
            row.get("note", ""),
            row.get("warehouse_qty", 0),
            row.get("warehouse_locations", ""),
            "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
            row.get("selected", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# REQ-PO5-001: _DISTRIBUTOR_LABEL_MAP includes warehouse codes
# ---------------------------------------------------------------------------


class TestDistributorLabelMapWarehouseCodes:
    """REQ-PO5-001: Warehouse display names are mapped to internal codes."""

    def test_label_map_includes_warehouse_korea(self):
        from order.excel_utils import _DISTRIBUTOR_LABEL_MAP

        assert "재고(한국)" in _DISTRIBUTOR_LABEL_MAP
        assert _DISTRIBUTOR_LABEL_MAP["재고(한국)"] == "warehouse_korea"

    def test_label_map_includes_warehouse_ca(self):
        from order.excel_utils import _DISTRIBUTOR_LABEL_MAP

        assert "재고(CA)" in _DISTRIBUTOR_LABEL_MAP
        assert _DISTRIBUTOR_LABEL_MAP["재고(CA)"] == "warehouse_ca"

    def test_label_map_includes_warehouse_nj(self):
        from order.excel_utils import _DISTRIBUTOR_LABEL_MAP

        assert "재고(NJ)" in _DISTRIBUTOR_LABEL_MAP
        assert _DISTRIBUTOR_LABEL_MAP["재고(NJ)"] == "warehouse_nj"


# ---------------------------------------------------------------------------
# REQ-PO5-002: parse_daily_review_excel recognizes warehouse labels
# ---------------------------------------------------------------------------


class TestParseDailyReviewExcelWarehouseCodes:
    """REQ-PO5-002: Parser correctly maps Korean warehouse labels to internal codes."""

    def test_parses_warehouse_korea_label(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_daily_review_excel([
            {"isbn": "9788901234567", "selected": "재고(한국)"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["sku"] == "9788901234567"
        assert results[0]["distributor"] == "warehouse_korea"

    def test_parses_warehouse_ca_label(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_daily_review_excel([
            {"isbn": "9788901234568", "selected": "재고(CA)"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["distributor"] == "warehouse_ca"

    def test_parses_warehouse_nj_label(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_daily_review_excel([
            {"isbn": "9788901234569", "selected": "재고(NJ)"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["distributor"] == "warehouse_nj"


# ---------------------------------------------------------------------------
# REQ-PO5-002/REQ-PO5-003: generate_daily_review_excel uses new column headers
# ---------------------------------------------------------------------------


class TestDailyReviewExcelHeaders:
    """REQ-PO5-002: Generated Excel has location columns instead of old warehouse columns."""

    def test_headers_contain_location_columns(self):
        from order.excel_utils import generate_daily_review_excel

        file_bytes = generate_daily_review_excel([])
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "재고(한국)" in headers
        assert "재고(CA)" in headers
        assert "재고(NJ)" in headers

    def test_headers_do_not_contain_old_columns(self):
        from order.excel_utils import generate_daily_review_excel

        file_bytes = generate_daily_review_excel([])
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "창고재고수량" not in headers
        assert "창고재고위치" not in headers

    def test_data_row_uses_new_stock_keys(self):
        from order.excel_utils import generate_daily_review_excel

        rows = [{
            "order_name": "#001", "sku": "9788901234567", "title": "Test", "quantity": 2,
            "location": "", "note": "",
            "korea_stock": 5, "ca_stock": 3, "nj_stock": 1,
            "bs_price": None, "ky_price": None, "bs_status": None,
            "ky_stock": None, "ky_status": None, "price_diff": None,
            "bs_arrival": None, "bs_returnable": None, "ky_available": None,
            "ky_returnable": None, "price_diff_alert": False,
            "publisher": None, "candidate_basis": None, "selected": "북센",
        }]
        file_bytes = generate_daily_review_excel(rows)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        korea_col_idx = headers.index("재고(한국)")
        ca_col_idx = headers.index("재고(CA)")
        nj_col_idx = headers.index("재고(NJ)")

        data_row = list(ws.iter_rows(values_only=True))[1]
        assert data_row[korea_col_idx] == 5
        assert data_row[ca_col_idx] == 3
        assert data_row[nj_col_idx] == 1


# ---------------------------------------------------------------------------
# SPEC-PURCHASE-ORDER-010 (T3): read-side reorder-candidate widening —
# DailyReviewExcelView (GET /api/purchase-orders/daily-review-excel/)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestDailyReviewExcelViewDamagedExchange:
    """AC-DMG-005 (scenario 1): damaged_exchange LineItems re-enter the Daily
    Review download regardless of existing PurchaseOrder linkage."""

    def test_damaged_exchange_linked_line_item_reexposed(self, auth_client):
        sku = "9791100000299"
        order = _make_order(shopify_order_id=87001)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        po = PurchaseOrder.objects.create(
            sku=sku, title="Book", distributor="booxen", quantity=1
        )
        po.line_items.add(li)
        li.purchase_status = "damaged_exchange"
        li.save(update_fields=["purchase_status"])

        res = auth_client.get(DAILY_REVIEW_EXCEL_URL)
        assert res.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        sku_col = headers.index("ISBN")
        skus_in_file = [row[sku_col] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert sku in skus_in_file

    def test_cs_required_linked_status_still_excluded(self, auth_client):
        sku = "9791100000298"
        order = _make_order(shopify_order_id=87002)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        li.purchase_status = "cs_required"
        li.save(update_fields=["purchase_status"])

        res = auth_client.get(DAILY_REVIEW_EXCEL_URL)
        assert res.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        sku_col = headers.index("ISBN")
        skus_in_file = [row[sku_col] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert sku not in skus_in_file


# ---------------------------------------------------------------------------
# REQ-PO5-008: VALID_DISTRIBUTORS includes warehouse codes
# ---------------------------------------------------------------------------


class TestValidDistributors:
    """REQ-PO5-008: VALID_DISTRIBUTORS set includes warehouse codes."""

    def test_valid_distributors_includes_warehouse_korea(self):
        from order.purchase_order_views import VALID_DISTRIBUTORS

        assert "warehouse_korea" in VALID_DISTRIBUTORS

    def test_valid_distributors_includes_warehouse_ca(self):
        from order.purchase_order_views import VALID_DISTRIBUTORS

        assert "warehouse_ca" in VALID_DISTRIBUTORS

    def test_valid_distributors_includes_warehouse_nj(self):
        from order.purchase_order_views import VALID_DISTRIBUTORS

        assert "warehouse_nj" in VALID_DISTRIBUTORS


# ---------------------------------------------------------------------------
# REQ-PO5-008: GenerateOrderFileView accepts warehouse distributor codes
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestGenerateOrderFileAcceptsWarehouseCodes:
    """REQ-PO5-008: generate-order-file accepts warehouse distributor codes (200, not 400)."""

    def _setup_sku(self, sku: str = "9788901234567", quantity: int = 2):
        order = _make_order(shopify_order_id=82001)
        _make_line_item(order, sku=sku, quantity=quantity, shopify_line_item_id=10)

    def test_warehouse_korea_returns_excel_or_warning(self, auth_client):
        """warehouse_korea is a valid distributor — not 400."""
        self._setup_sku()
        payload = {"distributor": "warehouse_korea", "skus": ["9788901234567"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        # Either Excel (200) or warning JSON (200) — both are OK; what matters is NOT 400
        assert res.status_code == 200

    def test_warehouse_ca_returns_200(self, auth_client):
        self._setup_sku("9788901234568")
        payload = {"distributor": "warehouse_ca", "skus": ["9788901234568"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 200

    def test_warehouse_nj_returns_200(self, auth_client):
        self._setup_sku("9788901234569")
        payload = {"distributor": "warehouse_nj", "skus": ["9788901234569"]}
        res = auth_client.post(GENERATE_URL, data=payload, format="json")
        assert res.status_code == 200


# ---------------------------------------------------------------------------
# REQ-PO5-004: UploadDailyReviewView deducts WarehouseStock
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadWarehouseDeductsStock:
    """REQ-PO5-004: Warehouse upload deducts LineItem quantities from WarehouseStock."""

    def test_upload_warehouse_deducts_stock(self, auth_client):
        """CA warehouse stock is decremented by sum of unordered LineItem quantities."""
        sku = "9788901234567"
        order = _make_order(shopify_order_id=83001)
        _make_line_item(order, sku=sku, quantity=2, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=10)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(
            UPLOAD_DAILY_URL,
            data={"file": file_obj},
            format="multipart",
        )
        assert res.status_code == 201

        stock = WarehouseStock.objects.get(isbn=sku, location="ca")
        assert stock.quantity == 8  # 10 - 2

    def test_upload_warehouse_deducts_korea_stock(self, auth_client):
        """Korea warehouse stock is decremented correctly."""
        sku = "9788901234570"
        order = _make_order(shopify_order_id=83002)
        _make_line_item(order, sku=sku, quantity=3, shopify_line_item_id=2)
        WarehouseStock.objects.create(isbn=sku, location="korea", quantity=20)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(한국)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")

        stock = WarehouseStock.objects.get(isbn=sku, location="korea")
        assert stock.quantity == 17  # 20 - 3

    def test_upload_warehouse_deducts_nj_stock(self, auth_client):
        """NJ warehouse stock is decremented correctly."""
        sku = "9788901234571"
        order = _make_order(shopify_order_id=83003)
        _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=3)
        WarehouseStock.objects.create(isbn=sku, location="nj", quantity=5)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(NJ)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")

        stock = WarehouseStock.objects.get(isbn=sku, location="nj")
        assert stock.quantity == 4  # 5 - 1


# ---------------------------------------------------------------------------
# REQ-PO5-004: Floor at zero
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadWarehouseFloorAtZero:
    """REQ-PO5-004: Quantity never goes negative; floored at 0."""

    def test_floor_at_zero_when_stock_insufficient(self, auth_client):
        """If LineItem quantity > current stock, result is 0 (not negative)."""
        sku = "9788901299001"
        order = _make_order(shopify_order_id=84001)
        _make_line_item(order, sku=sku, quantity=10, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=3)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")

        stock = WarehouseStock.objects.get(isbn=sku, location="ca")
        assert stock.quantity == 0  # Cannot go below 0

    def test_zero_initial_stock_stays_zero(self, auth_client):
        """Zero stock stays at zero after deduction."""
        sku = "9788901299002"
        order = _make_order(shopify_order_id=84002)
        _make_line_item(order, sku=sku, quantity=5, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="korea", quantity=0)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(한국)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")

        stock = WarehouseStock.objects.get(isbn=sku, location="korea")
        assert stock.quantity == 0


# ---------------------------------------------------------------------------
# REQ-PO5-005: LineItem.purchase_status set to "in_stock"
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadWarehouseSetsInStockStatus:
    """REQ-PO5-005: After warehouse upload, LineItems get purchase_status="in_stock"."""

    def test_sets_in_stock_status(self, auth_client):
        sku = "9788901299003"
        order = _make_order(shopify_order_id=85001)
        li = _make_line_item(order, sku=sku, quantity=2, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=10)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.purchase_status == "in_stock"

    def test_sets_in_stock_with_note(self, auth_client):
        """When note is provided, LineItemNote is created (SPEC-ORDER-010 migration)."""
        from order.models import LineItemNote
        sku = "9788901299004"
        order = _make_order(shopify_order_id=85002)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=2)
        WarehouseStock.objects.create(isbn=sku, location="korea", quantity=5)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(한국)", "note": "한국 창고에서 발송"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")

        li.refresh_from_db()
        assert li.purchase_status == "in_stock"
        assert LineItemNote.objects.filter(line_item=li, content="한국 창고에서 발송").exists()


# ---------------------------------------------------------------------------
# REQ-PO5-006: No PurchaseOrder created for warehouse selection
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadWarehouseNoPurchaseOrder:
    """REQ-PO5-006: Warehouse uploads do NOT create PurchaseOrder records."""

    def test_no_purchase_order_created(self, auth_client):
        sku = "9788901299005"
        order = _make_order(shopify_order_id=86001)
        _make_line_item(order, sku=sku, quantity=2, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="nj", quantity=10)

        po_count_before = PurchaseOrder.objects.count()

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(NJ)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        po_count_after = PurchaseOrder.objects.count()
        assert po_count_after == po_count_before  # No new PurchaseOrder

    def test_confirmed_count_incremented_for_warehouse(self, auth_client):
        """Warehouse entries increment confirmed_count."""
        sku = "9788901299006"
        order = _make_order(shopify_order_id=86002)
        _make_line_item(order, sku=sku, quantity=2, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=10)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        assert res.data["confirmed_count"] == 1


# ---------------------------------------------------------------------------
# REQ-PO5-007: Response includes confirmed_by_distributor
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadResponseConfirmedByDistributor:
    """REQ-PO5-007: Upload response includes confirmed_by_distributor breakdown."""

    def test_response_contains_confirmed_by_distributor_key(self, auth_client):
        """confirmed_by_distributor key must be present in response."""
        sku = "9788901299007"
        order = _make_order(shopify_order_id=87001)
        _make_line_item(order, sku=sku, quantity=2, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=10)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        assert "confirmed_by_distributor" in res.data

    def test_warehouse_entry_appears_in_confirmed_by_distributor(self, auth_client):
        """Warehouse SKU appears under its warehouse_* key."""
        sku = "9788901299008"
        order = _make_order(shopify_order_id=87002)
        _make_line_item(order, sku=sku, quantity=3, title="창고 도서", shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="korea", quantity=10)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(한국)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        cbd = res.data["confirmed_by_distributor"]
        assert "warehouse_korea" in cbd
        entries = cbd["warehouse_korea"]
        assert len(entries) == 1
        assert entries[0]["sku"] == sku
        assert entries[0]["quantity"] == 3

    def test_non_warehouse_entry_appears_in_confirmed_by_distributor(self, auth_client):
        """Non-warehouse SKU (booxen) appears under booxen key."""
        sku = "9788901299009"
        order = _make_order(shopify_order_id=87003)
        _make_line_item(order, sku=sku, quantity=2, title="북센 도서", shopify_line_item_id=1)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "북센"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        cbd = res.data["confirmed_by_distributor"]
        assert "booxen" in cbd
        assert len(cbd["booxen"]) == 1
        assert cbd["booxen"][0]["sku"] == sku

    def test_mixed_upload_groups_correctly(self, auth_client):
        """Mixed warehouse + regular upload groups all SKUs correctly."""
        sku_wh = "9788901299010"
        sku_bs = "9788901299011"

        order = _make_order(shopify_order_id=87004)
        _make_line_item(order, sku=sku_wh, quantity=2, shopify_line_item_id=1)
        _make_line_item(order, sku=sku_bs, quantity=1, shopify_line_item_id=2)
        WarehouseStock.objects.create(isbn=sku_wh, location="nj", quantity=10)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku_wh, "selected": "재고(NJ)"},
            {"isbn": sku_bs, "selected": "북센"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        cbd = res.data["confirmed_by_distributor"]
        assert "warehouse_nj" in cbd
        assert "booxen" in cbd
        assert res.data["confirmed_count"] == 2


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadWarehouseEdgeCases:
    """Edge cases for warehouse upload logic."""

    def test_warehouse_upload_no_existing_stock_record(self, auth_client):
        """If no WarehouseStock record exists, upload still processes LineItems (no error)."""
        sku = "9788901299020"
        order = _make_order(shopify_order_id=88001)
        li = _make_line_item(order, sku=sku, quantity=2, shopify_line_item_id=1)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        # Should not error — no WarehouseStock means no update but LineItem still processed
        assert res.status_code == 201
        li.refresh_from_db()
        assert li.purchase_status == "in_stock"

    def test_warehouse_skipped_if_no_unordered_line_items(self, auth_client):
        """If no unordered LineItems for the SKU, warehouse entry is skipped."""
        sku = "9788901299021"
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=10)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        assert res.data["skipped_count"] == 1

        # Stock should not be touched
        stock = WarehouseStock.objects.get(isbn=sku, location="ca")
        assert stock.quantity == 10


# ===========================================================================
# SPEC-PURCHASE-ORDER-008
#
# Covers:
#   REQ-PO8-001  Header row auto-detection (old row1 ISBN+선택 / new row3 Lineitem sku+선택)
#   REQ-PO8-002  Dual SKU column header name support
#   REQ-PO8-003  Headerless legend column never indexed as data
#   REQ-PO8-004  bs_price/ky_price/yes24_price parsing from new template columns
#   REQ-PO8-005  Note/memo source switch (Status vs 메모)
#   REQ-PO8-006  _DISTRIBUTOR_LABEL_MAP: YES24 -> yes24
#   REQ-PO8-007  _DISTRIBUTOR_LABEL_MAP: 재고 -> warehouse (generic)
#   REQ-PO8-008  Warehouse location resolved from Status value (new template)
#   REQ-PO8-009  LineItemNote assignee determined by resolved location
#   REQ-PO8-010  YES24 purchase-order confirmation branch
#   REQ-PO8-011  Unrecognized 선택 values (합계/total/check) skipped from PO logic
#   REQ-PO8-012  Legacy label map entries preserved
#   REQ-PO8-013  Old format regression safety
#   REQ-PO8-014  Vendor table upsert runs for every row with a SKU
#   REQ-PO8-015  BooxenData field mapping
#   REQ-PO8-016  KyoboData field mapping (existing fields)
#   REQ-PO8-017  KyoboData.list_price field + conditional inclusion
#   REQ-PO8-018  Yes24Data field mapping + list_price protection
# ===========================================================================


def _make_new_template_excel(
    rows: list[dict],
    *,
    include_kyobo_list_price: bool = False,
    junk_rows: int = 2,
) -> bytes:
    """
    Build a 'Daily Order Review Template' .xlsx matching the real external
    template structure (SPEC-PURCHASE-ORDER-008): the header is NOT row 1
    (simulated via leading junk rows), the SKU column is named
    'Lineitem sku' rather than 'ISBN', and a headerless legend column
    (containing unrelated reference values, never real row data) sits
    immediately to the right of '선택'.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    for _ in range(junk_rows):
        ws.append(["작업 파일 안내", None, None])

    header = [
        "", "Shipping Zip", "Date", "Name", "Lineitem sku", "Lineitem name", "Email",
        "Total", "location", "Status", "in Stock", "Stock Location", "SKU별 필요수량",
        "BOOXEN 공급가", "YES24 공급가", "YES24 재고상태", "교보 공급가",
        "BOOXEN 재고수량", "BOOXEN 재고상태", "교보 재고수량", "교보 재고상태",
        "비교시작", "Min Cost", "BOOXEN 가격비교", "YES24 가격비교", "교보 가격비교",
        "차이", "기준over", "BOOXEN 입고예정", "BOOXEN 반품", "교보 상품상태",
        "교보 분야", "교보 출판사", "교보 반품가능여부", "가격차이알림", "후보기준", "선택",
        "",  # headerless legend column immediately right of 선택 (REQ-PO8-003)
    ]
    if include_kyobo_list_price:
        header.append("교보 정가")
    ws.append(header)

    legend_values = [
        "total", "BOOXEN", "교보", "YES24", "주문취소", "재고",
        "주문보류", "CS필요", "타출판사", "합계", "check",
    ]

    for i, row in enumerate(rows):
        data = [""] * len(header)

        def set_col(name: str, value) -> None:
            data[header.index(name)] = value

        # REQ-PO8-018: default matches _make_order()'s default name
        # ("#8001") so existing single-order tests that never pass
        # order_name still exercise the same LineItem/order pairing they
        # did before the (order_name, sku) fix.
        set_col("Name", row.get("order_name", "#8001"))
        set_col("Lineitem sku", row.get("sku", ""))
        set_col("Status", row.get("status", ""))
        set_col("선택", row.get("selected", ""))
        for key, col_name in (
            ("bs_price", "BOOXEN 공급가"),
            ("ky_price", "교보 공급가"),
            ("yes24_price", "YES24 공급가"),
            ("bs_stock", "BOOXEN 재고수량"),
            ("bs_status", "BOOXEN 재고상태"),
            ("bs_arrival", "BOOXEN 입고예정"),
            ("bs_returnable", "BOOXEN 반품"),
            ("yes24_status", "YES24 재고상태"),
            ("ky_stock", "교보 재고수량"),
            ("ky_available", "교보 재고상태"),
            ("ky_status", "교보 상품상태"),
            ("ky_returnable", "교보 반품가능여부"),
            ("ky_publisher", "교보 출판사"),
        ):
            if key in row:
                set_col(col_name, row[key])
        if include_kyobo_list_price and "ky_list_price" in row:
            set_col("교보 정가", row["ky_list_price"])

        # Legend column: unrelated reference values, never real row data.
        data[header.index("선택") + 1] = legend_values[i % len(legend_values)]
        ws.append(data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# REQ-PO8-001/002/003: parse_daily_review_excel header auto-detection
# ---------------------------------------------------------------------------


class TestParseDailyReviewHeaderAutoDetection:
    """AC-001/AC-002/AC-003: header auto-detected for both formats; legend ignored."""

    def test_ac001_new_template_header_at_row_3_detected(self):
        """AC-001: header at row 3 (Lineitem sku + 선택) is recognized; data starts row 4."""
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_new_template_excel([
            {"sku": "9791124591055", "selected": "BOOXEN", "status": "정상"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["sku"] == "9791124591055"
        assert results[0]["distributor"] == "booxen"

    def test_ac002_legacy_header_at_row_1_still_works(self):
        """AC-002: existing fixture (1행 ISBN+선택) still parses without regression."""
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_daily_review_excel([
            {"isbn": "9788901234567", "selected": "북센"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["sku"] == "9788901234567"
        assert results[0]["distributor"] == "booxen"

    def test_ac003_legend_column_values_never_leak_into_data(self):
        """AC-003: the headerless legend column right of 선택 never pollutes any field."""
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_new_template_excel([
            {"sku": "9791124591055", "selected": "BOOXEN", "status": "한국재고"},
            {"sku": "9791124591056", "selected": "교보", "status": ""},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 2
        legend_values = {
            "total", "BOOXEN", "교보", "YES24", "주문취소", "재고",
            "주문보류", "CS필요", "타출판사", "합계", "check",
        }
        for row in results:
            # sku must be the real Lineitem sku, never a legend value
            assert row["sku"] not in legend_values
            # distributor must be a valid internal code or None, never a raw legend string
            assert row["distributor"] in (None, "booxen", "kyobo", "yes24", "warehouse")

    def test_header_not_found_raises_value_error_mentioning_both_sku_columns(self):
        """REQ-PO8-001: error message mentions both ISBN and Lineitem sku when no header matches."""
        from order.excel_utils import parse_daily_review_excel

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["엉뚱한 헤더1", "엉뚱한 헤더2"])
        ws.append(["data1", "data2"])
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError) as exc_info:
            parse_daily_review_excel(buf.getvalue())
        assert "ISBN" in str(exc_info.value)
        assert "Lineitem sku" in str(exc_info.value)

    def test_row_with_selected_text_but_no_sku_column_is_not_mistaken_for_header(self):
        """A row containing a '선택'-looking cell without a recognized SKU column
        (e.g. stray legend/instruction text before the real header) must not be
        mistaken for the header row — scanning continues to the true header."""
        from order.excel_utils import parse_daily_review_excel

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["선택", "참고용 안내"])  # false positive: has '선택' text, no SKU column
        ws.append(["Lineitem sku", "Name", "선택"])  # real header (REQ-PO8-018: Name required)
        ws.append(["9791124591055", "#1234", "BOOXEN"])
        buf = io.BytesIO()
        wb.save(buf)

        results = parse_daily_review_excel(buf.getvalue())
        assert len(results) == 1
        assert results[0]["sku"] == "9791124591055"
        assert results[0]["distributor"] == "booxen"

    def test_row_with_blank_sku_cell_is_skipped(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_new_template_excel([
            {"sku": "", "selected": "BOOXEN", "status": "정상"},
            {"sku": "9791124591056", "selected": "BOOXEN", "status": "정상"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["sku"] == "9791124591056"


# ---------------------------------------------------------------------------
# REQ-PO8-004: BOOXEN/교보/YES24 공급가 parsing (new template)
# ---------------------------------------------------------------------------


class TestParseDailyReviewNewTemplatePrices:
    def test_prices_parsed_from_new_template_columns(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_new_template_excel([
            {
                "sku": "9791124591055", "selected": "BOOXEN", "status": "정상",
                "bs_price": 9000, "ky_price": 9500, "yes24_price": 9800,
            },
        ])
        results = parse_daily_review_excel(file_bytes)
        assert results[0]["bs_price"] == 9000.0
        assert results[0]["ky_price"] == 9500.0
        assert results[0]["yes24_price"] == 9800.0

    def test_missing_price_columns_keep_keys_as_none(self):
        """REQ-PO8-004: old format (no BOOXEN/YES24 공급가 columns) keeps keys as None."""
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_daily_review_excel([
            {"isbn": "9788901234567", "selected": "북센"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert "yes24_price" in results[0]
        assert results[0]["yes24_price"] is None


# ---------------------------------------------------------------------------
# REQ-PO8-005: note/memo source column switch
# ---------------------------------------------------------------------------


class TestParseDailyReviewNoteSourceSwitch:
    def test_new_template_uses_status_column_as_note(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_new_template_excel([
            {"sku": "9791124591055", "selected": "재고", "status": "한국재고"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert results[0]["note"] == "한국재고"

    def test_legacy_format_uses_note_column(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_daily_review_excel([
            {"isbn": "9788901234567", "selected": "북센", "note": "레거시 메모"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert results[0]["note"] == "레거시 메모"


# ---------------------------------------------------------------------------
# REQ-PO8-006/007/012: _DISTRIBUTOR_LABEL_MAP additions + backward compat
# ---------------------------------------------------------------------------


class TestDistributorLabelMapNewCodes:
    def test_yes24_label_maps_to_yes24_code(self):
        from order.excel_utils import _DISTRIBUTOR_LABEL_MAP

        assert _DISTRIBUTOR_LABEL_MAP["YES24"] == "yes24"

    def test_generic_warehouse_label_maps_to_warehouse_code(self):
        from order.excel_utils import _DISTRIBUTOR_LABEL_MAP

        assert _DISTRIBUTOR_LABEL_MAP["재고"] == "warehouse"

    def test_legacy_mappings_preserved(self):
        """REQ-PO8-012: existing 8 mappings must remain untouched."""
        from order.excel_utils import _DISTRIBUTOR_LABEL_MAP

        expected = {
            "북센": "booxen",
            "교보": "kyobo",
            "처음교육": "choeumgoyuk",
            "아가페": "agape",
            "성서유니온": "sungseoyunion",
            "재고(한국)": "warehouse_korea",
            "재고(CA)": "warehouse_ca",
            "재고(NJ)": "warehouse_nj",
        }
        for label, code in expected.items():
            assert _DISTRIBUTOR_LABEL_MAP[label] == code


# ---------------------------------------------------------------------------
# REQ-PO8-011/014: rows with empty/unrecognized 선택 still returned (Part B)
# ---------------------------------------------------------------------------


class TestParseDailyReviewIncludesAllValidSkuRows:
    def test_blank_selected_row_still_returned(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_new_template_excel([
            {"sku": "9791124591055", "selected": "", "bs_price": 9000},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["distributor"] is None
        assert results[0]["note_type"] is None

    def test_unrecognized_selected_row_still_returned(self):
        """REQ-PO8-011: '합계' (legend value, not a real selection) still returned for Part B."""
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_new_template_excel([
            {"sku": "9791124591055", "selected": "합계", "bs_price": 9000},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["distributor"] is None
        assert results[0]["note_type"] is None


# ---------------------------------------------------------------------------
# CS note-type labels still recognized through the rewritten header/selection
# logic (pre-existing branch, exercised here to guard the refactor).
# ---------------------------------------------------------------------------


class TestParseDailyReviewCsNoteTypeStillRecognized:
    def test_cs_label_recognized_via_new_template(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_new_template_excel([
            {"sku": "9791100000099", "selected": "주문취소", "status": "고객 요청으로 취소"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["distributor"] is None
        assert results[0]["note_type"] == "주문취소"
        assert results[0]["note"] == "고객 요청으로 취소"


@pytest.mark.django_db
class TestUploadCsNoteTypeViaNewTemplate:
    def test_cs_label_updates_status_and_creates_cs_note(self, auth_client):
        """Guards the restructured CS branch: still reachable via the new
        template's Status-sourced note, still sets purchase_status and
        creates a LineItemNote assigned to CS."""
        from order.models import LineItemNote

        sku = "9791100000098"
        order = _make_order(shopify_order_id=90099)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "주문취소", "status": "고객 요청"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.purchase_status == "order_cancelled"

        note = LineItemNote.objects.get(line_item=li)
        assert note.content == "고객 요청"
        assert note.assignee == "CS"
        assert note.note_type == "주문취소"


# ---------------------------------------------------------------------------
# SPEC-PURCHASE-ORDER-010 (T2): _NOTE_TYPE_STATUS_MAP damaged_exchange entry
# ---------------------------------------------------------------------------


class TestParseDailyReviewDamagedExchangeRecognized:
    def test_damaged_exchange_label_recognized_as_note_type(self):
        """REQ-DMG-003: '파손/교환' in the 선택 column is recognized as a
        CS-type note_type (not a distributor), same mechanism as the
        existing four CS-type labels."""
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_daily_review_excel([
            {"isbn": "9791100000199", "selected": "파손/교환", "note": "박스 파손 확인"},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert len(results) == 1
        assert results[0]["distributor"] is None
        assert results[0]["note_type"] == "파손/교환"
        assert results[0]["note"] == "박스 파손 확인"


@pytest.mark.django_db
class TestUploadDamagedExchangeNoteTypeAutoApplies:
    """AC-DMG-003: Daily Review upload's CS branch auto-applies
    damaged_exchange via the existing map-lookup code path — no production
    code change needed in the branch itself, only the map entry (T2)."""

    def test_damaged_exchange_selection_sets_status_and_creates_note(self, auth_client):
        sku = "9791100000198"
        order = _make_order(shopify_order_id=90098)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "파손/교환", "note": "박스 파손, 교환 요청"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.purchase_status == "damaged_exchange"

        from order.models import LineItemNote

        note = LineItemNote.objects.get(line_item=li)
        assert note.content == "박스 파손, 교환 요청"
        assert note.assignee == "CS"
        assert note.note_type == "파손/교환"

    def test_damaged_exchange_selection_does_not_create_purchase_order(self, auth_client):
        """CS-type selections never create a PurchaseOrder (same as the
        other four CS labels)."""
        sku = "9791100000197"
        order = _make_order(shopify_order_id=90097)
        _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        po_count_before = PurchaseOrder.objects.count()

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "파손/교환"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")

        assert PurchaseOrder.objects.count() == po_count_before


# ---------------------------------------------------------------------------
# SPEC-PURCHASE-ORDER-010 (T3+T6): UploadDailyReviewView SKU-batch read-side
# widening + non-warehouse write-side auto-reset
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadDailyReviewDamagedExchangeReorder:
    """AC-DMG-005 (read-side) + AC-DMG-006/006C (write-side auto-reset,
    batch-scope isolation)."""

    def test_linked_damaged_exchange_sku_is_matched_not_skipped(self, auth_client):
        """T3: the SKU-batch query (UploadDailyReviewView.post()) must pick up
        an already-linked damaged_exchange LineItem instead of treating it as
        already-ordered (which would previously fall into skipped_count)."""
        sku = "9791100000296"
        order = _make_order(shopify_order_id=88001)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        old_po = PurchaseOrder.objects.create(
            sku=sku, title="Book", distributor="booxen", quantity=1
        )
        old_po.line_items.add(li)
        li.purchase_status = "damaged_exchange"
        li.save(update_fields=["purchase_status"])

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "북센"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        assert res.data["confirmed_count"] == 1
        assert res.data["skipped_count"] == 0

    def test_new_po_confirmation_resets_damaged_exchange_to_unordered(self, auth_client):
        """AC-DMG-006 (scenario 3): when the batch's new PurchaseOrder is
        created and linked, the damaged_exchange LineItem's purchase_status
        is auto-reset to unordered as part of the same write."""
        sku = "9791100000295"
        order = _make_order(shopify_order_id=88002)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        old_po = PurchaseOrder.objects.create(
            sku=sku, title="Book", distributor="booxen", quantity=1
        )
        old_po.line_items.add(li)
        li.purchase_status = "damaged_exchange"
        li.save(update_fields=["purchase_status"])

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "북센"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.purchase_status == "unordered"

        # Re-query the reorder-candidate list: no longer a candidate.
        res2 = auth_client.get(DAILY_REVIEW_EXCEL_URL)
        wb = openpyxl.load_workbook(io.BytesIO(res2.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        sku_col = headers.index("ISBN")
        skus_in_file = [row[sku_col] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert sku not in skus_in_file

    def test_reset_scoped_to_current_batch_only(self, auth_client):
        """AC-DMG-006C (scenario 3b): a damaged_exchange LineItem for a SKU
        NOT included in this upload's file is left untouched."""
        other_sku = "9791100000294"
        other_order = _make_order(shopify_order_id=88003, name="#8801")
        other_li = _make_line_item(other_order, sku=other_sku, quantity=1, shopify_line_item_id=1)
        other_po = PurchaseOrder.objects.create(
            sku=other_sku, title="Other Book", distributor="booxen", quantity=1
        )
        other_po.line_items.add(other_li)
        other_li.purchase_status = "damaged_exchange"
        other_li.save(update_fields=["purchase_status"])

        batch_sku = "9791100000293"
        batch_order = _make_order(shopify_order_id=88004, name="#8802")
        _make_line_item(batch_order, sku=batch_sku, quantity=1, shopify_line_item_id=1)

        file_bytes = _make_daily_review_excel([
            {"isbn": batch_sku, "selected": "북센"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        other_li.refresh_from_db()
        assert other_li.purchase_status == "damaged_exchange"

    def test_cs_branch_overwrites_damaged_exchange_regression(self, auth_client):
        """T6 regression: CS branch (~1219-1237) already unconditionally
        overwrites purchase_status, so a damaged_exchange LineItem selected
        with a different CS label ends up with that label's status, not a
        no-op stuck-at-damaged_exchange state."""
        sku = "9791100000292"
        order = _make_order(shopify_order_id=88005)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        li.purchase_status = "damaged_exchange"
        li.save(update_fields=["purchase_status"])

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "주문보류"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.purchase_status == "on_hold"

    def test_warehouse_branch_overwrites_damaged_exchange_regression(self, auth_client):
        """T6 regression: warehouse branch (~1250-1281) already unconditionally
        sets purchase_status="in_stock", so a damaged_exchange LineItem
        selected into the warehouse branch ends up "in_stock", unaffected by
        the new auto-reset logic (which only applies to the non-warehouse
        branch)."""
        sku = "9791100000291"
        order = _make_order(shopify_order_id=88006)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        li.purchase_status = "damaged_exchange"
        li.save(update_fields=["purchase_status"])
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=10)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.purchase_status == "in_stock"


# ---------------------------------------------------------------------------
# AC-004 .. AC-014: UploadDailyReviewView integration tests (new template)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadNewTemplateBooxenAndVendorUpsert:
    def test_ac004_booxen_selection_confirms_po_and_upserts_all_vendor_tables(self, auth_client):
        """AC-004: real sample SKU — BOOXEN selection creates PO and upserts
        BooxenData/KyoboData/Yes24Data simultaneously."""
        from order.models import BooxenData, KyoboData, Yes24Data

        sku = "9791124591055"
        order = _make_order(shopify_order_id=90001)
        _make_line_item(order, sku=sku, title="애니가 남긴 것", quantity=2, shopify_line_item_id=1)

        file_bytes = _make_new_template_excel([
            {
                "sku": sku, "selected": "BOOXEN", "status": "정상",
                "bs_price": 9000, "ky_price": 9500, "yes24_price": 9800,
            },
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        po = PurchaseOrder.objects.get(sku=sku)
        assert po.distributor == "booxen"

        assert BooxenData.objects.filter(sku=sku).exists()
        assert KyoboData.objects.filter(sku=sku).exists()
        assert Yes24Data.objects.filter(sku=sku).exists()

    def test_daily_review_upload_creates_po_with_confirmed_status(self, auth_client):
        """Daily Review upload reflects the distributor's actual confirmed
        response, so the PurchaseOrder it creates starts status="confirmed"
        (unlike ConfirmOrderView's status="pending" staging step)."""
        sku = "9791124591099"
        order = _make_order(shopify_order_id=90099)
        _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "BOOXEN", "status": "정상", "bs_price": 9000},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        po = PurchaseOrder.objects.get(sku=sku)
        assert po.status == "confirmed"


@pytest.mark.django_db
class TestUploadYes24Confirmation:
    def test_ac005_yes24_selection_confirms_po_with_parsed_price(self, auth_client):
        """AC-005: YES24 선택 시 PO 생성 + unit_price는 파싱된 YES24 공급가."""
        sku = "9791100000001"
        order = _make_order(shopify_order_id=90002)
        _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "YES24", "status": "정상", "yes24_price": 15000},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        po = PurchaseOrder.objects.get(sku=sku)
        assert po.distributor == "yes24"
        assert po.unit_price == Decimal("15000")

    def test_ac006_yes24_price_blank_falls_back_to_yes24data(self, auth_client):
        """AC-006: YES24 공급가 미기재 시 기존 Yes24Data.price로 폴백 (Part B upsert가
        이 폴백 값을 훼손하지 않아야 한다)."""
        from order.models import Yes24Data

        sku = "9791100000002"
        Yes24Data.objects.create(sku=sku, price=Decimal("12000"))
        order = _make_order(shopify_order_id=90003)
        _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "YES24", "status": "정상"},  # yes24_price blank
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        po = PurchaseOrder.objects.get(sku=sku)
        assert po.unit_price == Decimal("12000")


@pytest.mark.django_db
class TestUploadWarehouseStatusBasedLocation:
    def test_ac007_status_fullerton_resolves_ca_and_us_warehouse_assignee(self, auth_client):
        """AC-007: Status='Fullerton재고' -> ca 위치, 미국창고 담당자."""
        from order.models import LineItemNote

        sku = "9791100000003"
        order = _make_order(shopify_order_id=90004)
        li = _make_line_item(order, sku=sku, quantity=2, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=5)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "재고", "status": "Fullerton재고"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        stock = WarehouseStock.objects.get(isbn=sku, location="ca")
        assert stock.quantity == 3

        li.refresh_from_db()
        assert li.purchase_status == "in_stock"

        note = LineItemNote.objects.get(line_item=li)
        assert note.content == "Fullerton재고"
        assert note.assignee == "미국창고"

    def test_ac008_status_korea_resolves_korea_and_korea_warehouse_assignee(self, auth_client):
        """AC-008: Status='한국재고' -> korea 위치, 한국창고 담당자."""
        from order.models import LineItemNote

        sku = "9791100000004"
        order = _make_order(shopify_order_id=90005)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="korea", quantity=10)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "재고", "status": "한국재고"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        stock = WarehouseStock.objects.get(isbn=sku, location="korea")
        assert stock.quantity == 9

        note = LineItemNote.objects.get(line_item=li)
        assert note.content == "한국재고"
        assert note.assignee == "한국창고"

    def test_unrecognized_status_value_skips_warehouse_processing(self, auth_client):
        """REQ-PO8-008: unrecognized Status value under 재고 selection is skipped safely."""
        sku = "9791100000005"
        order = _make_order(shopify_order_id=90006)
        _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=10)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "재고", "status": "알수없는위치"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        assert res.data["skipped_count"] == 1

        stock = WarehouseStock.objects.get(isbn=sku, location="ca")
        assert stock.quantity == 10  # untouched

    def test_legacy_ca_warehouse_assignee_bug_fixed(self, auth_client):
        """REQ-PO8-009: legacy warehouse_ca now correctly logs 미국창고 (was 한국창고 bug)."""
        from order.models import LineItemNote

        sku = "9791100000006"
        order = _make_order(shopify_order_id=90007)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=5)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)", "note": "CA 창고 확인"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        note = LineItemNote.objects.get(line_item=li)
        assert note.assignee == "미국창고"

    def test_legacy_nj_warehouse_assignee_bug_fixed(self, auth_client):
        """REQ-PO8-009: legacy warehouse_nj now correctly logs 미국창고 (was 한국창고 bug)."""
        from order.models import LineItemNote

        sku = "9791100000007"
        order = _make_order(shopify_order_id=90008)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="nj", quantity=5)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(NJ)", "note": "NJ 창고 확인"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        note = LineItemNote.objects.get(line_item=li)
        assert note.assignee == "미국창고"

    def test_legacy_korea_warehouse_assignee_unchanged(self, auth_client):
        """REQ-PO8-009: legacy warehouse_korea keeps 한국창고 (no regression)."""
        from order.models import LineItemNote

        sku = "9791100000008"
        order = _make_order(shopify_order_id=90009)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="korea", quantity=5)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(한국)", "note": "한국 창고 확인"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        note = LineItemNote.objects.get(line_item=li)
        assert note.assignee == "한국창고"


@pytest.mark.django_db
class TestWarehouseConfirmedDistributorLocationSuffix:
    """Bug fix: generic 'warehouse' code must persist the *resolved* location
    into LineItem.confirmed_distributor.

    Before the fix, the note-lookup path resolved `loc` (korea/ca/nj) correctly
    for the WarehouseStock deduction and the LineItemNote assignee, but then
    wrote the bare, unsuffixed `distributor_code` ("warehouse") into
    confirmed_distributor — discarding the location. The frontend
    DISTRIBUTOR_LABELS dict already carries 창고(한국)/창고(CA)/창고(NJ) labels
    keyed on warehouse_korea/warehouse_ca/warehouse_nj, which were therefore
    unreachable.
    """

    def test_generic_warehouse_fullerton_sets_confirmed_distributor_ca(self, auth_client):
        sku = "9791100000021"
        order = _make_order(shopify_order_id=90021)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=5)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "재고", "status": "Fullerton재고"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.confirmed_distributor == "warehouse_ca"

    def test_generic_warehouse_korea_sets_confirmed_distributor_korea(self, auth_client):
        sku = "9791100000022"
        order = _make_order(shopify_order_id=90022)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="korea", quantity=5)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "재고", "status": "한국재고"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.confirmed_distributor == "warehouse_korea"

    def test_generic_warehouse_nj_sets_confirmed_distributor_nj(self, auth_client):
        sku = "9791100000023"
        order = _make_order(shopify_order_id=90023)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="nj", quantity=5)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "재고", "status": "NJ재고"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.confirmed_distributor == "warehouse_nj"

    def test_legacy_suffixed_code_confirmed_distributor_unchanged(self, auth_client):
        """Regression guard: the legacy _WAREHOUSE_LOCATION_MAP path already wrote
        the correct suffixed code and must keep writing `distributor_code` as-is."""
        sku = "9791100000024"
        order = _make_order(shopify_order_id=90024)
        li = _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=5)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "재고(CA)", "note": "CA 창고 확인"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        li.refresh_from_db()
        assert li.confirmed_distributor == "warehouse_ca"

    def test_confirmed_by_distributor_summary_still_keyed_by_bare_code(self, auth_client):
        """Regression guard: the response summary dict keys off the *original*
        distributor_code, not the location-resolved one. Only the persisted
        LineItem.confirmed_distributor field changes."""
        sku = "9791100000025"
        order = _make_order(shopify_order_id=90025)
        _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)
        WarehouseStock.objects.create(isbn=sku, location="ca", quantity=5)

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "재고", "status": "Fullerton재고"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        cbd = res.data["confirmed_by_distributor"]
        assert "warehouse" in cbd
        assert "warehouse_ca" not in cbd


@pytest.mark.django_db
class TestUploadUnrecognizedSelectedSkipsButUpsertsVendorData:
    def test_ac009_summary_row_skips_po_but_upserts_vendor_tables(self, auth_client):
        """AC-009: 선택='합계' -> PO/CS/창고 미실행, 벤더 upsert는 실행."""
        from order.models import BooxenData, KyoboData, Yes24Data

        sku = "9791100000009"
        po_count_before = PurchaseOrder.objects.count()

        file_bytes = _make_new_template_excel([
            {
                "sku": sku, "selected": "합계", "status": "",
                "bs_price": 9000, "ky_price": 9500, "yes24_price": 9800,
            },
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        assert PurchaseOrder.objects.count() == po_count_before
        assert BooxenData.objects.filter(sku=sku).exists()
        assert KyoboData.objects.filter(sku=sku).exists()
        assert Yes24Data.objects.filter(sku=sku).exists()

    def test_ac010_blank_selected_skips_po_but_upserts_vendor_tables(self, auth_client):
        """AC-010: 선택 값이 빈 문자열이어도 벤더 upsert는 독립적으로 실행."""
        from order.models import BooxenData, KyoboData, Yes24Data

        sku = "9791100000010"
        po_count_before = PurchaseOrder.objects.count()

        file_bytes = _make_new_template_excel([
            {
                "sku": sku, "selected": "", "status": "",
                "bs_price": 8000, "ky_price": 8500, "yes24_price": 8800,
            },
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        assert PurchaseOrder.objects.count() == po_count_before
        assert BooxenData.objects.get(sku=sku).price == Decimal("8000")
        assert KyoboData.objects.get(sku=sku).price == Decimal("8500")
        assert Yes24Data.objects.get(sku=sku).price == Decimal("8800")


@pytest.mark.django_db
class TestUploadKyoboFieldMapping:
    def test_ac011_kyobo_available_and_status_mapped_from_correct_columns(self, auth_client):
        """AC-011: 교보 재고상태(Y/N)->available, 교보 상품상태(정상/품절)->status,
        두 컬럼이 서로 뒤바뀌지 않아야 한다."""
        from order.models import KyoboData

        sku = "9791100000011"
        file_bytes = _make_new_template_excel([
            {
                "sku": sku, "selected": "", "status": "",
                "ky_available": "Y", "ky_status": "정상",
            },
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        kd = KyoboData.objects.get(sku=sku)
        assert kd.available is True
        assert kd.status == "정상"


@pytest.mark.django_db
class TestUploadKyoboListPrice:
    def test_ac012_list_price_column_absent_leaves_existing_value_unchanged(self, auth_client):
        """AC-012: 교보 정가 컬럼이 없는 현재 템플릿 업로드는 기존 list_price를 보존."""
        from order.models import KyoboData

        sku = "9791100000012"
        KyoboData.objects.create(sku=sku, list_price=Decimal("5000"))

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "", "status": "", "ky_price": 4500},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        kd = KyoboData.objects.get(sku=sku)
        assert kd.list_price == Decimal("5000")
        assert kd.price == Decimal("4500")

    def test_ac013_list_price_column_present_is_saved(self, auth_client):
        """AC-013: 교보 정가 컬럼이 추가된 워크북에서는 list_price가 저장된다."""
        from order.models import KyoboData

        sku = "9791100000013"
        file_bytes = _make_new_template_excel(
            [{"sku": sku, "selected": "", "status": "", "ky_list_price": 15000}],
            include_kyobo_list_price=True,
        )
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        kd = KyoboData.objects.get(sku=sku)
        assert kd.list_price == Decimal("15000")


@pytest.mark.django_db
class TestUploadYes24ListPriceProtected:
    def test_ac014_yes24_list_price_untouched_by_daily_review_upload(self, auth_client):
        """AC-014: Daily Review 업로드는 Yes24Data.price만 갱신하고 list_price는 보존."""
        from order.models import Yes24Data

        sku = "9791100000014"
        Yes24Data.objects.create(sku=sku, list_price=Decimal("20000"), price=Decimal("18000"))

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "", "status": "", "yes24_price": 19000},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        yd = Yes24Data.objects.get(sku=sku)
        assert yd.price == Decimal("19000")
        assert yd.list_price == Decimal("20000")


@pytest.mark.django_db
class TestUploadBooxenFieldMapping:
    def test_booxen_fields_mapped_from_new_template_columns(self, auth_client):
        """REQ-PO8-015: BooxenData field mapping (stock/status/arrival/returnable/available)."""
        from order.models import BooxenData

        sku = "9791100000015"
        file_bytes = _make_new_template_excel([
            {
                "sku": sku, "selected": "", "status": "",
                "bs_stock": 12, "bs_status": "정상",
                "bs_arrival": "2026-08-01", "bs_returnable": "가능",
            },
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        bd = BooxenData.objects.get(sku=sku)
        assert bd.stock == 12
        assert bd.status == "정상"
        assert bd.arrival == "2026-08-01"
        assert bd.returnable is True
        assert bd.available is True

    def test_booxen_returnable_false_when_not_available(self, auth_client):
        from order.models import BooxenData

        sku = "9791100000016"
        file_bytes = _make_new_template_excel([
            {
                "sku": sku, "selected": "", "status": "",
                "bs_stock": 0, "bs_returnable": "불가",
            },
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        bd = BooxenData.objects.get(sku=sku)
        assert bd.returnable is False
        assert bd.available is False


@pytest.mark.django_db
class TestUploadVendorUpsertIndependentOfLineItemExistence:
    def test_vendor_upsert_runs_even_without_unordered_line_item(self, auth_client):
        """REQ-PO8-014: no matching unordered LineItem exists at all — vendor
        upsert still runs (independent of the confirmation logic)."""
        from order.models import BooxenData

        sku = "9791100000017"
        assert not BooxenData.objects.filter(sku=sku).exists()

        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "북센", "status": "", "bs_price": 7000},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        assert res.data["skipped_count"] == 1  # no unordered LineItem -> PO logic skipped

        assert BooxenData.objects.get(sku=sku).price == Decimal("7000")


# ---------------------------------------------------------------------------
# Bug fixes (evaluator-active FAIL verdict review, follow-up cycle):
#   Bug 1 [CRITICAL]  Legacy-format upload must not wipe existing vendor data
#   Bug 2 [HIGH]      OverflowError on crafted numeric input (e.g. "inf")
#   Bug 3 [HIGH]      bs_price parsing regression for legacy-format files
#   Bug 4 [MEDIUM]    KyoboData.returnable coverage gap (no code bug)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
class TestUploadLegacyFormatPreservesVendorData:
    """Bug 1: legacy-format ('ISBN'+'선택' header) uploads must not overwrite
    existing BooxenData/KyoboData with None/False just because the legacy
    format's columns don't exist under the new-template's column names."""

    def test_legacy_upload_does_not_wipe_existing_booxen_and_kyobo_data(self, auth_client):
        """The legacy header has no 'BOOXEN 재고수량/재고상태/입고예정/반품' columns
        at all (it uses the differently-named '북센 ...' columns, which the
        parser does not look up), so every BooxenData field must survive
        untouched. Likewise '교보 상품상태' and '교보 출판사' (-> ky_status /
        ky_publisher) do not exist in the legacy header at all, so those two
        KyoboData fields must also survive untouched. ('교보 재고수량' /
        '교보 재고상태' DO share their column name with the new template, so a
        blank cell there legitimately overwrites — that is intentional
        column-exists-but-blank behavior, not part of this bug.)"""
        from order.models import BooxenData, KyoboData

        sku = "9791100000018"
        BooxenData.objects.create(
            sku=sku,
            stock=50,
            status="정상",
            arrival="2026-08-01",
            returnable=True,
            available=True,
            price=Decimal("5000"),
        )
        KyoboData.objects.create(
            sku=sku,
            status="정상",
            publisher="테스트출판사",
            price=Decimal("6000"),
        )

        # Blank '선택' still triggers the Part B vendor upsert (independent of
        # PO/CS confirmation) — see REQ-PO8-011/014.
        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": ""},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        bd = BooxenData.objects.get(sku=sku)
        assert bd.stock == 50
        assert bd.status == "정상"
        assert bd.arrival == "2026-08-01"
        assert bd.returnable is True
        assert bd.available is True

        kd = KyoboData.objects.get(sku=sku)
        assert kd.status == "정상"
        assert kd.publisher == "테스트출판사"


class TestIntOrNoneOverflow:
    """Bug 2: `_int_or_none` must gracefully degrade to None instead of
    raising OverflowError on values like 'inf' that `float()` parses fine
    but `int()` cannot convert."""

    def test_int_or_none_returns_none_for_infinity_string(self):
        from order.excel_utils import _int_or_none

        assert _int_or_none("inf") is None

    def test_int_or_none_returns_none_for_negative_infinity_string(self):
        from order.excel_utils import _int_or_none

        assert _int_or_none("-inf") is None


@pytest.mark.django_db
class TestUploadOverflowInputDoesNotCrash:
    """Bug 2: an 'inf' cell in a stock column must not produce an unhandled
    500 — the row should be processed with that field as None."""

    def test_infinite_stock_value_does_not_500_and_field_becomes_none(self, auth_client):
        from order.models import BooxenData

        sku = "9791100000019"
        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "", "status": "", "bs_stock": "inf"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        bd = BooxenData.objects.get(sku=sku)
        assert bd.stock is None


class TestParseLegacyBsPriceAlias:
    """Bug 3: legacy-format files use the Korean header '북센 공급가' for the
    Booxen supply price column, not 'BOOXEN 공급가' (new-template only)."""

    def test_legacy_bs_price_column_parsed(self):
        from order.excel_utils import parse_daily_review_excel

        file_bytes = _make_daily_review_excel([
            {"isbn": "9791100000020", "selected": "북센", "bs_price": 7500},
        ])
        results = parse_daily_review_excel(file_bytes)
        assert results[0]["bs_price"] == 7500.0


@pytest.mark.django_db
class TestUploadLegacyBsPriceFlowsToPurchaseOrder:
    """Bug 3: a real legacy '북센 공급가' value must flow through to
    PurchaseOrder.unit_price when 선택='북센' (not silently dropped to None)."""

    def test_legacy_bs_price_used_as_unit_price(self, auth_client):
        sku = "9791100000021"
        order = _make_order(shopify_order_id=90099)
        _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=1)

        file_bytes = _make_daily_review_excel([
            {"isbn": sku, "selected": "북센", "bs_price": 7500},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        po = PurchaseOrder.objects.get(sku=sku)
        assert po.unit_price == Decimal("7500")


@pytest.mark.django_db
class TestUploadKyoboReturnableMapping:
    """Bug 4 (coverage gap, not a code bug): KyoboData.returnable must be
    correctly mapped from 교보 반품가능여부's Y/N values in a new-template
    upload."""

    def test_kyobo_returnable_true_for_y(self, auth_client):
        from order.models import KyoboData

        sku = "9791100000022"
        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "", "status": "", "ky_returnable": "Y"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        assert KyoboData.objects.get(sku=sku).returnable is True

    def test_kyobo_returnable_false_for_n(self, auth_client):
        from order.models import KyoboData

        sku = "9791100000023"
        file_bytes = _make_new_template_excel([
            {"sku": sku, "selected": "", "status": "", "ky_returnable": "N"},
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201

        assert KyoboData.objects.get(sku=sku).returnable is False


# ===========================================================================
# SPEC-PURCHASE-ORDER-009
#
# Covers:
#   REQ-PO9-010  Query-count ceiling regression test (300~500 row fixture)
#   AC-001       Query count stays under a fixed ceiling, independent of
#                uploaded SKU row count
# ===========================================================================

_ACTIONABLE_CS_SELECTS = ["주문취소", "주문보류", "CS필요"]
_ACTIONABLE_WAREHOUSE_STATUSES = ["한국재고", "Fullerton재고", "NJ재고"]
_ACTIONABLE_WAREHOUSE_LOCATIONS = ["korea", "ca", "nj"]
_PO_BRANCH_SELECTS = ["BOOXEN", "교보", "YES24"]
_FILLER_SELECTS = ["", "합계"]
_N_FILLER = 6


def _make_bulk_daily_review_fixture(
    total_rows: int, sku_seed: str, order_name: str = "#8001"
) -> tuple[bytes, list[str]]:
    """
    Build a synthetic "Daily Order Review Template" upload with `total_rows`
    rows, for the REQ-PO9-010 query-count ceiling test.

    Composition (SPEC-PURCHASE-ORDER-009 fix cycle 2): a fixed, constant
    3 CS rows / 3 warehouse rows are routed to their respective branches,
    a fixed small ``_N_FILLER`` (6) rows have no matching LineItem and only
    exercise the vendor-table-sync + immediate-skip path, and the
    **remainder (`total_rows - 9 - _N_FILLER`) all route through the
    non-warehouse PurchaseOrder-creation branch** (booxen/kyobo/yes24,
    cycled).

    This mirrors the real production file's measured composition — 439 of
    447 rows (~98%) went through the non-warehouse PO-creation branch — so
    unlike the original fixture (which hardcoded the PO-branch count at a
    fixed 3 regardless of `total_rows`, leaving REQ-PO9-007's PO batching
    completely unmeasured by this test), the dominant branch's row count
    here actually scales with `total_rows`. `sku_seed` keeps SKUs (and
    therefore WarehouseStock rows, which are unique per isbn+location)
    distinct across multiple calls within the same test.

    Returns (file_bytes, actionable_skus) where actionable_skus is ordered
    [cs_sku_0..2, warehouse_sku_0..2, po_sku_0..N] (po_sku count scales with
    total_rows).
    """
    assert total_rows > 9 + _N_FILLER

    actionable_skus: list[str] = []
    rows: list[dict] = []

    for i, selected in enumerate(_ACTIONABLE_CS_SELECTS):
        sku = f"PERF{sku_seed}CS{i}"
        actionable_skus.append(sku)
        rows.append({"sku": sku, "selected": selected, "status": "", "bs_price": 1000})

    for i, wh_status in enumerate(_ACTIONABLE_WAREHOUSE_STATUSES):
        sku = f"PERF{sku_seed}WH{i}"
        actionable_skus.append(sku)
        rows.append({"sku": sku, "selected": "재고", "status": wh_status})

    n_po = total_rows - len(actionable_skus) - _N_FILLER
    for i in range(n_po):
        selected = _PO_BRANCH_SELECTS[i % len(_PO_BRANCH_SELECTS)]
        sku = f"PERF{sku_seed}PO{i}"
        actionable_skus.append(sku)
        row = {"sku": sku, "selected": selected, "status": ""}
        # Vary field presence across rows (REQ-PO9-002): at scale, this must
        # stay a small, fixed number of bulk_create() groups, never one
        # group per row.
        if i % 3 == 0:
            row["bs_price"] = 1200 + i
        if i % 4 == 0:
            row["ky_price"] = 3400 + i
        if i % 5 == 0:
            row["bs_stock"] = 10
        if selected == "BOOXEN" and "bs_price" not in row:
            row["bs_price"] = 5000
        rows.append(row)

    for i in range(_N_FILLER):
        selected = _FILLER_SELECTS[i % len(_FILLER_SELECTS)]
        rows.append({"sku": f"PERF{sku_seed}F{i}", "selected": selected, "status": ""})

    # REQ-PO8-018: every row must name the Order it belongs to — applied
    # uniformly here since this fixture only ever seeds a single order (see
    # _seed_actionable_line_items below).
    for row in rows:
        row["order_name"] = order_name

    file_bytes = _make_new_template_excel(rows)
    return file_bytes, actionable_skus


@pytest.mark.django_db
class TestUploadDailyReviewQueryCountCeiling:
    """REQ-PO9-010/AC-001: total Django query count during
    UploadDailyReviewView.post() stays under a fixed, small ceiling and does
    not scale with the number of uploaded SKU rows (SPEC-PURCHASE-ORDER-009)."""

    def _seed_actionable_line_items(self, actionable_skus: list[str], order_id: int) -> None:
        order = _make_order(shopify_order_id=order_id, name=f"#{order_id}")
        for idx, sku in enumerate(actionable_skus):
            _make_line_item(order, sku=sku, quantity=2, shopify_line_item_id=idx + 1)
        # actionable_skus layout: [cs x3, warehouse x3, po x(total_rows-9-_N_FILLER)]
        # — see _make_bulk_daily_review_fixture.
        warehouse_skus = actionable_skus[3:6]
        for sku, location in zip(warehouse_skus, _ACTIONABLE_WAREHOUSE_LOCATIONS):
            WarehouseStock.objects.create(isbn=sku, location=location, quantity=100)

    def _upload_and_capture(self, auth_client, total_rows: int, sku_seed: str, order_id: int):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        file_bytes, actionable_skus = _make_bulk_daily_review_fixture(
            total_rows, sku_seed, order_name=f"#{order_id}"
        )
        self._seed_actionable_line_items(actionable_skus, order_id)

        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        with CaptureQueriesContext(connection) as ctx:
            res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        # confirmed_count == every actionable row (CS + warehouse + PO);
        # the fixed _N_FILLER rows have no matching LineItem and are skipped.
        assert res.data["confirmed_count"] == len(actionable_skus)
        return res, len(ctx.captured_queries)

    def test_query_count_stays_under_ceiling_at_300_rows(self, auth_client):
        _, query_count = self._upload_and_capture(
            auth_client, 300, sku_seed="A300", order_id=97001
        )
        # Ceiling raised from the original 30 (SPEC-PURCHASE-ORDER-009 fix
        # cycle 1) to account for the now-batched PurchaseOrder creation
        # (REQ-PO9-007 fix cycle 2): one bulk_create() for the POs, one
        # SELECT to re-resolve their PKs (bulk_create() does not populate
        # `.pk` on this project's MySQL backend — see the REQ-PO9-007
        # investigation note in purchase_order_views.py), and one
        # bulk_create() for the M2M through-table rows. These are 3 fixed
        # additional queries, not per-row, so the ceiling is still a small
        # constant — just a slightly higher one than before this branch was
        # batched.
        assert query_count < 35, f"expected under 35 queries, got {query_count}"

    def test_query_count_does_not_scale_with_row_count(self, auth_client):
        """AC-001: 300 vs 500 rows must not produce proportionally more
        queries — only a small constant delta (incidental variation), never
        linear growth with SKU count. With this fixture, the PO-creation
        branch (previously fixed at 3 rows regardless of total_rows) now
        scales with total_rows, so this assertion is meaningful against the
        REQ-PO9-007 PurchaseOrder batching added in fix cycle 2."""
        _, count_300 = self._upload_and_capture(
            auth_client, 300, sku_seed="B300", order_id=97101
        )
        _, count_500 = self._upload_and_capture(
            auth_client, 500, sku_seed="B500", order_id=97201
        )
        assert count_500 < 35, f"expected under 35 queries at 500 rows, got {count_500}"
        assert count_500 - count_300 <= 3, (
            "query count scaled with row count instead of staying constant: "
            f"300 rows -> {count_300} queries, 500 rows -> {count_500} queries"
        )


@pytest.mark.django_db
class TestUploadPurchaseOrderLineItemsM2MCorrectness:
    """AC-008: batched PurchaseOrder creation (REQ-PO9-007) must link each PO
    to exactly the LineItems that belong to its own SKU — never mixed up
    across SKUs by the created_at-window re-query correlation strategy."""

    def test_batched_po_creation_links_correct_line_items_per_sku(self, auth_client):
        order = _make_order(shopify_order_id=98001, name="#98001")

        # Varying LineItem counts per SKU (1, 2, or 3) across all three
        # non-warehouse distributors, so a cross-SKU mix-up in the
        # created_at/highest-pk correlation would be detectable via a
        # wrong line_items count or membership for at least one SKU.
        distributors = ["BOOXEN", "교보", "YES24"]
        skus_and_counts: list[tuple[str, int]] = []
        rows = []
        li_counter = 0
        for i in range(120):
            sku = f"9791{i:09d}"
            li_count = (i % 3) + 1  # 1, 2, or 3 LineItems
            skus_and_counts.append((sku, li_count))
            for _ in range(li_count):
                li_counter += 1
                _make_line_item(order, sku=sku, quantity=1, shopify_line_item_id=li_counter)
            distributor = distributors[i % 3]
            row = {"sku": sku, "selected": distributor, "status": "정상", "order_name": "#98001"}
            if distributor == "BOOXEN":
                row["bs_price"] = 9000
            elif distributor == "교보":
                row["ky_price"] = 9500
            else:
                row["yes24_price"] = 9800
            rows.append(row)

        file_bytes = _make_new_template_excel(rows)
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        assert res.data["confirmed_count"] == 120

        for sku, expected_count in skus_and_counts:
            po = PurchaseOrder.objects.get(sku=sku)
            linked_skus = {li.sku for li in po.line_items.all()}
            assert linked_skus == {sku}, (
                f"PurchaseOrder for {sku} linked to wrong SKU(s): {linked_skus}"
            )
            assert po.line_items.count() == expected_count, (
                f"PurchaseOrder for {sku} expected {expected_count} line items, "
                f"got {po.line_items.count()}"
            )


# ===========================================================================
# REQ-PO8-018 (bugfix)
#
# Real production data (Daily Order Review Template 20260810.xlsx, SKU
# 9788998441012): order #37893 selected 재고/Fullerton재고 (warehouse) and
# order #37918 selected BOOXEN, for the SAME SKU. Because Part A previously
# collapsed rows by SKU alone, the later row (#37918/BOOXEN) silently won,
# and BOTH orders' LineItems ended up linked to the SAME PurchaseOrder.
#
# Covers:
#   AC-018-1  two orders sharing one SKU with DIFFERENT '선택' decisions no
#             longer collapse into one — each order's own decision applies
#             only to its own LineItem.
#   AC-018-2  two orders sharing one SKU with the SAME distributor still
#             batch into one PurchaseOrder with combined quantity (the
#             cross-order batching behavior must not regress).
# ===========================================================================


@pytest.mark.django_db
class TestUploadDailyReviewCrossOrderSameSku:
    """AC-018-1/AC-018-2: Part A resolves independently per (order_name,
    sku) — different decisions for the same SKU across different orders no
    longer collide, while same-distributor batching across orders is
    preserved."""

    def test_different_decisions_for_same_sku_across_orders_do_not_collide(self, auth_client):
        """AC-018-1: reproduces the real-world bug. order_a picks a
        warehouse code, order_b picks a real distributor, for the SAME SKU
        in the SAME upload — each LineItem must land in its own, independent
        outcome."""
        sku = "9788998441012"
        order_a = _make_order(shopify_order_id=37893, name="#A1")
        order_b = _make_order(shopify_order_id=37918, name="#A2")
        li_a = _make_line_item(order_a, sku=sku, quantity=1, shopify_line_item_id=1)
        li_b = _make_line_item(order_b, sku=sku, quantity=2, shopify_line_item_id=2)

        file_bytes = _make_new_template_excel([
            {
                "sku": sku, "selected": "재고", "status": "Fullerton재고",
                "order_name": "#A1",
            },
            {
                "sku": sku, "selected": "BOOXEN", "status": "",
                "order_name": "#A2", "bs_price": 9000,
            },
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        assert res.data["confirmed_count"] == 2

        li_a.refresh_from_db()
        li_b.refresh_from_db()

        # order_a: warehouse branch — in_stock, NOT linked to any PurchaseOrder.
        assert li_a.purchase_status == "in_stock"
        assert li_a.confirmed_distributor == "warehouse_ca"
        assert li_a.purchase_orders.count() == 0

        # order_b: non-warehouse branch — linked to exactly one PurchaseOrder,
        # and that PO must NOT include order_a's LineItem.
        assert li_b.purchase_orders.count() == 1
        po = li_b.purchase_orders.get()
        assert po.distributor == "booxen"
        linked_lis = list(po.line_items.all())
        assert linked_lis == [li_b]
        assert li_a not in linked_lis

    def test_same_distributor_across_orders_still_batches_into_one_po(self, auth_client):
        """AC-018-2: two different orders selecting the SAME distributor for
        the SAME SKU must still merge into one PurchaseOrder with combined
        quantity — the pre-existing batching behavior is not a regression
        target of the (order_name, sku) fix."""
        sku = "9788998441099"
        order_a = _make_order(shopify_order_id=37920, name="#B1")
        order_b = _make_order(shopify_order_id=37921, name="#B2")
        li_a = _make_line_item(order_a, sku=sku, quantity=3, shopify_line_item_id=1)
        li_b = _make_line_item(order_b, sku=sku, quantity=5, shopify_line_item_id=2)

        file_bytes = _make_new_template_excel([
            {
                "sku": sku, "selected": "BOOXEN", "status": "",
                "order_name": "#B1", "bs_price": 9000,
            },
            {
                "sku": sku, "selected": "BOOXEN", "status": "",
                "order_name": "#B2", "bs_price": 9000,
            },
        ])
        file_obj = io.BytesIO(file_bytes)
        file_obj.name = "daily_review.xlsx"
        res = auth_client.post(UPLOAD_DAILY_URL, data={"file": file_obj}, format="multipart")
        assert res.status_code == 201
        assert res.data["confirmed_count"] == 2

        pos = list(PurchaseOrder.objects.filter(sku=sku))
        assert len(pos) == 1, f"expected exactly one PurchaseOrder for {sku}, got {len(pos)}"
        po = pos[0]
        assert po.distributor == "booxen"
        assert po.quantity == 8  # 3 (order_a) + 5 (order_b)

        linked_lis = set(po.line_items.all())
        assert linked_lis == {li_a, li_b}


class TestParseDailyReviewMissingNameColumnRejected:
    """AC-018-3: a file missing the order-name column (Name / 주문번호) is
    rejected the same way a file missing the SKU column is (REQ-PO8-018)."""

    def test_missing_name_column_raises_value_error(self):
        from order.excel_utils import parse_daily_review_excel

        wb = openpyxl.Workbook()
        ws = wb.active
        # New-template-style header WITHOUT a "Name" column.
        ws.append(["Lineitem sku", "선택"])
        ws.append(["9791124591055", "BOOXEN"])
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError) as exc_info:
            parse_daily_review_excel(buf.getvalue())
        assert "Name" in str(exc_info.value)
        assert "주문번호" in str(exc_info.value)
