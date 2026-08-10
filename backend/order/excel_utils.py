"""
Excel generation and parsing utilities for purchase order workflows.

# @MX:ANCHOR: [AUTO] Public API for Excel I/O; called by generate-order-file and upload-vendor-file views
# @MX:REASON: High fan-in (2+ view callers) and external data format boundary
"""

import io
from decimal import Decimal

import openpyxl
import xlrd
from openpyxl import Workbook

# OLE2 Compound Document magic bytes — identifies legacy .xls (BIFF) format
_XLS_MAGIC = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

# Booxen result file (.xls): row 0 = timestamp, row 1 = headers (encoding-unreliable).
# Column positions confirmed from sample file analysis.
_BOOXEN_COL_TITLE = 1      # 도서명
_BOOXEN_COL_PRICE = 6      # 출고가 (단가)
_BOOXEN_COL_STOCK = 7      # 재고량
_BOOXEN_COL_RETURNABLE = 10  # 반품 ('가능'/'불가')
_BOOXEN_COL_STATUS = 11    # 상태 ('정상', '품절' 등)
_BOOXEN_COL_ISBN = 14      # ISBN
_BOOXEN_COL_ARRIVAL = 15   # 입고예정

# Kyobo result file (.xlsx): row 0 = "엑셀주문" (skip), row 1 = headers, row 2+ = data.
_KYOBO_COL_ISBN = 1          # ISBN
_KYOBO_COL_STATUS = 4        # 상품상태 ('정상', '품절' 등)
_KYOBO_COL_AVAILABLE = 5     # 출고여부 ('Y'/'N')
_KYOBO_COL_RETURNABLE = 6    # 반품가능여부 ('Y'/'N')
_KYOBO_COL_PUBLISHER = 8     # 출판사
_KYOBO_COL_QTY = 11          # 주문수량 (출고가 계산에 사용)
_KYOBO_COL_TOTAL_PRICE = 14  # 출고가합 (÷ 주문수량 = 단가)
_KYOBO_COL_STOCK = 15        # 보유재고

# YES24 result file (.xlsx): row 0 = headers (no title row), row 1+ = data.
# Column positions confirmed from sample file analysis.
_YES24_COL_LIST_PRICE = 6    # 정가
_YES24_COL_ISBN = 8          # ISBN
_YES24_COL_STATUS = 11       # 유통상태
_YES24_COL_PRICE = 13        # 공급가


def generate_order_excel(skus_data: list[dict], distributor: str) -> bytes:
    """
    Generate an Excel (.xlsx) file for purchase order submission to a distributor.

    Args:
        skus_data: List of dicts with keys: sku (str), title (str), total_quantity (int).
        distributor: Distributor name (used only for context; not written to file).

    Returns:
        Raw bytes of the generated .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "발주서"

    if distributor == "booxen":
        ws.append(["ISBN", "주문수량", "도서명", "출판사", "저자", "정가"])
        for row in skus_data:
            ws.append([row["sku"], row["total_quantity"], "", "", "", ""])
    elif distributor == "kyobo":
        ws.append(["ISBN", "수량"])
        for row in skus_data:
            ws.append([row["sku"], row["total_quantity"]])
    elif distributor == "yes24":
        ws.append(["번호", "도서명", "ISBN", "출판사", "정가", "수량"])
        for row in skus_data:
            ws.append(["", "", row["sku"], "", "", row["total_quantity"]])
    else:
        ws.append(["ISBN", "도서명", "수량"])
        for row in skus_data:
            ws.append([row["sku"], row["title"], row["total_quantity"]])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_line_item_notes_excel(notes: list[dict]) -> bytes:
    """Generate an Excel (.xlsx) file for 타출판사 line item notes export."""
    wb = Workbook()
    ws = wb.active
    ws.title = "타출판사 메모"

    ws.append(["주문명", "SKU", "도서명", "메모 내용", "작성자", "작성일", "유통사"])
    for note in notes:
        ws.append([
            note.get("order_name") or "",
            note.get("line_item_sku") or "",
            note.get("line_item_title") or "",
            note.get("content") or "",
            note.get("author_username") or "",
            note.get("created_at") or "",
            note.get("confirmed_distributor") or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_vendor_excel(file_bytes: bytes, distributor: str) -> list[dict]:
    """
    Parse a vendor-supplied Excel file into comparison records.

    Supports .xls (xlrd, booxen) and .xlsx (openpyxl, kyobo), detected via magic bytes.

    Returns:
        List of dicts with keys:
          sku, available, price,
          stock, returnable, status,
          arrival (booxen only)

    Raises:
        ValueError: If the file is empty or unreadable.
    """
    if file_bytes[:8] == _XLS_MAGIC:
        return _parse_booxen_xls(file_bytes)
    if distributor == "kyobo":
        return _parse_kyobo_xlsx(file_bytes)
    if distributor == "yes24":
        return _parse_yes24_xlsx(file_bytes)
    return _parse_generic_xlsx(file_bytes)


def _parse_booxen_xls(file_bytes: bytes) -> list[dict]:
    """
    Parse a Booxen result .xls file.

    Structure: row 0 = timestamp/title, row 1 = headers (encoding unreliable),
    row 2+ = data. Column positions are fixed (confirmed from sample file).
    """
    try:
        wb = xlrd.open_workbook(file_contents=file_bytes)
    except Exception as exc:
        raise ValueError(f"Cannot read .xls file: {exc}") from exc

    ws = wb.sheet_by_index(0)
    if ws.nrows < 3:
        raise ValueError("Empty file")

    required_max_col = max(
        _BOOXEN_COL_ISBN,
        _BOOXEN_COL_PRICE,
        _BOOXEN_COL_STOCK,
        _BOOXEN_COL_RETURNABLE,
        _BOOXEN_COL_STATUS,
        _BOOXEN_COL_ARRIVAL,
    )

    results = []
    for row_num in range(2, ws.nrows):  # skip row 0 (timestamp) and row 1 (headers)
        row = ws.row_values(row_num)
        if len(row) <= required_max_col:
            continue

        raw_sku = row[_BOOXEN_COL_ISBN]
        if raw_sku is None or raw_sku == "":
            continue
        # xlrd reads numeric ISBN cells as float
        sku = str(int(raw_sku)) if isinstance(raw_sku, float) else str(raw_sku).strip()
        if not sku:
            continue

        raw_stock = row[_BOOXEN_COL_STOCK]
        stock: int | None = None
        try:
            stock = int(float(raw_stock)) if raw_stock not in (None, "") else None
        except (TypeError, ValueError):
            stock = None
        available = (stock is not None and stock > 0)

        raw_price = row[_BOOXEN_COL_PRICE]
        price: float | None = None
        try:
            price = float(raw_price) if raw_price not in (None, "") else None
        except (TypeError, ValueError):
            price = None

        # 반품: '가능' → True, anything else → False
        raw_returnable = row[_BOOXEN_COL_RETURNABLE]
        returnable = str(raw_returnable).strip() == "가능" if raw_returnable not in (None, "") else None

        raw_status = row[_BOOXEN_COL_STATUS]
        status = str(raw_status).strip() if raw_status not in (None, "") else None

        raw_arrival = row[_BOOXEN_COL_ARRIVAL]
        arrival = str(raw_arrival).strip() if raw_arrival not in (None, "") else None

        results.append({
            "sku": sku,
            "available": available,
            "price": price,
            "stock": stock,
            "returnable": returnable,
            "status": status,
            "publisher": None,
            "ordered_qty": None,
            "total_price": None,
            "arrival": arrival,
        })

    return results


def _parse_kyobo_xlsx(file_bytes: bytes) -> list[dict]:
    """
    Parse a Kyobo result .xlsx file.

    Structure: row 0 = "엑셀주문" (title, skip), row 1 = headers, row 2+ = data.
    Column positions are fixed (confirmed from sample file analysis).
    출고가 = 출고가합 (col 14) ÷ 주문수량 (col 11)
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read .xlsx file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        raise ValueError("Empty file")

    required_max_col = max(
        _KYOBO_COL_ISBN,
        _KYOBO_COL_STATUS,
        _KYOBO_COL_AVAILABLE,
        _KYOBO_COL_RETURNABLE,
        _KYOBO_COL_PUBLISHER,
        _KYOBO_COL_QTY,
        _KYOBO_COL_TOTAL_PRICE,
        _KYOBO_COL_STOCK,
    )

    results = []
    for row in rows[2:]:  # skip row 0 (title) and row 1 (headers)
        if len(row) <= required_max_col:
            continue

        raw_sku = row[_KYOBO_COL_ISBN]
        sku = str(raw_sku).strip() if raw_sku is not None else None
        if not sku or not sku.isdigit():
            continue

        # 출고여부 'Y' → available
        raw_available = row[_KYOBO_COL_AVAILABLE]
        available = str(raw_available).strip().upper() == "Y" if raw_available is not None else False

        # 출고가합 저장 + 출고가(단가) = 출고가합 / 주문수량
        raw_total = row[_KYOBO_COL_TOTAL_PRICE]
        raw_qty = row[_KYOBO_COL_QTY]
        total_price: float | None = None
        ordered_qty: int | None = None
        price: float | None = None
        try:
            total_price = float(raw_total) if raw_total not in (None, "") else None
            qty_f = float(raw_qty) if raw_qty not in (None, "") else None
            if qty_f is not None:
                ordered_qty = int(qty_f)
            if total_price is not None and qty_f and qty_f > 0:
                price = total_price / qty_f
        except (TypeError, ValueError):
            pass

        raw_stock = row[_KYOBO_COL_STOCK]
        stock: int | None = None
        try:
            stock = int(float(raw_stock)) if raw_stock not in (None, "") else None
        except (TypeError, ValueError):
            stock = None

        raw_returnable = row[_KYOBO_COL_RETURNABLE]
        returnable = str(raw_returnable).strip().upper() == "Y" if raw_returnable is not None else None

        raw_status = row[_KYOBO_COL_STATUS]
        status = str(raw_status).strip() if raw_status not in (None, "") else None

        raw_publisher = row[_KYOBO_COL_PUBLISHER]
        publisher = str(raw_publisher).strip() if raw_publisher not in (None, "") else None

        results.append({
            "sku": sku,
            "available": available,
            "price": price,
            "stock": stock,
            "returnable": returnable,
            "status": status,
            "publisher": publisher,
            "ordered_qty": ordered_qty,
            "total_price": total_price,
            "arrival": None,
        })

    return results


def _parse_yes24_xlsx(file_bytes: bytes) -> list[dict]:
    """
    Parse a YES24 result .xlsx file.

    Structure: row 0 = headers (no title row, unlike booxen/kyobo), row 1+ = data.
    Column positions are fixed (confirmed from sample file analysis).
    YES24 has no availability/stock/returnable data, so `available` is always None.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read .xlsx file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Empty file")

    required_max_col = max(
        _YES24_COL_ISBN,
        _YES24_COL_LIST_PRICE,
        _YES24_COL_PRICE,
        _YES24_COL_STATUS,
    )

    results = []
    for row in rows[1:]:  # skip row 0 (headers) only
        if len(row) <= required_max_col:
            continue

        raw_sku = row[_YES24_COL_ISBN]
        sku = str(raw_sku).strip() if raw_sku is not None else None
        if not sku or not sku.isdigit():
            continue

        raw_list_price = row[_YES24_COL_LIST_PRICE]
        list_price: float | None = None
        try:
            list_price = float(raw_list_price) if raw_list_price not in (None, "") else None
        except (TypeError, ValueError):
            list_price = None

        raw_price = row[_YES24_COL_PRICE]
        price: float | None = None
        try:
            price = float(raw_price) if raw_price not in (None, "") else None
        except (TypeError, ValueError):
            price = None

        raw_status = row[_YES24_COL_STATUS]
        status = str(raw_status).strip() if raw_status not in (None, "") else None

        results.append({
            "sku": sku,
            "available": None,
            "price": price,
            "list_price": list_price,
            "status": status,
        })

    return results


def _parse_generic_xlsx(file_bytes: bytes) -> list[dict]:
    """Fallback parser for unknown .xlsx distributors. Uses header-name detection."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read .xlsx file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty file")

    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    isbn_idx = next((i for i, h in enumerate(header) if "isbn" in h or "sku" in h), 0)
    avail_idx = next((i for i, h in enumerate(header) if "재고" in h or "available" in h), 1)
    price_idx = next((i for i, h in enumerate(header) if "단가" in h or "price" in h), 2)

    results = []
    for row in rows[1:]:
        max_idx = max(isbn_idx, avail_idx, price_idx)
        if len(row) <= max_idx:
            continue

        raw_sku = row[isbn_idx]
        sku = str(raw_sku).strip() if raw_sku is not None else None
        if not sku:
            continue

        raw_avail = row[avail_idx]
        available = bool(raw_avail) if raw_avail is not None else False

        raw_price = row[price_idx]
        price: float | None = None
        try:
            price = float(raw_price) if raw_price is not None else None
        except (TypeError, ValueError):
            price = None

        results.append({
            "sku": sku, "available": available, "price": price,
            "stock": None, "returnable": None, "status": None,
            "publisher": None, "ordered_qty": None, "total_price": None, "arrival": None,
        })

    return results


def _result(
    selected: str | None,
    basis: str,
    price_diff: Decimal | None,
    price_diff_alert: bool,
) -> dict:
    """Return a standardized auto_select_distributor result dict."""
    return {
        "selected_distributor": selected,
        "candidate_basis": basis,
        "price_diff": price_diff,
        "price_diff_alert": price_diff_alert,
    }


def _compare_both(
    bs_price: Decimal | None,
    ky_price: Decimal | None,
    bs_ret: bool | None,
    ky_ret: bool | None,
) -> tuple[str, str]:
    """
    Select distributor when both vendors have sufficient stock (Step 2-A).

    Returns (selected_distributor, candidate_basis).
    """
    if bs_price is not None and ky_price is not None:
        if bs_price < ky_price:
            return "booxen", "양사재고/북센저가"
        if ky_price < bs_price:
            return "kyobo", "양사재고/교보저가"
        # Same price → check returnability
        if bs_ret is True and ky_ret is not True:
            return "booxen", "양사재고/동가/북센반품"
        if ky_ret is True and bs_ret is not True:
            return "kyobo", "양사재고/동가/교보반품"
        return "booxen", "양사재고/동가/반품동일"

    if bs_price is None and ky_price is not None:
        return "kyobo", "양사재고/교보가격만확인"
    if ky_price is None and bs_price is not None:
        return "booxen", "양사재고/북센가격만확인"
    return "booxen", "양사재고/가격없음"


def _no_stock_logic(
    bs_status: str,
    ky_status: str,
    bs_price: Decimal | None,
    ky_price: Decimal | None,
    bs_ret: bool | None,
    ky_ret: bool | None,
) -> tuple[str, str]:
    """
    Select distributor when neither vendor has sufficient stock (Step 2-D then 2-E).

    Returns (selected_distributor, candidate_basis).
    """
    basis = "양사재고없음"
    bs_cheaper = (
        bs_price is not None
        and ky_price is not None
        and bs_price <= ky_price
    )

    # Step 2-D: status/price heuristics
    if bs_status == "정상":
        if bs_cheaper:
            selected = "booxen"
        else:
            selected = "kyobo" if ky_status == "정상" else "check_required"
    elif ky_status in ("정상", "주문판매"):
        selected = "kyobo"
    else:
        selected = "check_required"

    # Step 2-E: kyobo-returnable override
    if ky_ret is True and bs_ret is not True:
        if ky_status == "정상":
            selected = "kyobo"
        else:
            selected = "check_required"

    return selected, basis


def auto_select_distributor(
    vc: "VendorComparison",
    total_qty: int,
    korea_stock: int = 0,
    ca_stock: int = 0,
    nj_stock: int = 0,
    vendor_rules: list[tuple[str, str]] | None = None,
) -> dict:
    """
    Determine the best distributor for a VendorComparison record.

    Implements a 5-step decision tree:
      Step 0: DistributorVendorRule override (agape / choeumgoyuk)
      Step 1: Warehouse stock priority (korea / west)
      Step 2: Vendor comparison (stock, price, returnability, status)
      Step 3: Price diff alert calculation

    Args:
        vc: VendorComparison instance (unsaved OK — only reads field values).
        total_qty: Total quantity needed for this SKU.
        korea_stock: Korea warehouse stock quantity.
        ca_stock: CA warehouse stock quantity.
        nj_stock: NJ warehouse stock quantity.
        vendor_rules: Pre-fetched list of (publisher_name, distributor_code) tuples.

    Returns:
        dict with keys:
            selected_distributor: str code or None
            candidate_basis: str label describing the selection reason
            price_diff: Decimal (booxen_price - kyobo_price) or None
            price_diff_alert: bool
    """
    # Lazy import to avoid circular dependency at module level
    # (VendorComparison is in models.py which imports nothing from excel_utils)
    bs_price: Decimal | None = vc.booxen_price
    ky_price: Decimal | None = vc.kyobo_price
    price_diff: Decimal | None = (
        (bs_price - ky_price)
        if (bs_price is not None and ky_price is not None)
        else None
    )

    # ------------------------------------------------------------------
    # Step 0: DistributorVendorRule override
    # ------------------------------------------------------------------
    if vendor_rules and vc.kyobo_publisher:
        for pub_name, dist_code in vendor_rules:
            if dist_code == "agape" and "아가페" in vc.kyobo_publisher:
                return _result("agape", "아가페규칙", price_diff, False)
            if dist_code == "choeumgoyuk" and vc.kyobo_publisher == pub_name:
                return _result("choeumgoyuk", "처음교육규칙", price_diff, False)
            if dist_code == "sungseoyunion" and vc.kyobo_publisher == pub_name:
                return _result("sungseoyunion", "성서유니온규칙", price_diff, False)

    # ------------------------------------------------------------------
    # Step 1: Warehouse stock priority
    # ------------------------------------------------------------------
    if korea_stock >= total_qty:
        return _result("warehouse", "재고우선", price_diff, False)
    if ca_stock >= total_qty or nj_stock >= total_qty:
        return _result("warehouse_west", "서부창고확인", price_diff, False)

    # ------------------------------------------------------------------
    # Step 2: Vendor comparison
    # ------------------------------------------------------------------
    bs_stock: int = vc.booxen_stock or 0
    ky_stock: int = vc.kyobo_stock or 0
    bs_ret: bool | None = vc.booxen_returnable
    ky_ret: bool | None = vc.kyobo_returnable
    bs_status: str = vc.booxen_status or ""
    ky_status: str = vc.kyobo_status or ""

    bs_enough = bs_stock >= total_qty
    ky_enough = ky_stock >= total_qty

    if bs_enough and ky_enough:
        selected, basis = _compare_both(bs_price, ky_price, bs_ret, ky_ret)
    elif bs_enough and not ky_enough:
        selected, basis = "booxen", "북센재고우선"
    elif ky_enough and not bs_enough:
        selected, basis = "kyobo", "교보재고우선"
    else:
        selected, basis = _no_stock_logic(
            bs_status, ky_status, bs_price, ky_price, bs_ret, ky_ret
        )

    # ------------------------------------------------------------------
    # Step 3: Price diff alert
    # ------------------------------------------------------------------
    price_diff_alert = False
    if price_diff is not None and abs(price_diff) >= 3000:
        if (
            selected == "check_required"
            or (selected == "booxen" and bs_price is not None and ky_price is not None and bs_price > ky_price)
            or (selected == "kyobo" and ky_price is not None and bs_price is not None and ky_price > bs_price)
        ):
            price_diff_alert = True

    return _result(selected, basis, price_diff, price_diff_alert)


# Distributor display name ↔ internal code mappings
_DISTRIBUTOR_LABEL_MAP: dict[str, str] = {
    "북센": "booxen",
    "교보": "kyobo",
    "처음교육": "choeumgoyuk",
    "아가페": "agape",
    "성서유니온": "sungseoyunion",
    "재고(한국)": "warehouse_korea",
    "재고(CA)": "warehouse_ca",
    "재고(NJ)": "warehouse_nj",
    # SPEC-PURCHASE-ORDER-008: new external template labels (AC-004, REQ-PO8-006/007)
    "BOOXEN": "booxen",  # new template legend uses English "BOOXEN" instead of "북센"
    "YES24": "yes24",
    "재고": "warehouse",  # generic warehouse code; location resolved from Status (REQ-PO8-008)
}

# Maps '선택' column CS-type values to purchase_status codes
_NOTE_TYPE_STATUS_MAP: dict[str, str] = {
    "주문취소": "order_cancelled",
    "주문보류": "on_hold",
    "CS필요": "cs_required",
    "타출판사": "other_publisher",
    # SPEC-PURCHASE-ORDER-010 REQ-DMG-003: re-enters the reorder queue via
    # ConfirmOrderView/UploadDailyReviewView's widened read-side eligibility.
    "파손/교환": "damaged_exchange",
}

_DAILY_REVIEW_HEADERS = [
    "주문번호", "ISBN", "제목", "수량", "주문위치", "메모",
    "재고(한국)", "재고(CA)", "재고(NJ)",
    "북센 공급가", "교보 공급가",
    "북센 재고수량", "북센 재고상태",
    "교보 재고수량", "교보 재고상태",
    "공급가차이",
    "북센 입고예정", "북센 반품가능여부",
    "교보 출고여부", "교보 반품가능여부",
    "가격차이알림",
    "출판사", "선택근거", "선택",
]


def generate_daily_review_excel(rows: list[dict]) -> bytes:
    """
    Generate Daily Order Review Excel with 24 columns.

    Each dict in `rows` should have:
        order_name, sku, title, quantity, location, note,
        korea_stock (int), ca_stock (int), nj_stock (int),
        bs_price (float|None), ky_price (float|None),
        bs_status (str|None), ky_stock (int|None), ky_status (str|None),
        price_diff (float|None), bs_arrival (str|None),
        bs_returnable (bool|None), ky_available (bool|None), ky_returnable (bool|None),
        price_diff_alert (bool), candidate_basis (str|None), selected (str)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Order Review"
    ws.append(_DAILY_REVIEW_HEADERS)

    for row in rows:
        bs_returnable = "Y" if row.get("bs_returnable") is True else ("N" if row.get("bs_returnable") is False else "")
        ky_returnable = "Y" if row.get("ky_returnable") is True else ("N" if row.get("ky_returnable") is False else "")
        ky_available = "Y" if row.get("ky_available") is True else ("N" if row.get("ky_available") is False else "")
        price_diff_alert = "Y" if row.get("price_diff_alert") else "N"

        ws.append([
            row.get("order_name") or "",
            row.get("sku") or "",
            row.get("title") or "",
            row.get("quantity") or 0,
            row.get("location") or "",
            row.get("other_publisher_memo") or row.get("note") or "",
            row.get("korea_stock") or 0,
            row.get("ca_stock") or 0,
            row.get("nj_stock") or 0,
            row.get("bs_price") if row.get("bs_price") is not None else "",
            row.get("ky_price") if row.get("ky_price") is not None else "",
            row.get("bs_stock") if row.get("bs_stock") is not None else "",
            row.get("bs_status") or "",
            row.get("ky_stock") if row.get("ky_stock") is not None else "",
            row.get("ky_status") or "",
            row.get("price_diff") if row.get("price_diff") is not None else "",
            row.get("bs_arrival") or "",
            bs_returnable,
            ky_available,
            ky_returnable,
            price_diff_alert,
            row.get("publisher") or "",
            row.get("candidate_basis") or "",
            row.get("selected") or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(row: tuple, idx: int | None):
    """Safely read a cell by index; returns None when idx is unset or out of range."""
    if idx is None or len(row) <= idx:
        return None
    return row[idx]


def _str_or_none(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _float_or_none(value) -> float | None:
    try:
        return float(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _int_or_none(value) -> int | None:
    f = _float_or_none(value)
    if f is None:
        return None
    try:
        return int(f)
    except (OverflowError, ValueError):
        # e.g. float("inf")/"nan" parse fine but cannot convert to int —
        # degrade gracefully to None instead of propagating an unhandled 500.
        return None


def parse_daily_review_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse an uploaded Daily Review Excel file.

    Supports two header layouts, auto-detected by scanning rows for the first
    one that contains a '선택' cell together with either a 'Lineitem sku' cell
    (external "Daily Order Review Template", header typically on row 3) or an
    'ISBN' cell (legacy self-generated format, header on row 1) — REQ-PO8-001.
    All column lookups are name-based (header.index()), never positional, so
    the headerless legend column that sits immediately right of '선택' in the
    new template is never read as data (REQ-PO8-003).

    The note/memo source column also depends on which SKU column matched: the
    new template has no '메모' column and uses 'Status' instead; the legacy
    format keeps using '메모' (REQ-PO8-005).

    Rows are returned for every non-empty SKU regardless of whether '선택' is
    populated or recognized (REQ-PO8-011, REQ-PO8-014) — Part B vendor-table
    sync consumes every such row independently of purchase-order confirmation.

    Returns:
        List of dicts:
            sku, distributor, note_type, note,
            bs_price, ky_price, yes24_price (always present, value None when
                the cell/column is unset — safe because callers only act on
                these when not None),
            bs_stock, bs_status, bs_arrival, bs_returnable, bs_available,
            yes24_status,
            ky_stock, ky_available, ky_status, ky_returnable, ky_publisher,
            ky_list_price
                — each of these keys is present ONLY when its source header
                column exists in this file (same convention as ky_list_price
                below), so a legacy-format upload (which lacks these columns
                under these names) never overwrites existing vendor-table
                values with None/False (Bug-1-fix, REQ-PO8-014 hardening).
        distributor is the internal code (booxen, kyobo, yes24, choeumgoyuk,
        agape, sungseoyunion, warehouse, warehouse_korea/ca/nj) or None when
        '선택' is empty or unrecognized.

    Raises:
        ValueError: If file is unreadable or no header row is found.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read Excel file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("파일이 비어 있습니다.")

    header: list[str] | None = None
    header_row_idx: int | None = None
    sku_header_name: str | None = None
    for idx, raw_row in enumerate(rows):
        candidate = [str(h).strip() if h is not None else "" for h in raw_row]
        if "선택" not in candidate:
            continue
        if "Lineitem sku" in candidate:
            sku_header_name = "Lineitem sku"
        elif "ISBN" in candidate:
            sku_header_name = "ISBN"
        else:
            continue
        header = candidate
        header_row_idx = idx
        break

    if header is None or header_row_idx is None or sku_header_name is None:
        raise ValueError(
            "올바른 Daily Review 형식의 파일이 아닙니다 (ISBN 또는 Lineitem sku, 선택 컬럼 필요)."
        )

    sku_idx = header.index(sku_header_name)
    selected_idx = header.index("선택")
    is_new_template = sku_header_name == "Lineitem sku"

    if is_new_template:
        note_idx = header.index("Status") if "Status" in header else None
    else:
        note_idx = header.index("메모") if "메모" in header else None

    # REQ-PO8-004/Bug-3-fix: new template uses "BOOXEN 공급가"; the legacy
    # self-generated format (_DAILY_REVIEW_HEADERS) uses "북센 공급가" for the
    # same column. Check the new-template name first, fall back to legacy —
    # same dual-alias pattern already used for the SKU column.
    if "BOOXEN 공급가" in header:
        bs_price_idx = header.index("BOOXEN 공급가")
    elif "북센 공급가" in header:
        bs_price_idx = header.index("북센 공급가")
    else:
        bs_price_idx = None
    ky_price_idx = header.index("교보 공급가") if "교보 공급가" in header else None
    yes24_price_idx = header.index("YES24 공급가") if "YES24 공급가" in header else None

    bs_stock_idx = header.index("BOOXEN 재고수량") if "BOOXEN 재고수량" in header else None
    bs_status_idx = header.index("BOOXEN 재고상태") if "BOOXEN 재고상태" in header else None
    bs_arrival_idx = header.index("BOOXEN 입고예정") if "BOOXEN 입고예정" in header else None
    bs_returnable_idx = header.index("BOOXEN 반품") if "BOOXEN 반품" in header else None

    yes24_status_idx = header.index("YES24 재고상태") if "YES24 재고상태" in header else None

    ky_stock_idx = header.index("교보 재고수량") if "교보 재고수량" in header else None
    ky_available_idx = header.index("교보 재고상태") if "교보 재고상태" in header else None
    ky_status_idx = header.index("교보 상품상태") if "교보 상품상태" in header else None
    ky_returnable_idx = header.index("교보 반품가능여부") if "교보 반품가능여부" in header else None
    ky_publisher_idx = header.index("교보 출판사") if "교보 출판사" in header else None
    ky_list_price_idx = header.index("교보 정가") if "교보 정가" in header else None

    results = []
    for row in rows[header_row_idx + 1:]:
        if len(row) <= max(sku_idx, selected_idx):
            continue

        raw_sku = row[sku_idx]
        sku = str(raw_sku).strip() if raw_sku is not None else ""
        if not sku:
            continue

        selected_label = _str_or_none(_cell(row, selected_idx)) or ""

        distributor_code = _DISTRIBUTOR_LABEL_MAP.get(selected_label) if selected_label else None
        note_type: str | None = None
        if selected_label and not distributor_code and selected_label in _NOTE_TYPE_STATUS_MAP:
            note_type = selected_label

        note = _str_or_none(_cell(row, note_idx))

        bs_price = _float_or_none(_cell(row, bs_price_idx))
        ky_price = _float_or_none(_cell(row, ky_price_idx))
        yes24_price = _float_or_none(_cell(row, yes24_price_idx))

        bs_stock = _int_or_none(_cell(row, bs_stock_idx))
        bs_status = _str_or_none(_cell(row, bs_status_idx))
        bs_arrival = _str_or_none(_cell(row, bs_arrival_idx))
        # REQ-PO8-015: '가능' -> True, any other value (including blank) -> False.
        bs_returnable = _str_or_none(_cell(row, bs_returnable_idx)) == "가능"
        bs_available = bs_stock is not None and bs_stock > 0

        yes24_status = _str_or_none(_cell(row, yes24_status_idx))

        ky_stock = _int_or_none(_cell(row, ky_stock_idx))
        raw_ky_available = _cell(row, ky_available_idx)
        ky_available = (
            str(raw_ky_available).strip().upper() == "Y" if raw_ky_available is not None else None
        )
        ky_status = _str_or_none(_cell(row, ky_status_idx))
        raw_ky_returnable = _cell(row, ky_returnable_idx)
        ky_returnable = (
            str(raw_ky_returnable).strip().upper() == "Y" if raw_ky_returnable is not None else None
        )
        ky_publisher = _str_or_none(_cell(row, ky_publisher_idx))

        result = {
            "sku": sku,
            "distributor": distributor_code,
            "note_type": note_type,
            "note": note,
            "bs_price": bs_price,
            "ky_price": ky_price,
            "yes24_price": yes24_price,
        }
        # Bug-1-fix (REQ-PO8-014 hardening): only include a Part B vendor
        # field in the result dict when its source column actually exists in
        # this file's header. Legacy-format files (ISBN+선택, row 1) never had
        # these BOOXEN/YES24/일부 교보 columns under these exact names, so the
        # key must be entirely absent rather than present-with-None — the
        # view relies on key absence to leave existing DB values untouched.
        # Same "only present if the column existed" convention already used
        # for ky_list_price above.
        if bs_stock_idx is not None:
            result["bs_stock"] = bs_stock
            result["bs_available"] = bs_available
        if bs_status_idx is not None:
            result["bs_status"] = bs_status
        if bs_arrival_idx is not None:
            result["bs_arrival"] = bs_arrival
        if bs_returnable_idx is not None:
            result["bs_returnable"] = bs_returnable
        if yes24_status_idx is not None:
            result["yes24_status"] = yes24_status
        if ky_stock_idx is not None:
            result["ky_stock"] = ky_stock
        if ky_available_idx is not None:
            result["ky_available"] = ky_available
        if ky_status_idx is not None:
            result["ky_status"] = ky_status
        if ky_returnable_idx is not None:
            result["ky_returnable"] = ky_returnable
        if ky_publisher_idx is not None:
            result["ky_publisher"] = ky_publisher
        # REQ-PO8-017: only include list_price when the '교보 정가' column exists,
        # so an upload without it never clears an existing KyoboData.list_price.
        if ky_list_price_idx is not None:
            result["ky_list_price"] = _float_or_none(_cell(row, ky_list_price_idx))

        results.append(result)

    return results


# ---------------------------------------------------------------------------
# SPEC-ORDER-011: logistics_status uploads (vendor-shipment-confirmation,
# warehouse-receiving-results) — PLACEHOLDER SKU-only schema.
#
# No real vendor/warehouse template exists yet for either upload (spec.md
# "확인이 필요한 가정"). Both files are parsed with a minimal SKU-only schema:
# a single required column whose header contains "sku" or "isbn"
# (case-insensitive, first match wins) — every other column is ignored.
# Replace this parser (and the column-matching rule) once the real templates
# are available; the calling views (UploadVendorShipmentView /
# UploadWarehouseReceiptView) only depend on each parsed row exposing a
# "sku" key, so swapping the parser body will not require view changes.
# ---------------------------------------------------------------------------


def _parse_sku_only_xlsx(file_bytes: bytes) -> list[dict]:
    """
    Shared placeholder parser for parse_vendor_shipment_excel() and
    parse_warehouse_receipt_excel(). Mirrors _parse_generic_xlsx's
    header-name detection, but the SKU/ISBN column is REQUIRED (raises
    ValueError when no header matches, unlike _parse_generic_xlsx's
    positional fallback) since this is the only column consumed.

    Blank-SKU rows are skipped uncounted (never added to the returned list)
    — the calling view's skipped_count only reflects SKUs that failed to
    match a LineItem, not blank rows in the source file.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read .xlsx file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty file")

    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    sku_idx = next((i for i, h in enumerate(header) if "sku" in h or "isbn" in h), None)
    if sku_idx is None:
        raise ValueError("No SKU/ISBN column found in header")

    results = []
    for row in rows[1:]:
        if len(row) <= sku_idx:
            continue

        raw_sku = row[sku_idx]
        sku = str(raw_sku).strip() if raw_sku is not None else ""
        if not sku:
            continue

        results.append({"sku": sku})

    return results


# @MX:NOTE: [AUTO] Column alias lists ("주문번호"/"order", "sku"/"isbn",
# "렉번호"/"rack") intentionally mirror the case-insensitive substring
# convention already used by _parse_sku_only_xlsx above, extended from 1 to
# 3 required columns (SPEC-ORDER-013 결정 D). The order-number alias also
# accepts an exact "name" header match (Shopify's bare "Name" column) and
# excludes any header containing "date" from the loose "order" substring
# match, so "Order Date" is never mistaken for the order-number column
# (bug fix, real-world "발송정리" export file).
def parse_rack_number_excel(file_bytes: bytes) -> list[dict]:
    """
    SPEC-ORDER-013 REQ-RACK-005/005a (design correction, same-day follow-up):
    parse an uploaded rack-number Excel file. Header row (row 0) is scanned
    for three required columns — order identifier, SKU, rack number — using
    case-insensitive substring matching against each column's expected name
    variants, independent of left-to-right position (REQ-RACK-005). The
    order-identifier column also matches on an exact "name" header (Shopify
    order-name column alias) and excludes any header containing "date" from
    the loose "order" substring match, so a leading "Order Date" column is
    never mistaken for the order-identifier column (bug fix). Raises
    ValueError when any of the three columns cannot be located
    (REQ-RACK-005a).

    Order-identifier cell values are read as a raw string and whitespace-
    stripped only — no "#" stripping and no integer conversion (design fix).
    `Order.name` (Shopify's order display name, e.g. "#37349") is what this
    value is matched against downstream, and it is stored WITH the leading
    "#". Order.name and Order.order_number usually correlate numerically for
    normal Shopify orders but diverge for manually-entered "EB"-prefixed
    orders (e.g. name="#EB10011778" with an unrelated order_number=1778),
    where there is no way to derive one from the other. Matching against
    Order.name as an exact string is therefore the only correct approach.

    Data rows: a row is skipped (not included in the result) only when its
    SKU is blank. A row whose order-identifier cell is itself blank/empty is
    still included, with `order_name` set to `None` — it is NOT dropped
    here, so that `UploadRackNumberView` can count it toward `skipped_count`
    (REQ-RACK-006a treats it the same as an "Order not found" skip) rather
    than it silently vanishing from both `matched_count` and
    `skipped_count` (AC-RACK-007). The rack_number value is NOT required —
    an empty string is preserved as an explicit "clear this rack_number"
    value rather than being skipped (plan.md M3).

    Returns:
        List of dicts: {"order_name": str | None, "sku": str, "rack_number": str}
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read .xlsx file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty file")

    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    order_idx = next(
        (
            i
            for i, h in enumerate(header)
            if "주문번호" in h or h == "name" or ("order" in h and "date" not in h)
        ),
        None,
    )
    sku_idx = next((i for i, h in enumerate(header) if "sku" in h or "isbn" in h), None)
    rack_idx = next((i for i, h in enumerate(header) if "렉번호" in h or "rack" in h), None)

    if order_idx is None or sku_idx is None or rack_idx is None:
        raise ValueError(
            "필수 컬럼(주문번호/SKU/렉번호)을 찾을 수 없습니다."
        )

    results = []
    for row in rows[1:]:
        if len(row) <= max(order_idx, sku_idx, rack_idx):
            continue

        # @MX:NOTE: [AUTO] order_name is intentionally NOT used to skip the
        # row here (design fix, REQ-RACK-006a/AC-RACK-007) — a blank
        # order-identifier cell yields order_name=None and the row still
        # flows through to `results` so UploadRackNumberView can count it
        # as skipped instead of dropping it invisibly. The value is read
        # verbatim (string identity, "#" preserved) since it is matched
        # against Order.name downstream, not parsed as a number — Order.name
        # and Order.order_number diverge for "EB"-prefixed manual orders.
        raw_order = _cell(row, order_idx)
        order_name = _str_or_none(raw_order)

        raw_sku = _cell(row, sku_idx)
        sku = str(raw_sku).strip() if raw_sku is not None else ""
        if not sku:
            continue

        raw_rack = _cell(row, rack_idx)
        rack_number = str(raw_rack).strip() if raw_rack is not None else ""

        results.append({"order_name": order_name, "sku": sku, "rack_number": rack_number})

    return results


def parse_outbound_excel(file_bytes: bytes) -> list[dict]:
    """
    SPEC-ORDER-015 REQ-OUTBOUND-012/012a: parse an uploaded outbound-processing
    Excel file. The header row (row 0) is scanned for three required columns —
    order name, SKU, quantity — using the same case-insensitive substring
    convention as `parse_rack_number_excel` above, so column order does not
    matter. Raises ValueError when any of the three cannot be located
    (REQ-OUTBOUND-012a); `UploadOutboundView` converts that into HTTP 422.

    Order-name cell values are read as a raw whitespace-stripped string with
    no "#" stripping and no int conversion — they are matched against
    `Order.name` ("#37349", "#EB10011778") downstream, never against
    `Order.order_number`.

    Data rows with a blank SKU are skipped, which also drops the trailing
    all-empty rows openpyxl commonly reports.

    Returns:
        List of dicts: {"name": str, "sku": str, "total": int} — the exact
        row shape `_process_outbound_rows` consumes from the manual-entry
        endpoint, so both endpoints share one contract.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read .xlsx file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty file")

    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    name_idx = next((i for i, h in enumerate(header) if "name" in h or "주문번호" in h), None)
    sku_idx = next((i for i, h in enumerate(header) if "sku" in h), None)
    total_idx = next((i for i, h in enumerate(header) if "total" in h or "수량" in h), None)

    if name_idx is None or sku_idx is None or total_idx is None:
        raise ValueError("필수 컬럼(주문번호/SKU/수량)을 찾을 수 없습니다.")

    results = []
    for row in rows[1:]:
        # No explicit short-row guard: openpyxl pads every row of
        # iter_rows(values_only=True) to the sheet's max_column, and `_cell`
        # returns None for an out-of-range index regardless, so a short row
        # simply yields a blank SKU and is dropped just below.
        raw_sku = _cell(row, sku_idx)
        sku = str(raw_sku).strip() if raw_sku is not None else ""
        if not sku:
            continue

        raw_name = _cell(row, name_idx)
        name = str(raw_name).strip() if raw_name is not None else ""

        raw_total = _cell(row, total_idx)
        try:
            total = int(raw_total)
        except (TypeError, ValueError):
            total = 0

        results.append({"name": name, "sku": sku, "total": total})

    return results


def parse_vendor_shipment_excel(file_bytes: bytes) -> list[dict]:
    """
    SPEC-ORDER-011 REQ-LOGI-003: parse an uploaded vendor-shipment-
    confirmation file. PLACEHOLDER schema — see _parse_sku_only_xlsx.
    """
    return _parse_sku_only_xlsx(file_bytes)


def parse_warehouse_receipt_excel(file_bytes: bytes) -> list[dict]:
    """
    SPEC-ORDER-011 REQ-LOGI-005: parse an uploaded warehouse-receiving-
    results file. PLACEHOLDER schema — see _parse_sku_only_xlsx.
    """
    return _parse_sku_only_xlsx(file_bytes)
